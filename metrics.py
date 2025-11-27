from database import get_db_connection, close_db_connection
from flask_jwt_extended import get_jwt_identity
from flask import current_app

def execute_metric_query(sql_query, client_id, metric_name="metric"):
    """
    Generic function to execute a parameterized metric query for a specific client.
    
    Args:
        sql_query (str): The SQL query to execute with %s placeholder for client_id
        client_id (int): The client ID to use in the query
        metric_name (str): Name of the metric for error reporting
        
    Returns:
        float: The calculated metric value or a special value indicating no data/error
    """
    try:
        client_id = int(client_id)
        if client_id <= 0:
            print(f"Warning: {metric_name} - client_id must be a positive integer, got {client_id}")
            return None
    except (ValueError, TypeError):
        print(f"Warning: {metric_name} - client_id must be a positive integer, got {client_id}")
        return None

    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Parameterize the query to prevent SQL injection
        # Count only %s parameter placeholders, not % in ILIKE patterns
        import re
        param_count = len(re.findall(r'(?<!%)%s(?!%)', sql_query))
        
        # Debug logging for parameter counting
        try:
            current_app.logger.info(f"DEBUG: {metric_name} query found {param_count} %s placeholders")
            # Show first 200 chars of query for debugging
            query_preview = sql_query[:200] + "..." if len(sql_query) > 200 else sql_query
            current_app.logger.info(f"DEBUG: Query preview: {query_preview}")
        except RuntimeError:
            print(f"DEBUG: {metric_name} query found {param_count} %s placeholders")
            query_preview = sql_query[:200] + "..." if len(sql_query) > 200 else sql_query
            print(f"DEBUG: Query preview: {query_preview}")
        
        if param_count > 0:
            params = tuple(client_id for _ in range(param_count))
        else:
            params = ()
        
        # Debug logging
        try:
            current_app.logger.info(f"DEBUG: Executing {metric_name} query for client {client_id} with {param_count} parameters: {params}")
        except RuntimeError:
            print(f"DEBUG: Executing {metric_name} query for client {client_id} with {param_count} parameters: {params}")
        
        cursor.execute(sql_query, params)
        
        result = cursor.fetchone()
        cursor.close()
        
        # Debug logging
        try:
            current_app.logger.info(f"DEBUG: {metric_name} query result: {result}")
        except RuntimeError:
            print(f"DEBUG: {metric_name} query result: {result}")
        
        # Handle different result scenarios
        if result is None:
            # No rows returned - this is normal for some ratio queries when there's no data
            print(f"Info: No data returned for {metric_name} for client_id {client_id}")
            return None
        
        # Check if result has at least one column
        if not result or len(result) == 0:
            print(f"Info: Empty result row for {metric_name} for client_id {client_id}")
            return None
        
        # Safely access the first value
        try:
            value = result[0]
        except IndexError:
            print(f"Info: No columns in result for {metric_name} for client_id {client_id}")
            return None
        
        if value is None:
            return None  # Null value in database
        
        # Try to convert to float, handle conversion errors
        try:
            return float(value)
        except (ValueError, TypeError):
            print(f"Warning: Could not convert {metric_name} value '{value}' to float for client_id {client_id}")
            return None
        
    except Exception as e:
        print(f"Error calculating {metric_name} for client_id {client_id}: {e}")
        # Return None instead of raising exception to allow graceful handling
        return None
    finally:
        if connection:
            close_db_connection(connection)

def calculate_net_worth_for_user():
    """
    Calculates net worth for the authenticated user using a parameterized query.
    The client_id is extracted from the JWT token of the authenticated user.
    
    Returns:
        float: The calculated net worth value
        
    Raises:
        ValueError: If client_id is invalid
        Exception: For database-related errors
    """
    # Get the client_id from the authenticated user's JWT token
    client_id = get_jwt_identity()
    
    sql_query = """
        WITH assets AS (
            SELECT COALESCE(SUM(value), 0) AS total_value FROM core.holdings WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(total_value), 0) FROM core.real_estate_assets WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(amount), 0) FROM core.businesses WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(total_value), 0) FROM core.investment_deposit_accounts WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(total_value), 0) FROM core.personal_property_accounts WHERE client_id = %s
        ),
        liabilities AS (
            SELECT COALESCE(SUM(ABS(total_value)), 0) AS total_liabilities
            FROM core.liability_note_accounts
            WHERE client_id = %s
        ),
        asset_summary AS (
            SELECT SUM(total_value) AS total_assets FROM assets
        )
        SELECT total_assets - total_liabilities AS net_worth FROM asset_summary, liabilities;
    """
    
    return execute_metric_query(sql_query, client_id, "net_worth")

def calculate_portfolio_value_for_user():
    """Calculate portfolio value for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH portfolio_components AS (
            -- Holdings
            SELECT COALESCE(SUM(value), 0) AS portfolio_value
            FROM core.holdings
            WHERE client_id = %s
            
            UNION ALL
            
            -- Investment deposit accounts
            SELECT COALESCE(SUM(total_value), 0)
            FROM core.investment_deposit_accounts
            WHERE client_id = %s
        )
        SELECT SUM(portfolio_value) AS total_portfolio_value
        FROM portfolio_components;
    """
    return execute_metric_query(sql_query, client_id, "portfolio_value")

def calculate_real_estate_value_for_user():
    """Calculate real estate value for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(total_value),0) as real_estate_value
        FROM core.real_estate_assets
        WHERE client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "real_estate_value")

def calculate_debt_for_user():
    """Calculate debt for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
        ABS(COALESCE(SUM(total_value),0))
        FROM core.liability_note_accounts
        WHERE client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "debt")

def calculate_equity_for_user():
    """Calculate equity for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH equity_holdings AS (
            SELECT
                client_id,
                COALESCE(SUM(value),0) as equity_holdings_value
            FROM core.holdings
            WHERE asset_class IN ('largecap', 'smallcap', 'largevalue', 'smallvalue', 'internat', 'emerging', 'ips') -- Filters for equity asset classes: largecap, smallcap, largevalue, smallvalue, internat, emerging, ips
                AND client_id = %s
                AND value IS NOT NULL
            GROUP BY client_id
        ),
        investment_equity AS (
            SELECT
                client_id,
                COALESCE(SUM(holdings_value),0) as investment_equity_value
            FROM core.investment_deposit_accounts
            WHERE fact_type_name IN ('Taxable Investment', 'Roth IRA', 'Qualified Retirement') -- Filters for investment account types: Taxable Investment, Roth IRA, Qualified Retirement
                AND client_id = %s
                AND holdings_value IS NOT NULL
            GROUP BY client_id
        )

        SELECT
            (COALESCE(eh.equity_holdings_value, 0) + COALESCE(ie.investment_equity_value, 0)) as total_equity
        FROM equity_holdings eh
        FULL OUTER JOIN investment_equity ie ON eh.client_id = ie.client_id;
    """
    return execute_metric_query(sql_query, client_id, "equity")

def calculate_fixed_income_for_user():
    """Calculate fixed income for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(value),0) as fixed_income_total
        FROM core.holdings
        WHERE asset_class IN ('highyldbond', 'inttermmun', 'investbond', 'shortermbond', 'shortermmun')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "fixed_income")

def calculate_cash_for_user():
    """Calculate cash for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH holdings_cash AS (
            SELECT
                client_id,
                SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) as cash_from_holdings
            FROM core.holdings
            WHERE asset_class = 'cash' AND value IS NOT NULL
            AND client_id = %s
            GROUP BY client_id
        ),
        investment_cash AS (
            SELECT
                client_id,
                SUM(COALESCE(cash_balance, 0)) as cash_from_investments
            FROM core.investment_deposit_accounts
            WHERE fact_type_name = 'Cash Alternative'
                AND cash_balance IS NOT NULL
                AND client_id = %s
            GROUP BY client_id
        )

        SELECT
            COALESCE(h.cash_from_holdings, 0) + COALESCE(i.cash_from_investments, 0) as total_cash
        FROM holdings_cash h
        FULL OUTER JOIN investment_cash i ON h.client_id = i.client_id;
    """
    return execute_metric_query(sql_query, client_id, "cash")

def calculate_earned_income_for_user():
    """Calculate earned income for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as earned_income
        FROM core.incomes
        WHERE income_type IN ('Salary')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "earned_income")

def calculate_social_security_income_for_user():
    """Calculate social security income for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as social_income
        FROM core.incomes
        WHERE income_type IN ('SocialSecurity')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "social_security_income")

def calculate_pension_income_for_user():
    """Calculate pension income for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as pension_income
        FROM core.incomes
        WHERE income_type IN ('Pension') -- No pension income_type currently in the data
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "pension_income")

def calculate_real_estate_income_for_user():
    """Calculate real estate income for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as real_estate_income
        FROM core.incomes
        WHERE income_type IN ('Real Estate') -- No Real Estate income_type currently in the data
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "real_estate_income")

def calculate_business_income_for_user():
    """Calculate business income for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as business_income
        FROM core.incomes
        WHERE income_type IN ('Business') -- No Business income_type currently in the data
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "business_income")

def calculate_total_income_for_user():
    """Calculate total income for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH income_breakdown AS (
            SELECT
                client_id,
                -- Earned Income
                COALESCE(SUM(CASE WHEN income_type = 'Salary' THEN current_year_amount ELSE 0 END), 0) as earned_income,
                -- Social Income
                COALESCE(SUM(CASE WHEN income_type = 'SocialSecurity' THEN current_year_amount ELSE 0 END), 0) as social_income,
                -- Pension Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Pension' THEN current_year_amount ELSE 0 END), 0) as pension_income,
                -- Real Estate Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Real Estate' THEN current_year_amount ELSE 0 END), 0) as real_estate_income,
                -- Business Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Business' THEN current_year_amount ELSE 0 END), 0) as business_income,
                -- Other Income (exists as 'Other Income' in database)
                COALESCE(SUM(CASE WHEN income_type = 'Other' THEN current_year_amount ELSE 0 END), 0) as other_income,
                -- Total Income = Sum of all components
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s  -- Filter for specific client
              AND current_year_amount IS NOT NULL
            GROUP BY client_id
        )
        SELECT
            total_income
        FROM income_breakdown;
    """
    return execute_metric_query(sql_query, client_id, "total_income")

def calculate_current_year_giving_for_user():
    """Calculate current year giving for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(annual_amount), 0) AS current_year_giving
        FROM core.expenses
        WHERE client_id = %s
            AND type = 'Spending'
            AND sub_type = 'GivingAndPhilanthropy'
            AND annual_amount > 0
            -- Check if expense overlaps with current year
            AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
            AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE));
    """
    return execute_metric_query(sql_query, client_id, "current_year_giving")

def calculate_current_year_savings_for_user():
    """Calculate current year savings for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(calculated_annual_amount_usd),0) as current_year_savings
        FROM core.savings
        WHERE start_type = 'Active'  -- Only include currently active savings plans
          AND client_id = %s;  -- Filter for specific client (replace with actual client ID)
    """
    return execute_metric_query(sql_query, client_id, "current_year_savings")

def calculate_current_year_debt_for_user():
    """Calculate current year debt for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH active_debts AS (
            -- Calculate annual debt payments for all active loans
            SELECT
                client_id,
                account_name,
                total_value,
                interest_rate,
                loan_term_in_years,
                payment_frequency,
                loan_date,
                -- Calculate annual payment using amortization formula or simplified estimate
                CASE
                  WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                    -- Standard amortization: monthly payment * 12
                    ABS(total_value) * (interest_rate / 12) /
                    (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                  ELSE
                    -- For loans missing rate/term, assume 12-month repayment
                    ABS(total_value) / 12
                END as annual_payment
            FROM core.liability_note_accounts
            WHERE client_id = %s -- Filter for specific client
              AND total_value < 0  -- Only include actual debt (negative values)
              AND repayment_type = 'PrincipalAndInterest'  -- Only active debt being serviced
              -- DYNAMIC CURRENT YEAR LOGIC:
              AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)  -- Loan originated before or in current year
              AND (loan_term_in_years IS NULL OR
                   EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))  -- Still active in current year
        )
        SELECT
            ROUND(SUM(annual_payment),2) as current_year_debt
        FROM active_debts;
    """
    return execute_metric_query(sql_query, client_id, "current_year_debt")

def calculate_current_year_taxes_for_user():
    """Calculate current year taxes for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) as current_year_taxes
        FROM core.incomes
        WHERE client_id = %s -- Filter for specific client
          AND current_year_amount IS NOT NULL;
    """
    return execute_metric_query(sql_query, client_id, "current_year_taxes")

def calculate_current_year_living_expenses_for_user():
    """Calculate current year living expenses for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
        FROM core.expenses
        WHERE client_id = %s
            AND type = 'Living'
            AND annual_amount > 0
            -- Check if expense overlaps with current year
            AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
            AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
            -- Ensure logical date ranges
            AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date);
    """
    return execute_metric_query(sql_query, client_id, "current_year_living_expenses")

def calculate_total_expenses_for_user():
    """Calculate total expenses for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH
        -- Giving Expense
        giving_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_giving
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Spending'
                AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        ),

        -- Savings Expense
        savings_expense AS (
            SELECT
                COALESCE(SUM(calculated_annual_amount_usd), 0) as current_year_savings
            FROM core.savings
            WHERE start_type = 'Active'
              AND client_id = %s
        ),

        -- Debt Expense
        debt_expense AS (
            WITH active_debts AS (
                SELECT
                    client_id,
                    CASE
                      WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                        ABS(total_value) * (interest_rate / 12) /
                        (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                      ELSE
                        ABS(total_value) / 12
                    END as annual_payment
                FROM core.liability_note_accounts
                WHERE client_id = %s
                  AND total_value < 0
                  AND repayment_type = 'PrincipalAndInterest'
                  AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                  AND (loan_term_in_years IS NULL OR
                       EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))
            )
            SELECT
                COALESCE(SUM(annual_payment), 0) as current_year_debt
            FROM active_debts
        ),

        -- Tax Expense
        tax_expense AS (
            SELECT
                ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) as current_year_taxes
            FROM core.incomes
            WHERE client_id = %s
              AND current_year_amount IS NOT NULL
        ),

        -- Living Expense
        living_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Living'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                -- Ensure logical date ranges
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
        )

        -- Final Total Expense Calculation
        SELECT
            -- Total Expense = Sum of all components
            ROUND((COALESCE(g.current_year_giving, 0) +
             COALESCE(s.current_year_savings, 0) +
             COALESCE(d.current_year_debt, 0) +
             COALESCE(t.current_year_taxes, 0) +
             COALESCE(l.current_year_living_expenses, 0)),2) as total_expense
        FROM giving_expense g, savings_expense s, debt_expense d, tax_expense t, living_expense l;
    """
    return execute_metric_query(sql_query, client_id, "total_expenses")

def calculate_margin_for_user():
    """Calculate margin for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH
        -- Total Income Calculation
        income_breakdown AS (
            SELECT
                client_id,
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s  -- Filter for specific client
              AND current_year_amount IS NOT NULL
            GROUP BY client_id
        ),

        -- Giving Expense
        giving_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_giving
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Spending'
                AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        ),

        -- Savings Expense
        savings_expense AS (
            SELECT
                COALESCE(SUM(calculated_annual_amount_usd), 0) as current_year_savings
            FROM core.savings
            WHERE start_type = 'Active'
              AND client_id = %s
        ),

        -- Debt Expense
        debt_expense AS (
            WITH active_debts AS (
                SELECT
                    client_id,
                    CASE
                      WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                        ABS(total_value) * (interest_rate / 12) /
                        (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                      ELSE
                        ABS(total_value) / 12
                    END as annual_payment
                FROM core.liability_note_accounts
                WHERE client_id = %s
                  AND total_value < 0
                  AND repayment_type = 'PrincipalAndInterest'
                  AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                  AND (loan_term_in_years IS NULL OR
                       EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))
            )
            SELECT
                COALESCE(SUM(annual_payment), 0) as current_year_debt
            FROM active_debts
        ),

        -- Tax Expense
        tax_expense AS (
            SELECT
                ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) as current_year_taxes
            FROM core.incomes
            WHERE client_id = %s
              AND current_year_amount IS NOT NULL
        ),

        -- Living Expense
        living_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Living'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                -- Ensure logical date ranges
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
        ),

        -- Total Expense Calculation
        total_expense_calc AS (
            SELECT
                ROUND((COALESCE(g.current_year_giving, 0) +
                       COALESCE(s.current_year_savings, 0) +
                       COALESCE(d.current_year_debt, 0) +
                       COALESCE(t.current_year_taxes, 0) +
                       COALESCE(l.current_year_living_expenses, 0)), 2) as total_expense
            FROM giving_expense g, savings_expense s, debt_expense d, tax_expense t, living_expense l
        )

        -- Final Margin Calculation (Total Income - Total Expenses)
        SELECT
            -- Margin = Total Income - Total Expenses, rounded to 2 decimal places
            ROUND((COALESCE(i.total_income, 0) - COALESCE(e.total_expense, 0)), 2) as margin
        FROM income_breakdown i, total_expense_calc e;
    """
    return execute_metric_query(sql_query, client_id, "margin")

def calculate_life_insurance_for_user():
    """Calculate life insurance for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(death_benefit), 0) as life_insurance_metric
        FROM core.life_insurance_annuity_accounts
        WHERE fact_type_name = 'Life Insurance'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "life_insurance")

def calculate_disability_for_user():
    """Calculate disability for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT COALESCE(SUM(benefit_amount), 0) as disability_metric
        FROM core.disability_ltc_insurance_accounts
        WHERE fact_type_name IN ('Disability Policy', 'Business Disability Policy')
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "disability")

def calculate_ltc_for_user():
    """Calculate LTC for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(benefit_amount),0) as ltc_metric
        FROM core.disability_ltc_insurance_accounts
        WHERE sub_type = 'PersonalLT'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "ltc")

def calculate_umbrella_for_user():
    """Calculate umbrella for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(maximum_annual_benefit),0) as umbrella_metric
        FROM core.property_casualty_insurance_accounts
        WHERE sub_type = 'Umbrella'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "umbrella")

def calculate_business_insurance_for_user():
    """Calculate business insurance for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(benefit_amount), 0) AS business_insurance
        FROM core.disability_ltc_insurance_accounts
        WHERE sub_type = 'BusinessReducingTerm'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "business_insurance")

def calculate_flood_insurance_for_user():
    """Calculate flood insurance for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            COALESCE(SUM(maximum_annual_benefit), 0) as flood_insurance_metric
        FROM core.property_casualty_insurance_accounts
        WHERE sub_type = 'Flood'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "flood_insurance")

def calculate_at_risk_for_user():
    """Calculate at risk for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH taxable AS (
            SELECT COALESCE(SUM(total_value), 0) AS taxable_investments_usd
            FROM core.investment_deposit_accounts
            WHERE client_id = %s
              AND fact_type_name = 'Taxable Investment'   -- excludes IRA, 401k, 529 etc.
        ),
        umbrella AS (
            SELECT COALESCE(SUM(maximum_annual_benefit), 0) AS umbrella_coverage_usd
            FROM core.property_casualty_insurance_accounts
            WHERE client_id = %s
              AND sub_type = 'Umbrella'
        )
        SELECT (taxable.taxable_investments_usd - umbrella.umbrella_coverage_usd) AS at_risk_usd
        FROM taxable, umbrella;
    """
    return execute_metric_query(sql_query, client_id, "at_risk")

def calculate_retirement_ratio_for_user():
    """Calculate retirement ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH
client_info AS (
    SELECT
        c.client_id,
        EXTRACT(YEAR FROM AGE(CURRENT_DATE, c.hh_date_of_birth)) AS current_age,
        65 AS retirement_age
    FROM core.clients c
    WHERE c.client_id = %s
),
future_income_pv AS (
    SELECT
        i.client_id,
        SUM(
            CASE
                WHEN (ci.retirement_age - ci.current_age) > 0 AND
                     (i.end_type IS NULL OR
                      i.end_type != 'Age' OR
                      (i.end_type = 'Age' AND i.end_value > ci.retirement_age))
                THEN i.annual_amount *
                     (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
                ELSE 0
            END
        ) AS pv_future_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    WHERE i.deleted IS NULL OR i.deleted = false
    GROUP BY i.client_id
),
future_expenses_pv AS (
    SELECT
        e.client_id,
        SUM(
            CASE
                WHEN (ci.retirement_age - ci.current_age) > 0 AND
                     (e.end_type IS NULL OR
                      e.end_type != 'Age' OR
                      (e.end_type = 'Age' AND
                       EXTRACT(YEAR FROM AGE(e.end_actual_date, e.start_actual_date)) > (ci.retirement_age - ci.current_age)))
                THEN e.annual_amount *
                     (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
                ELSE 0
            END
        ) AS pv_future_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
    GROUP BY e.client_id
),
current_assets AS (
    SELECT
        client_id,
        SUM(total_value) AS current_assets
    FROM (
        SELECT client_id, total_value FROM core.investment_deposit_accounts
        UNION ALL
        SELECT client_id, total_value FROM core.real_estate_assets
        UNION ALL
        SELECT client_id, total_value FROM core.personal_property_accounts
    ) all_assets
    GROUP BY client_id
),
retirement_savings AS (
    SELECT
        client_id,
        SUM(COALESCE(calculated_annual_amount_usd, fixed_amount_usd)) AS retirement_savings
    FROM core.savings
    WHERE destination ~* 'retirement|401k|ira' OR
          account_id ~* 'retirement|401k|ira'
    GROUP BY client_id
),
current_liabilities AS (
    SELECT
        client_id,
        SUM(total_value) AS current_liabilities
    FROM core.liability_note_accounts
    GROUP BY client_id
)
SELECT
    ROUND(
        (
            COALESCE(fi.pv_future_income, 0) +
            COALESCE(ca.current_assets, 0) +
            COALESCE(rs.retirement_savings, 0)
        ) /
        NULLIF(
            (COALESCE(fe.pv_future_expenses, 0) + COALESCE(cl.current_liabilities, 0)),
            0
        ),
        2
    ) AS retirement_ratio
FROM client_info ci
LEFT JOIN future_income_pv fi ON ci.client_id = fi.client_id
LEFT JOIN future_expenses_pv fe ON ci.client_id = fe.client_id
LEFT JOIN current_assets ca ON ci.client_id = ca.client_id
LEFT JOIN retirement_savings rs ON ci.client_id = rs.client_id
LEFT JOIN current_liabilities cl ON ci.client_id = cl.client_id
WHERE ci.current_age < ci.retirement_age;
    """
    result = execute_metric_query(sql_query, client_id, "retirement_ratio")
    try:
        current_app.logger.info(f"DEBUG: Retirement ratio for client {client_id}: {result}")
    except RuntimeError:
        print(f"DEBUG: Retirement ratio for client {client_id}: {result}")
    return result

def calculate_survivor_ratio_for_user():
    """Calculate survivor ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH
        -- Get client information and calculate current age
        client_info AS (
            SELECT
                c.client_id,
                EXTRACT(YEAR FROM AGE(CURRENT_DATE, c.hh_date_of_birth)) AS current_age
            FROM core.clients c
            WHERE c.client_id = %s
        ),

        -- Calculate present value of future income streams (post-death scenario)
        future_income_pv AS (
            SELECT
                i.client_id,
                SUM(
                    CASE
                        -- Income that continues after death (spouse income, survivor benefits, etc.)
                        WHEN (i.end_type = 'SpousesDeath' OR i.owner_type = 'Spouse') AND
                             (i.end_value IS NULL OR i.end_value > EXTRACT(YEAR FROM CURRENT_DATE))
                        THEN i.annual_amount *
                             (1 - POWER(1.0/1.04, 20)) / 0.04  -- 20-year planning horizon
                        ELSE 0
                    END
                ) AS pv_future_income
            FROM core.incomes i
            JOIN client_info ci ON i.client_id = ci.client_id
            WHERE (i.deleted IS NULL OR i.deleted = false)
            GROUP BY i.client_id
        ),

        -- Calculate present value of future expenses (post-death scenario)
        future_expenses_pv AS (
            SELECT
                e.client_id,
                SUM(
                    CASE
                        -- Expenses that continue after death (ongoing living expenses, etc.)
                        WHEN e.end_type != 'AtSecondDeath' AND
                             (e.end_actual_date IS NULL OR e.end_actual_date > CURRENT_DATE)
                        THEN e.annual_amount *
                             (1 - POWER(1.0/1.04, 20)) / 0.04  -- 20-year planning horizon
                        ELSE 0
                    END
                ) AS pv_future_expenses
            FROM core.expenses e
            JOIN client_info ci ON e.client_id = ci.client_id
            GROUP BY e.client_id  -- Fixed: was i.client_id, now e.client_id
        ),

        -- Current assets
        current_assets AS (
            SELECT
                client_id,
                SUM(total_value) AS current_assets
            FROM (
                SELECT client_id, total_value FROM core.investment_deposit_accounts
                UNION ALL
                SELECT client_id, total_value FROM core.real_estate_assets
                UNION ALL
                SELECT client_id, total_value FROM core.personal_property_accounts
            ) all_assets
            GROUP BY client_id
        ),

        -- Life insurance death benefits
        life_insurance AS (
            SELECT
                client_id,
                SUM(death_benefit) AS life_insurance_value
            FROM core.life_insurance_annuity_accounts
            WHERE death_benefit IS NOT NULL AND death_benefit > 0
            GROUP BY client_id
        ),

        -- Current liabilities (note: these are negative values in the data)
        current_liabilities AS (
            SELECT
                client_id,
                ABS(SUM(total_value)) AS current_liabilities  -- Convert to positive for calculation
            FROM core.liability_note_accounts
            GROUP BY client_id
        )

        -- Final survivor ratio calculation
        SELECT
            ROUND(
                (
                    COALESCE(li.life_insurance_value, 0) +
                    COALESCE(fi.pv_future_income, 0) +
                    COALESCE(ca.current_assets, 0)
                ) /
                NULLIF(
                    (COALESCE(fe.pv_future_expenses, 0) + COALESCE(cl.current_liabilities, 0)),
                    0
                ),
                2
            ) AS survivor_ratio
        FROM client_info ci
        LEFT JOIN future_income_pv fi ON ci.client_id = fi.client_id
        LEFT JOIN future_expenses_pv fe ON ci.client_id = fe.client_id
        LEFT JOIN current_assets ca ON ci.client_id = ca.client_id
        LEFT JOIN life_insurance li ON ci.client_id = li.client_id
        LEFT JOIN current_liabilities cl ON ci.client_id = cl.client_id;
    """
    result = execute_metric_query(sql_query, client_id, "survivor_ratio")
    try:
        current_app.logger.info(f"DEBUG: Survivor ratio for client {client_id}: {result}")
    except RuntimeError:
        print(f"DEBUG: Survivor ratio for client {client_id}: {result}")
    return result

def calculate_education_ratio_for_user():
    """Calculate education ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
       WITH
-- Get client information
client_info AS (
    SELECT
        c.client_id
    FROM core.clients c
    WHERE c.client_id = %s
),

-- Calculate present value of education savings (annual contributions)
education_savings_pv AS (
    SELECT
        s.client_id,
        SUM(
            CASE
                WHEN s.destination ~* 'education'
                THEN COALESCE(s.calculated_annual_amount_usd, s.fixed_amount_usd) *
                     (1 - POWER(1.0/1.04, 10)) / 0.04  -- 10-year education planning horizon
                ELSE 0
            END
        ) AS pv_education_savings
    FROM core.savings s
    JOIN client_info ci ON s.client_id = ci.client_id
    GROUP BY s.client_id
),

-- Current education account balances
education_accounts AS (
    SELECT
        client_id,
        SUM(total_value) AS education_account_balances
    FROM (
        -- Investment accounts with education subtype
        SELECT client_id, total_value
        FROM core.investment_deposit_accounts
        WHERE sub_type ~* 'education'

        UNION ALL

        -- Personal property accounts (all included as they may contain education assets)
        SELECT client_id, total_value
        FROM core.personal_property_accounts
    ) edu_accounts
    GROUP BY client_id
),

-- Calculate present value of future education expenses
education_expenses_pv AS (
    SELECT
        e.client_id,
        SUM(
            CASE
                WHEN e.type ~* 'education' OR
                     e.sub_type ~* 'education' OR
                     e.expense_item ~* 'education'
                THEN e.annual_amount *
                     (1 - POWER(1.0/1.04, 10)) / 0.04  -- 10-year education planning horizon
                ELSE 0
            END
        ) AS pv_education_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
    GROUP BY e.client_id
)

-- Final education ratio calculation
SELECT
    ROUND(
        (
            COALESCE(es.pv_education_savings, 0) +
            COALESCE(ea.education_account_balances, 0)
        ) /
        NULLIF(COALESCE(ee.pv_education_expenses, 0), 0),
        2
    ) AS education_ratio
FROM client_info ci
LEFT JOIN education_savings_pv es ON ci.client_id = es.client_id
LEFT JOIN education_accounts ea ON ci.client_id = ea.client_id
LEFT JOIN education_expenses_pv ee ON ci.client_id = ee.client_id;

    """
    result = execute_metric_query(sql_query, client_id, "education_ratio")
    try:
        current_app.logger.info(f"DEBUG: Education ratio for client {client_id}: {result}")
    except RuntimeError:
        print(f"DEBUG: Education ratio for client {client_id}: {result}")
    return result

def calculate_new_cars_ratio_for_user():
    """Calculate new cars ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH
-- Get client information
client_info AS (
    SELECT
        c.client_id
    FROM core.clients c
    WHERE c.client_id = %s
),

-- Current taxable account balances (investment accounts flagged as taxable)
taxable_accounts AS (
    SELECT
        client_id,
        SUM(total_value) AS taxable_account_value
    FROM core.investment_deposit_accounts
    WHERE sub_type ~* 'taxable'
       OR account_name ~* 'taxable|brokerage'
    GROUP BY client_id
),

-- Calculate present value of taxable savings (annual contributions to taxable accounts)
taxable_savings_pv AS (
    SELECT
        s.client_id,
        SUM(
            CASE
                WHEN NOT (s.destination ~* 'retirement|education')
                THEN COALESCE(s.calculated_annual_amount_usd, s.fixed_amount_usd) *
                     (1 - POWER(1.0/1.04, 5)) / 0.04  -- 5-year car planning horizon
                ELSE 0
            END
        ) AS pv_taxable_savings
    FROM core.savings s
    JOIN client_info ci ON s.client_id = ci.client_id
    GROUP BY s.client_id
),

-- Calculate present value of future car expenses
car_expenses_pv AS (
    SELECT
        e.client_id,
        SUM(
            CASE
                WHEN e.expense_item ~* 'car|vehicle|auto' OR
                     e.type ~* 'car|vehicle|auto' OR
                     e.sub_type ~* 'car|vehicle|auto'
                THEN e.annual_amount *
                     (1 - POWER(1.0/1.04, 5)) / 0.04  -- 5-year car planning horizon
                ELSE 0
            END
        ) AS pv_car_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
    GROUP BY e.client_id
)

-- Final new cars ratio calculation
SELECT
    ROUND(
        (
            COALESCE(ta.taxable_account_value, 0) +
            COALESCE(ts.pv_taxable_savings, 0)
        ) /
        NULLIF(COALESCE(ce.pv_car_expenses, 0), 0),
        2
    ) AS new_cars_ratio
FROM client_info ci
LEFT JOIN taxable_accounts ta ON ci.client_id = ta.client_id
LEFT JOIN taxable_savings_pv ts ON ci.client_id = ts.client_id
LEFT JOIN car_expenses_pv ce ON ci.client_id = ce.client_id;

    """
    result = execute_metric_query(sql_query, client_id, "new_cars_ratio")
    try:
        current_app.logger.info(f"DEBUG: New cars ratio for client {client_id}: {result}")
    except RuntimeError:
        print(f"DEBUG: New cars ratio for client {client_id}: {result}")
    return result

def calculate_ltc_ratio_for_user():
    """Calculate LTC ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        WITH
-- Get client information
client_info AS (
    SELECT
        c.client_id
    FROM core.clients c
    WHERE c.client_id = %s
),

-- Calculate present value of all future income streams
future_income_pv AS (
    SELECT
        i.client_id,
        SUM(
            CASE
                WHEN (i.deleted IS NULL OR i.deleted = false)
                THEN i.annual_amount *
                     (1 - POWER(1.0/1.04, 20)) / 0.04  -- 20-year planning horizon for LTC
                ELSE 0
            END
        ) AS pv_future_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    GROUP BY i.client_id
),

-- Current assets (investment + real estate + personal property)
current_assets AS (
    SELECT
        client_id,
        SUM(total_value) AS total_assets
    FROM (
        SELECT client_id, total_value FROM core.investment_deposit_accounts
        UNION ALL
        SELECT client_id, total_value FROM core.real_estate_assets
        UNION ALL
        SELECT client_id, total_value FROM core.personal_property_accounts
    ) all_assets
    GROUP BY client_id
),

-- Calculate present value of future regular expenses (excluding LTC)
future_expenses_pv AS (
    SELECT
        e.client_id,
        SUM(
            CASE
                WHEN NOT (e.type ~* 'ltc' OR e.expense_item ~* 'long term care')
                THEN e.annual_amount *
                     (1 - POWER(1.0/1.04, 20)) / 0.04  -- 20-year planning horizon
                ELSE 0
            END
        ) AS pv_future_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
    GROUP BY e.client_id
),

-- Calculate present value of future LTC expenses (premiums only)
ltc_expenses_pv AS (
    SELECT
        l.client_id,
        SUM(
            CASE
                WHEN l.sub_type ~* 'ltc' OR l.fact_type_name ~* 'long term care'
                THEN COALESCE(l.annual_premium, 0) *
                     (1 - POWER(1.0/1.04, 20)) / 0.04  -- 20-year planning horizon
                ELSE 0
            END
        ) AS pv_ltc_expenses
    FROM core.disability_ltc_insurance_accounts l
    JOIN client_info ci ON l.client_id = ci.client_id
    GROUP BY l.client_id
)

-- Final LTC ratio calculation
SELECT
    ROUND(
        (
            COALESCE(fi.pv_future_income, 0) +
            COALESCE(ca.total_assets, 0)
        ) /
        NULLIF(
            (COALESCE(fe.pv_future_expenses, 0) + COALESCE(le.pv_ltc_expenses, 0)),
            0
        ),
        2
    ) AS ltc_ratio
FROM client_info ci
LEFT JOIN future_income_pv fi ON ci.client_id = fi.client_id
LEFT JOIN current_assets ca ON ci.client_id = ca.client_id
LEFT JOIN future_expenses_pv fe ON ci.client_id = fe.client_id
LEFT JOIN ltc_expenses_pv le ON ci.client_id = le.client_id;

    """
    result = execute_metric_query(sql_query, client_id, "ltc_ratio")
    try:
        current_app.logger.info(f"DEBUG: LTC ratio for client {client_id}: {result}")
    except RuntimeError:
        print(f"DEBUG: LTC ratio for client {client_id}: {result}")
    return result

def calculate_ltd_ratio_for_user():
    """Calculate LTD ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
       WITH
-- Get client information
client_info AS (
    SELECT
        c.client_id
    FROM core.clients c
    WHERE c.client_id = %s
),

-- LTD value (current value/benefit amount of LTD policies)
ltd_value AS (
    SELECT
        client_id,
        SUM(COALESCE(benefit_amount, 0)) AS ltd_value
    FROM core.disability_ltc_insurance_accounts
    WHERE fact_type_name ~* 'disability'
    GROUP BY client_id
),

-- Current earned income (salary/wage income for current year)
earned_income AS (
    SELECT
        client_id,
        SUM(COALESCE(current_year_amount, 0)) AS earned_income
    FROM core.incomes
    WHERE income_type = 'Salary'
      AND (deleted IS NULL OR deleted = false)
    GROUP BY client_id
)

-- Final LTD ratio calculation
SELECT
    ROUND(
        COALESCE(l.ltd_value, 0) / NULLIF(COALESCE(e.earned_income, 0), 0),
        2
    ) AS ltd_ratio
FROM client_info ci
LEFT JOIN ltd_value l ON ci.client_id = l.client_id
LEFT JOIN earned_income e ON ci.client_id = e.client_id;

    """
    result = execute_metric_query(sql_query, client_id, "ltd_ratio")
    try:
        current_app.logger.info(f"DEBUG: LTD ratio for client {client_id}: {result}")
    except RuntimeError:
        print(f"DEBUG: LTD ratio for client {client_id}: {result}")
    return result

def calculate_savings_ratio_for_user():
    """Calculate savings ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            ROUND(COALESCE(savings.current_year_savings, 0) / NULLIF(income.total_income, 0),2) as savings_ratio
        FROM (
            SELECT
                COALESCE(SUM(calculated_annual_amount_usd), 0) as current_year_savings
            FROM core.savings
            WHERE start_type = 'Active'
              AND client_id = %s
        ) savings
        CROSS JOIN (
            SELECT
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s
              AND current_year_amount IS NOT NULL
        ) income;
    """
    return execute_metric_query(sql_query, client_id, "savings_ratio")

def calculate_giving_ratio_for_user():
    """Calculate giving ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            ROUND(COALESCE(giving.current_year_giving, 0) / NULLIF(income.total_income, 0),2) as giving_ratio
        FROM (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_giving
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Spending'
                AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        ) giving
        CROSS JOIN (
            SELECT
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s
              AND current_year_amount IS NOT NULL
        ) income;
    """
    return execute_metric_query(sql_query, client_id, "giving_ratio")

def calculate_reserves_ratio_for_user():
    """Calculate reserves ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            ROUND((COALESCE(cash.total_cash, 0) / NULLIF(expenses.current_year_living_expenses, 0)) * 0.5, 2) as reserves
        FROM (
            WITH holdings_cash AS (
                SELECT
                    client_id,
                    SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) as cash_from_holdings
                FROM core.holdings
                WHERE asset_class = 'cash' AND value IS NOT NULL
                AND client_id = %s
                GROUP BY client_id
            ),
            investment_cash AS (
                SELECT
                    client_id,
                    SUM(COALESCE(cash_balance, 0)) as cash_from_investments
                FROM core.investment_deposit_accounts
                WHERE fact_type_name = 'Cash Alternative'
                    AND cash_balance IS NOT NULL
                    AND client_id = %s
                GROUP BY client_id
            )

            SELECT
                COALESCE(h.cash_from_holdings, 0) + COALESCE(i.cash_from_investments, 0) as total_cash
            FROM holdings_cash h
            FULL OUTER JOIN investment_cash i ON h.client_id = i.client_id
        ) cash
        CROSS JOIN (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Living'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                -- Ensure logical date ranges
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
        ) expenses;
    """
    return execute_metric_query(sql_query, client_id, "reserves_ratio")

def calculate_debt_ratio_for_user():
    """Calculate debt ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            ROUND(house_equity / NULLIF(house_value, 0), 2) as debt
        FROM (
            SELECT
                COALESCE(SUM(rea.total_value), 0) as house_value,
                COALESCE(SUM(rea.total_value), 0) - COALESCE(SUM(CASE WHEN lna.sub_type = 'Mortgage' THEN ABS(lna.total_value) ELSE 0 END), 0) as house_equity
            FROM core.real_estate_assets rea
            LEFT JOIN core.liability_note_accounts lna ON rea.account_id = lna.real_estate_id
            WHERE rea.sub_type = 'Residence'
              AND rea.client_id = %s
              AND (lna.sub_type = 'Mortgage' OR lna.sub_type IS NULL)
            GROUP BY rea.client_id
        ) house_data;
    """
    return execute_metric_query(sql_query, client_id, "debt_ratio")

def calculate_diversification_ratio_for_user():
    """Calculate diversification ratio for the authenticated user."""
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            ROUND(1 - (largest_holding / NULLIF(total_portfolio, 0)), 2) as diversification
        FROM (
            SELECT
                (SELECT MAX(value)
                 FROM core.holdings
                 WHERE client_id = %s
                   AND value IS NOT NULL
                   AND value > 0) as largest_holding,
                (SELECT SUM(portfolio_value)
                 FROM (
                     SELECT COALESCE(SUM(value), 0) AS portfolio_value
                     FROM core.holdings
                     WHERE client_id = %s
                     
                     UNION ALL
                     
                     SELECT COALESCE(SUM(total_value), 0)
                    FROM core.investment_deposit_accounts
                    WHERE client_id = %s
                 ) portfolio_components) as total_portfolio
        ) calculations;
    """
    return execute_metric_query(sql_query, client_id, "diversification_ratio")

# Client-scoped Wisdom Index ratios (admin/summary usage)
def calculate_savings_ratio_for_client(client_id):
    """Calculate savings ratio for a specific client."""
    sql_query = """
        SELECT
            ROUND(COALESCE(savings.current_year_savings, 0) / NULLIF(income.total_income, 0),2) as savings_ratio
        FROM (
            SELECT
                COALESCE(SUM(calculated_annual_amount_usd), 0) as current_year_savings
            FROM core.savings
            WHERE start_type = 'Active'
              AND client_id = %s
        ) savings
        CROSS JOIN (
            SELECT
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s
              AND current_year_amount IS NOT NULL
        ) income;
    """
    return execute_metric_query(sql_query, client_id, "savings_ratio_for_client")

def calculate_giving_ratio_for_client(client_id):
    """Calculate giving ratio for a specific client."""
    sql_query = """
        SELECT
            ROUND(COALESCE(giving.current_year_giving, 0) / NULLIF(income.total_income, 0),2) as giving_ratio
        FROM (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_giving
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Spending'
                AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        ) giving
        CROSS JOIN (
            SELECT
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s
              AND current_year_amount IS NOT NULL
        ) income;
    """
    return execute_metric_query(sql_query, client_id, "giving_ratio_for_client")

def calculate_reserves_ratio_for_client(client_id):
    """Calculate reserves ratio for a specific client."""
    sql_query = """
        SELECT
            ROUND((COALESCE(cash.total_cash, 0) / NULLIF(expenses.current_year_living_expenses, 0)) * 0.5, 2) as reserves
        FROM (
            WITH holdings_cash AS (
                SELECT
                    client_id,
                    SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) as cash_from_holdings
                FROM core.holdings
                WHERE asset_class = 'cash' AND value IS NOT NULL
                AND client_id = %s
                GROUP BY client_id
            ),
            investment_cash AS (
                SELECT
                    client_id,
                    SUM(COALESCE(cash_balance, 0)) as cash_from_investments
                FROM core.investment_deposit_accounts
                WHERE fact_type_name = 'Cash Alternative'
                    AND cash_balance IS NOT NULL
                    AND client_id = %s
                GROUP BY client_id
            )

            SELECT
                COALESCE(h.cash_from_holdings, 0) + COALESCE(i.cash_from_investments, 0) as total_cash
            FROM holdings_cash h
            FULL OUTER JOIN investment_cash i ON h.client_id = i.client_id
        ) cash
        CROSS JOIN (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Living'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                -- Ensure logical date ranges
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
        ) expenses;
    """
    return execute_metric_query(sql_query, client_id, "reserves_ratio_for_client")

def calculate_debt_ratio_for_client(client_id):
    """Calculate debt ratio for a specific client."""
    sql_query = """
        SELECT
            ROUND(house_equity / NULLIF(house_value, 0), 2) as debt
        FROM (
            SELECT
                COALESCE(SUM(rea.total_value), 0) as house_value,
                COALESCE(SUM(rea.total_value), 0) - COALESCE(SUM(CASE WHEN lna.sub_type = 'Mortgage' THEN ABS(lna.total_value) ELSE 0 END), 0) as house_equity
            FROM core.real_estate_assets rea
            LEFT JOIN core.liability_note_accounts lna ON rea.account_id = lna.real_estate_id
            WHERE rea.sub_type = 'Residence'
              AND rea.client_id = %s
              AND (lna.sub_type = 'Mortgage' OR lna.sub_type IS NULL)
            GROUP BY rea.client_id
        ) house_data;
    """
    return execute_metric_query(sql_query, client_id, "debt_ratio_for_client")

def calculate_diversification_ratio_for_client(client_id):
    """Calculate diversification ratio for a specific client."""
    sql_query = """
        SELECT
            ROUND(1 - (largest_holding / NULLIF(total_portfolio, 0)), 2) as diversification
        FROM (
            SELECT
                (SELECT MAX(value)
                 FROM core.holdings
                 WHERE client_id = %s
                   AND value IS NOT NULL
                   AND value > 0) as largest_holding,
                (SELECT SUM(portfolio_value)
                 FROM (
                     SELECT COALESCE(SUM(value), 0) AS portfolio_value
                     FROM core.holdings
                     WHERE client_id = %s
                     
                     UNION ALL
                     
                     SELECT COALESCE(SUM(total_value), 0)
                     FROM core.investment_deposit_accounts
                     WHERE client_id = %s
                 ) portfolio_components) as total_portfolio
        ) calculations;
    """
    return execute_metric_query(sql_query, client_id, "diversification_ratio_for_client")

def execute_chart_query(sql_query, client_id, chart_name="chart"):
    """
    Generic function to execute a parameterized chart query for a specific client.
    
    Args:
        sql_query (str): The SQL query to execute with %s placeholder for client_id
        client_id (int): The client ID to use in the query
        chart_name (str): Name of the chart for error reporting
        
    Returns:
        list: List of dictionaries with chart data
    """
    try:
        client_id = int(client_id)
        if client_id <= 0:
            print(f"Warning: {chart_name} - client_id must be a positive integer, got {client_id}")
            return []
    except (ValueError, TypeError):
        print(f"Warning: {chart_name} - client_id must be a positive integer, got {client_id}")
        return []

    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Parameterize the query to prevent SQL injection
        import re
        param_count = len(re.findall(r'(?<!%)%s(?!%)', sql_query))
        
        # Debug logging for parameter counting
        try:
            current_app.logger.info(f"DEBUG: {chart_name} query found {param_count} %s placeholders")
            # Show first 200 chars of query for debugging
            query_preview = sql_query[:200] + "..." if len(sql_query) > 200 else sql_query
            current_app.logger.info(f"DEBUG: Query preview: {query_preview}")
        except RuntimeError:
            print(f"DEBUG: {chart_name} query found {param_count} %s placeholders")
            query_preview = sql_query[:200] + "..." if len(sql_query) > 200 else sql_query
            print(f"DEBUG: Query preview: {query_preview}")
        
        if param_count > 0:
            params = tuple(client_id for _ in range(param_count))
        else:
            params = ()
        
        # Debug logging
        try:
            current_app.logger.info(f"DEBUG: Executing {chart_name} query for client {client_id} with {param_count} parameters: {params}")
        except RuntimeError:
            print(f"DEBUG: Executing {chart_name} query for client {client_id} with {param_count} parameters: {params}")
        
        cursor.execute(sql_query, params)
        
        results = cursor.fetchall()
        
        # CRITICAL: Get column names BEFORE closing cursor
        # In psycopg2, cursor.description becomes unavailable after cursor.close()
        column_names = [desc[0] for desc in cursor.description] if cursor.description else []
        
        cursor.close()
        
        # Debug logging
        try:
            current_app.logger.info(f"DEBUG: {chart_name} query returned {len(results)} rows")
        except RuntimeError:
            print(f"DEBUG: {chart_name} query returned {len(results)} rows")
        
        # Convert results to list of dictionaries
        chart_data = []
        if results:
            # Debug log column names
            try:
                current_app.logger.info(f"DEBUG: {chart_name} column_names: {column_names}")
            except RuntimeError:
                print(f"DEBUG: {chart_name} column_names: {column_names}")
            
            for row in results:
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(column_names):
                        # Convert to appropriate Python types
                        # Handle Decimal objects from PostgreSQL numeric types
                        from decimal import Decimal
                        if isinstance(value, Decimal):
                            row_dict[column_names[i]] = float(value)
                        elif isinstance(value, (int, float)):
                            row_dict[column_names[i]] = float(value) if '.' in str(value) else int(value)
                        else:
                            row_dict[column_names[i]] = value
                chart_data.append(row_dict)
        
        # Debug log final chart data
        try:
            current_app.logger.info(f"DEBUG: {chart_name} returning chart_data: {chart_data}")
        except RuntimeError:
            print(f"DEBUG: {chart_name} returning chart_data: {chart_data}")
        
        return chart_data
        
    except Exception as e:
        print(f"Error executing {chart_name} query for client_id {client_id}: {e}")
        return []
    finally:
        if connection:
            close_db_connection(connection)

def get_income_chart_data_for_user():
    """
    Get income data for bar chart visualization for the authenticated user.
    
    Returns:
        list: List of dictionaries with income categories and amounts
    """
    client_id = get_jwt_identity()
    
    sql_query = """
    -- Combined Income Metrics for Bar Chart Visualization
    -- Returns each income category with its amount for a specific client

    -- Earned Income
    SELECT
        'Earned Income' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Salary')
        AND client_id = %s

    UNION ALL

    -- Social Security Income
    SELECT
        'Social Security' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('SocialSecurity')
        AND client_id = %s

    UNION ALL

    -- Pension Income
    SELECT
        'Pension' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Pension')
        AND client_id = %s

    UNION ALL

    -- Real Estate Income
    SELECT
        'Real Estate' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Real Estate')
        AND client_id = %s

    UNION ALL

    -- Business Income
    SELECT
        'Business' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Business')
        AND client_id = %s

    ORDER BY income_category;
    """
    
    return execute_chart_query(sql_query, client_id, "income_bar_chart")

def get_expense_chart_data_for_user():
    """
    Get expense data for pie chart visualization for the authenticated user.
    
    Returns:
        list: List of dictionaries with expense categories and amounts
    """
    client_id = get_jwt_identity()
    
    sql_query = """
    -- Combined Expense Metrics for Pie Chart Visualization
    -- Returns each expense category with its amount for a specific client

    -- Giving Expense
    SELECT
        'Giving' AS expense_category,
        COALESCE(SUM(annual_amount), 0) AS amount
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Spending'
        AND sub_type = 'GivingAndPhilanthropy'
        AND annual_amount > 0
        -- Check if expense overlaps with current year
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))

    UNION ALL

    -- Savings Expense
    SELECT
        'Savings' AS expense_category,
        COALESCE(SUM(calculated_annual_amount_usd), 0) AS amount
    FROM core.savings
    WHERE start_type = 'Active'  -- Only include currently active savings plans
      AND client_id = %s

    UNION ALL

    -- Debt Expense
    SELECT
        'Debt' AS expense_category,
        ROUND(SUM(
            CASE
              WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                ABS(total_value) * (interest_rate / 12) /
                (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
              ELSE
                ABS(total_value) / 12
            END), 2) AS amount
    FROM core.liability_note_accounts
    WHERE client_id = %s
      AND total_value < 0  -- Only include actual debt (negative values)
      AND repayment_type = 'PrincipalAndInterest'  -- Only active debt being serviced
      -- DYNAMIC CURRENT YEAR LOGIC:
      AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)  -- Loan originated before or in current year
      AND (loan_term_in_years IS NULL OR
           EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))  -- Still active in current year

    UNION ALL

    -- Tax Expense
    SELECT
        'Taxes' AS expense_category,
        ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) AS amount
    FROM core.incomes
    WHERE client_id = %s
      AND current_year_amount IS NOT NULL

    UNION ALL

    -- Living Expense
    SELECT
        'Living' AS expense_category,
        COALESCE(SUM(annual_amount), 0) AS amount
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Living'
        AND annual_amount > 0
        -- Check if expense overlaps with current year
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        -- Ensure logical date ranges
        AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)

    ORDER BY expense_category;
    """
    
    return execute_chart_query(sql_query, client_id, "expense_pie_chart")

def get_income_chart_data_for_client(client_id):
    """
    Get income data for bar chart visualization for a specific client (admin function).
    
    Args:
        client_id (int): The client ID to get data for
        
    Returns:
        list: List of dictionaries with income categories and amounts
    """
    sql_query = """
    -- Combined Income Metrics for Bar Chart Visualization
    -- Returns each income category with its amount for a specific client

    -- Earned Income
    SELECT
        'Earned Income' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Salary')
        AND client_id = %s

    UNION ALL

    -- Social Security Income
    SELECT
        'Social Security' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('SocialSecurity')
        AND client_id = %s

    UNION ALL

    -- Pension Income
    SELECT
        'Pension' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Pension')
        AND client_id = %s

    UNION ALL

    -- Real Estate Income
    SELECT
        'Real Estate' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Real Estate')
        AND client_id = %s

    UNION ALL

    -- Business Income
    SELECT
        'Business' AS income_category,
        COALESCE(SUM(current_year_amount), 0) AS amount
    FROM core.incomes
    WHERE income_type IN ('Business')
        AND client_id = %s

    ORDER BY income_category;
    """
    
    return execute_chart_query(sql_query, client_id, "income_bar_chart_for_client")

def get_expense_chart_data_for_client(client_id):
    """
    Get expense data for pie chart visualization for a specific client (admin function).
    
    Args:
        client_id (int): The client ID to get data for
        
    Returns:
        list: List of dictionaries with expense categories and amounts
    """
    sql_query = """
    -- Combined Expense Metrics for Pie Chart Visualization
    -- Returns each expense category with its amount for a specific client

    -- Giving Expense
    SELECT
        'Giving' AS expense_category,
        COALESCE(SUM(annual_amount), 0) AS amount
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Spending'
        AND sub_type = 'GivingAndPhilanthropy'
        AND annual_amount > 0
        -- Check if expense overlaps with current year
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))

    UNION ALL

    -- Savings Expense
    SELECT
        'Savings' AS expense_category,
        COALESCE(SUM(calculated_annual_amount_usd), 0) AS amount
    FROM core.savings
    WHERE start_type = 'Active'  -- Only include currently active savings plans
      AND client_id = %s

    UNION ALL

    -- Debt Expense
    SELECT
        'Debt' AS expense_category,
        ROUND(SUM(
            CASE
              WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                ABS(total_value) * (interest_rate / 12) /
                (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
              ELSE
                ABS(total_value) / 12
            END), 2) AS amount
    FROM core.liability_note_accounts
    WHERE client_id = %s
      AND total_value < 0  -- Only include actual debt (negative values)
      AND repayment_type = 'PrincipalAndInterest'  -- Only active debt being serviced
      -- DYNAMIC CURRENT YEAR LOGIC:
      AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)  -- Loan originated before or in current year
      AND (loan_term_in_years IS NULL OR
           EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))  -- Still active in current year

    UNION ALL

    -- Tax Expense
    SELECT
        'Taxes' AS expense_category,
        ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) AS amount
    FROM core.incomes
    WHERE client_id = %s
      AND current_year_amount IS NOT NULL

    UNION ALL

    -- Living Expense
    SELECT
        'Living' AS expense_category,
        COALESCE(SUM(annual_amount), 0) AS amount
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Living'
        AND annual_amount > 0
        -- Check if expense overlaps with current year
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        -- Ensure logical date ranges
        AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)

    ORDER BY expense_category;
    """
    
    return execute_chart_query(sql_query, client_id, "expense_pie_chart_for_client")

def get_treemap_chart_data_for_client(client_id):
    """
    Get treemap chart data for a specific client (admin function).
    
    Args:
        client_id (int): The client ID to get data for
        
    Returns:
        list: List of dictionaries with treemap data
    """
    sql_query = """
    -- Optimized Treemap Query
    SELECT
        'Equity' as category,
        COALESCE(h.equity_holdings, 0) + COALESCE(i.investment_equity, 0) as value
    FROM
        (SELECT client_id, COALESCE(SUM(value),0) as equity_holdings
         FROM core.holdings
         WHERE asset_class IN ('largecap', 'smallcap', 'largevalue', 'smallvalue', 'internat', 'emerging', 'ips')
             AND client_id = %s AND value IS NOT NULL
         GROUP BY client_id) h
    FULL OUTER JOIN
        (SELECT client_id, COALESCE(SUM(holdings_value),0) as investment_equity
         FROM core.investment_deposit_accounts
         WHERE fact_type_name IN ('Taxable Investment', 'Roth IRA', 'Qualified Retirement')
             AND client_id = %s AND holdings_value IS NOT NULL
         GROUP BY client_id) i ON h.client_id = i.client_id
    WHERE COALESCE(h.client_id, i.client_id) = %s

    UNION ALL

    SELECT
        'Cash',
        COALESCE(h.cash_holdings, 0) + COALESCE(i.cash_alternatives, 0)
    FROM
        (SELECT client_id, SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) as cash_holdings
         FROM core.holdings
         WHERE asset_class = 'cash' AND client_id = %s
         GROUP BY client_id) h
    FULL OUTER JOIN
        (SELECT client_id, SUM(COALESCE(cash_balance, 0)) as cash_alternatives
         FROM core.investment_deposit_accounts
         WHERE fact_type_name = 'Cash Alternative' AND client_id = %s
         GROUP BY client_id) i ON h.client_id = i.client_id
    WHERE COALESCE(h.client_id, i.client_id) = %s

    UNION ALL

    SELECT 'Real Estate', COALESCE(SUM(total_value),0)
    FROM core.real_estate_assets
    WHERE client_id = %s
    GROUP BY client_id

    UNION ALL

    SELECT 'Fixed Income', COALESCE(SUM(value),0)
    FROM core.holdings
    WHERE asset_class IN ('highyldbond', 'inttermmun', 'investbond', 'shortermbond', 'shortermmun')
        AND client_id = %s
    GROUP BY client_id

    ORDER BY category;
    """
    
    return execute_chart_query(sql_query, client_id, "treemap_chart_for_client")

def get_wisdom_index_chart_data_for_client(client_id):
    """
    Get Wisdom Index ratios data for bar chart visualization for a specific client (admin function).
    
    Args:
        client_id (int): The client ID to get data for
        
    Returns:
        list: List of dictionaries with Wisdom Index ratios
    """
    sql_query = """
-- Unified Bar Chart Query for 9 Financial Ratios
-- Returns data in format: metric_name, metric_value

WITH
-- Client information for age-based calculations
client_info AS (
    SELECT
        client_id,
        EXTRACT(YEAR FROM AGE(CURRENT_DATE, hh_date_of_birth)) AS current_age,
        65 AS retirement_age
    FROM core.clients
    WHERE client_id = %s
),

-- Common data aggregations
income_data AS (
    SELECT
        SUM(CASE WHEN income_type = 'Salary' THEN current_year_amount ELSE 0 END) AS earned_income,
        SUM(current_year_amount) AS total_income
    FROM core.incomes
    WHERE client_id = %s AND current_year_amount IS NOT NULL
),

savings_data AS (
    SELECT SUM(calculated_annual_amount_usd) AS current_year_savings
    FROM core.savings
    WHERE client_id = %s AND start_type = 'Active'
),

giving_data AS (
    SELECT SUM(annual_amount) AS current_year_giving
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Spending'
        AND sub_type = 'GivingAndPhilanthropy'
        AND annual_amount > 0
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
),

living_expenses_data AS (
    SELECT SUM(annual_amount) AS current_year_living_expenses
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Living'
        AND annual_amount > 0
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
),

cash_data AS (
    SELECT
        SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) AS cash_from_holdings,
        SUM(value) AS total_holdings,
        MAX(value) AS largest_holding
    FROM core.holdings
    WHERE client_id = %s AND value IS NOT NULL AND value > 0
),

investment_cash_data AS (
    SELECT
        SUM(COALESCE(cash_balance, 0)) AS cash_from_investments,
        SUM(total_value) AS total_investments
    FROM core.investment_deposit_accounts
    WHERE client_id = %s
),

real_estate_data AS (
    SELECT
        SUM(rea.total_value) AS house_value,
        SUM(rea.total_value) - COALESCE(SUM(CASE WHEN lna.sub_type = 'Mortgage' THEN ABS(lna.total_value) ELSE 0 END), 0) AS house_equity
    FROM core.real_estate_assets rea
    LEFT JOIN core.liability_note_accounts lna ON rea.account_id = lna.real_estate_id
    WHERE rea.sub_type = 'Residence'
        AND rea.client_id = %s
        AND (lna.sub_type = 'Mortgage' OR lna.sub_type IS NULL)
    GROUP BY rea.client_id
),

life_insurance_data AS (
    SELECT SUM(death_benefit) AS life_insurance_value
    FROM core.life_insurance_annuity_accounts
    WHERE client_id = %s
        AND death_benefit IS NOT NULL 
        AND death_benefit > 0
),

disability_data AS (
    SELECT SUM(COALESCE(benefit_amount, 0)) AS ltd_value
    FROM core.disability_ltc_insurance_accounts
    WHERE client_id = %s
        AND fact_type_name ~* 'disability'
),

current_assets_data AS (
    SELECT SUM(total_value) AS current_assets
    FROM (
        SELECT total_value FROM core.investment_deposit_accounts WHERE client_id = %s
        UNION ALL
        SELECT total_value FROM core.real_estate_assets WHERE client_id = %s
        UNION ALL
        SELECT total_value FROM core.personal_property_accounts WHERE client_id = %s
    ) all_assets
),

current_liabilities_data AS (
    SELECT ABS(SUM(total_value)) AS current_liabilities
    FROM core.liability_note_accounts
    WHERE client_id = %s
),

-- Present value calculations for retirement ratio
retirement_future_income_pv AS (
    SELECT SUM(
        CASE
            WHEN (ci.retirement_age - ci.current_age) > 0 AND
                 (i.end_type IS NULL OR i.end_type != 'Age' OR (i.end_type = 'Age' AND i.end_value > ci.retirement_age))
            THEN i.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
            ELSE 0
        END
    ) AS pv_retirement_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    WHERE (i.deleted IS NULL OR i.deleted = false)
    GROUP BY ci.client_id
),

retirement_future_expenses_pv AS (
    SELECT SUM(
        CASE
            WHEN (ci.retirement_age - ci.current_age) > 0 AND
                 (e.end_type IS NULL OR e.end_type != 'Age' OR
                  (e.end_type = 'Age' AND EXTRACT(YEAR FROM AGE(e.end_actual_date, e.start_actual_date)) > (ci.retirement_age - ci.current_age)))
            THEN e.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
            ELSE 0
        END
    ) AS pv_retirement_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
    GROUP BY ci.client_id
),

retirement_savings_data AS (
    SELECT SUM(COALESCE(calculated_annual_amount_usd, fixed_amount_usd)) AS retirement_savings
    FROM core.savings
    WHERE client_id = %s
        AND (destination ~* 'retirement|401k|ira' OR account_id ~* 'retirement|401k|ira')
),

-- Present value calculations for survivor ratio
future_income_survivor_pv AS (
    SELECT SUM(
        CASE
            WHEN (i.end_type = 'SpousesDeath' OR i.owner_type = 'Spouse') AND
                 (i.end_value IS NULL OR i.end_value > EXTRACT(YEAR FROM CURRENT_DATE))
            THEN i.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_survivor_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    WHERE (i.deleted IS NULL OR i.deleted = false)
),

future_expenses_survivor_pv AS (
    SELECT SUM(
        CASE
            WHEN e.end_type != 'AtSecondDeath' AND
                 (e.end_actual_date IS NULL OR e.end_actual_date > CURRENT_DATE)
            THEN e.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_survivor_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
),

-- Present value calculations for LTC ratio
future_income_ltc_pv AS (
    SELECT SUM(
        CASE
            WHEN (i.deleted IS NULL OR i.deleted = false)
            THEN i.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_ltc_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    WHERE (i.deleted IS NULL OR i.deleted = false)
),

future_expenses_ltc_pv AS (
    SELECT SUM(
        CASE
            WHEN NOT (e.type ~* 'ltc' OR e.expense_item ~* 'long term care')
            THEN e.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_ltc_regular_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
),

ltc_expenses_data AS (
    SELECT SUM(
        CASE
            WHEN sub_type ~* 'ltc' OR fact_type_name ~* 'long term care'
            THEN COALESCE(annual_premium, 0) * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_ltc_expenses
    FROM core.disability_ltc_insurance_accounts
    WHERE client_id = %s
)

-- Final ratios calculation
SELECT 'Savings Ratio' AS metric_name,
       ROUND(COALESCE(sd.current_year_savings, 0) / NULLIF(id.total_income, 0), 2) AS metric_value
FROM savings_data sd, income_data id

UNION ALL

SELECT 'Giving Ratio' AS metric_name,
       ROUND(COALESCE(gd.current_year_giving, 0) / NULLIF(id.total_income, 0), 2) AS metric_value
FROM giving_data gd, income_data id

UNION ALL

SELECT 'Reserves Ratio' AS metric_name,
       ROUND((COALESCE(cd.cash_from_holdings, 0) + COALESCE(icd.cash_from_investments, 0)) / NULLIF(led.current_year_living_expenses, 0) * 0.5, 2) AS metric_value
FROM cash_data cd, investment_cash_data icd, living_expenses_data led

UNION ALL

SELECT 'Debt Ratio' AS metric_name,
       ROUND(COALESCE(red.house_equity, 0) / NULLIF(red.house_value, 0), 2) AS metric_value
FROM real_estate_data red

UNION ALL

SELECT 'Diversification Ratio' AS metric_name,
       ROUND(1 - (COALESCE(cd.largest_holding, 0) / NULLIF((COALESCE(cd.total_holdings, 0) + COALESCE(icd.total_investments, 0)), 0)), 2) AS metric_value
FROM cash_data cd, investment_cash_data icd

UNION ALL

SELECT 'Survivor Ratio' AS metric_name,
       ROUND(
           (COALESCE(lid.life_insurance_value, 0) + COALESCE(fipv.pv_survivor_income, 0) + COALESCE(cad.current_assets, 0)) /
           NULLIF((COALESCE(fepv.pv_survivor_expenses, 0) + COALESCE(cld.current_liabilities, 0)), 0),
           2
       ) AS metric_value
FROM life_insurance_data lid, future_income_survivor_pv fipv, current_assets_data cad, future_expenses_survivor_pv fepv, current_liabilities_data cld

UNION ALL

SELECT 'Retirement Ratio' AS metric_name,
    CASE
        WHEN ci.current_age < ci.retirement_age THEN
            ROUND(
                (COALESCE(ripv.pv_retirement_income, 0) + COALESCE(cad.current_assets, 0) + COALESCE(rs.retirement_savings, 0)) /
                NULLIF((COALESCE(repv.pv_retirement_expenses, 0) + COALESCE(cld.current_liabilities, 0)), 0),
                2
            )
        ELSE NULL
    END AS metric_value
FROM client_info ci, retirement_future_income_pv ripv, current_assets_data cad, retirement_savings_data rs, retirement_future_expenses_pv repv, current_liabilities_data cld

UNION ALL

SELECT 'LTD Ratio' AS metric_name,
       ROUND(COALESCE(dd.ltd_value, 0) / NULLIF(id.earned_income, 0), 2) AS metric_value
FROM disability_data dd, income_data id

UNION ALL

SELECT 'LTC Ratio' AS metric_name,
       ROUND(
           (COALESCE(filpv.pv_ltc_income, 0) + COALESCE(cad.current_assets, 0)) /
           NULLIF((COALESCE(feipv.pv_ltc_regular_expenses, 0) + COALESCE(ltc.pv_ltc_expenses, 0)), 0),
           2
       ) AS metric_value
FROM future_income_ltc_pv filpv, current_assets_data cad, future_expenses_ltc_pv feipv, ltc_expenses_data ltc

ORDER BY metric_name;
    """
    
    return execute_chart_query(sql_query, client_id, "wisdom_index_chart_for_client")

def get_client_profile():
    """
    Get client profile information for the authenticated user.
    
    Returns:
        dict: Client profile data or None if not available
    """
    client_id = get_jwt_identity()
    sql_query = """
        SELECT
            c.client_id,
            c.first_name,
            c.last_name,
            c.hh_date_of_birth,
            c.gender,
            c.marital_status,
            c.citizenship,
            c.spouse_first_name,
            c.spouse_last_name,
            c.spouse_dob,
            c.spouse_cell_phone,
            c.address1,
            c.city,
            c.state_or_province,
            c.postal_code,
            c.home_phone,
            c.business_phone,
            c.cell_phone,
            c.emp_name,
            c.emp_job_title,
            c.emp_years_employed,
            u.email
        FROM core.clients c
        LEFT JOIN core.users u ON c.client_id = u.client_id
        WHERE c.client_id = %s;
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id,))
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            # Format date fields
            def format_date(date_str):
                if date_str:
                    try:
                        from datetime import datetime
                        return datetime.strptime(date_str, '%Y-%m-%d').strftime('%B %d, %Y')
                    except:
                        return date_str
                return None
            
            return {
                'client_id': result[0],
                'first_name': result[1],
                'last_name': result[2],
                'full_name': f"{result[1]} {result[2]}" if result[1] and result[2] else None,
                'date_of_birth': format_date(result[3]),
                'gender': result[4],
                'marital_status': result[5],
                'citizenship': result[6],
                'spouse_first_name': result[7],
                'spouse_last_name': result[8],
                'spouse_full_name': f"{result[7]} {result[8]}" if result[7] and result[8] else None,
                'spouse_date_of_birth': format_date(result[9]),
                'spouse_cell_phone': result[10],
                'address1': result[11],
                'city': result[12],
                'state': result[13],
                'postal_code': result[14],
                'home_phone': result[15],
                'business_phone': result[16],
                'cell_phone': result[17],
                'employer_name': result[18],
                'job_title': result[19],
                'years_employed': result[20],
                'email': result[21]
            }
        return None
    except Exception as e:
        print(f"Error getting client profile: {e}")
        return None
    finally:
        if connection:
            close_db_connection(connection)

def update_client_profile(profile_data):
    """
    Update client profile information for the authenticated user.
    
    Args:
        profile_data (dict): Dictionary containing profile fields to update
        
    Returns:
        dict: Updated profile data or None if failed
    """
    client_id = get_jwt_identity()
    
    # Build dynamic UPDATE query based on provided fields
    update_fields = []
    params = []
    
    # Map frontend field names to database column names
    field_mapping = {
        'first_name': 'first_name',
        'last_name': 'last_name',
        'date_of_birth': 'hh_date_of_birth',
        'gender': 'gender',
        'marital_status': 'marital_status',
        'citizenship': 'citizenship',
        'spouse_first_name': 'spouse_first_name',
        'spouse_last_name': 'spouse_last_name',
        'spouse_date_of_birth': 'spouse_dob',
        'spouse_cell_phone': 'spouse_cell_phone',
        'address1': 'address1',
        'city': 'city',
        'state': 'state_or_province',
        'postal_code': 'postal_code',
        'home_phone': 'home_phone',
        'business_phone': 'business_phone',
        'cell_phone': 'cell_phone',
        'employer_name': 'emp_name',
        'job_title': 'emp_job_title',
        'years_employed': 'emp_years_employed'
    }
    
    # Build UPDATE query dynamically
    for field_name, db_column in field_mapping.items():
        if field_name in profile_data:
            # Handle date fields specially to avoid setting NULL when empty string
            if field_name in ['date_of_birth', 'spouse_date_of_birth']:
                # Only update date fields if they have a non-empty value
                if profile_data[field_name] and profile_data[field_name].strip():
                    update_fields.append(f"{db_column} = %s")
                    from datetime import datetime
                    try:
                        # Convert string date to PostgreSQL date format
                        date_obj = datetime.strptime(profile_data[field_name], '%Y-%m-%d')
                        params.append(date_obj.strftime('%Y-%m-%d'))
                    except (ValueError, TypeError):
                        # If date parsing fails, set to None
                        params.append(None)
                # If date field is empty, don't include it in the UPDATE query
                # This preserves existing date values in the database
            elif profile_data[field_name] is not None:
                # For non-date fields, update if not None
                update_fields.append(f"{db_column} = %s")
                params.append(profile_data[field_name])
            # For fields that are None, don't include them in the UPDATE query
            # This preserves existing values in the database
    
    if not update_fields:
        print("No valid fields to update")
        return None
    
    # Add client_id to params
    params.append(client_id)
    
    sql_query = f"""
        UPDATE core.clients
        SET {', '.join(update_fields)}
        WHERE client_id = %s
        RETURNING
            client_id, first_name, last_name, hh_date_of_birth, gender,
            marital_status, citizenship, spouse_first_name, spouse_last_name,
            spouse_dob, spouse_cell_phone, address1, city, state_or_province, postal_code,
            home_phone, business_phone, cell_phone, emp_name, emp_job_title,
            emp_years_employed
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, params)
        connection.commit()
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            # Get email from users table
            email_query = "SELECT email FROM core.users WHERE client_id = %s"
            cursor = connection.cursor()
            cursor.execute(email_query, (client_id,))
            email_result = cursor.fetchone()
            cursor.close()
            
            # Format date fields
            def format_date(date_str):
                if date_str:
                    try:
                        from datetime import datetime
                        return datetime.strptime(date_str, '%Y-%m-%d').strftime('%B %d, %Y')
                    except:
                        return date_str
                return None
            
            return {
                'client_id': result[0],
                'first_name': result[1],
                'last_name': result[2],
                'full_name': f"{result[1]} {result[2]}" if result[1] and result[2] else None,
                'date_of_birth': format_date(result[3]),
                'gender': result[4],
                'marital_status': result[5],
                'citizenship': result[6],
                'spouse_first_name': result[7],
                'spouse_last_name': result[8],
                'spouse_full_name': f"{result[7]} {result[8]}" if result[7] and result[8] else None,
                'spouse_date_of_birth': format_date(result[9]),
                'spouse_cell_phone': result[10],
                'address1': result[11],
                'city': result[12],
                'state': result[13],
                'postal_code': result[14],
                'home_phone': result[15],
                'business_phone': result[16],
                'cell_phone': result[17],
                'employer_name': result[18],
                'job_title': result[19],
                'years_employed': result[20],
                'email': email_result[0] if email_result else None
            }
        return None
        
    except Exception as e:
        print(f"Error updating client profile: {e}")
        if connection:
            connection.rollback()
        return None
    finally:
        if connection:
            close_db_connection(connection)

def get_all_user_metrics_for_export(client_id):
    """
    Aggregate all financial metrics for export.
    Uses the single-query helper to avoid 30+ sequential metric queries.
    """
    return get_all_metrics_for_client(client_id)

def get_chart_data_for_export(client_id):
    """
    Get chart data for export - both income and expense breakdowns.
    """
    return {
        'income': get_income_chart_data_for_user(),
        'expense': get_expense_chart_data_for_user()
    }

def create_metrics_sheet(sheet, metrics_data):
    """
    Create the Financial Metrics sheet with proper formatting and organization.
    """
    from openpyxl.styles import Font, Alignment
    from datetime import datetime
    
    # Headers
    sheet['A1'] = 'Wisdom Index Financial Export'
    sheet['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    # Category headers with styling
    row = 4
    for category, data in metrics_data.items():
        sheet.cell(row=row, column=1, value=category.replace('_', ' ').title())
        sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        
        # Metric names and values
        # Add bold header row for each category section
        row += 1
        metric_header_cell = sheet.cell(row=row, column=1, value="Metric")
        value_header_cell = sheet.cell(row=row, column=2, value="Value")
        metric_header_cell.font = Font(bold=True)
        value_header_cell.font = Font(bold=True)
        value_header_cell.alignment = Alignment(horizontal="right")

        for metric_name, metric_value in data.items():
            row += 1
            name_cell = sheet.cell(row=row, column=1, value=metric_name.replace('_', ' ').title())
            value_cell = sheet.cell(row=row, column=2, value=metric_value)
            value_cell.number_format = '#,##0.00'
            value_cell.alignment = Alignment(horizontal="right")

        row += 2  # Add spacing between categories

    # Auto-adjust column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20

def create_chart_sheet(sheet, chart_data, title):
    """
    Create a chart data sheet with categories and values.
    """
    from openpyxl.styles import Font, Alignment
    
    sheet['A1'] = title
    sheet['A1'].font = Font(bold=True, size=14)

    sheet['A2'] = 'Category'
    sheet['B2'] = 'Amount'
    sheet['A2'].font = sheet['B2'].font = Font(bold=True)
    
    # Right-align the "Amount" header
    sheet['B2'].alignment = Alignment(horizontal='right')

    # Data rows
    row = 3
    for item in chart_data:
        sheet.cell(row=row, column=1, value=item.get('income_category') or item.get('expense_category') or item.get('category'))
        amount_cell = sheet.cell(row=row, column=2, value=float(item.get('amount', 0)))
        amount_cell.number_format = '#,##0.00'
        # Right-align the amount values
        amount_cell.alignment = Alignment(horizontal='right')
        row += 1

    # Auto-adjust column widths
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 15

# Metric Targets Management Functions

def get_metric_target_for_user(metric_name):
    """
    Get the most recent target value for a specific metric for the authenticated user.
    
    Args:
        metric_name (str): The name of the metric
        
    Returns:
        float: Target value or None if not set
    """
    client_id = get_jwt_identity()
    
    sql_query = """
        SELECT target_value
        FROM core.metric_targets
        WHERE client_id = %s AND metric_name = %s
        ORDER BY created_at DESC
        LIMIT 1
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id, metric_name))
        result = cursor.fetchone()
        cursor.close()
        
        if result and result[0] is not None:
            return float(result[0])
        return None
        
    except Exception as e:
        print(f"Error getting target for {metric_name}: {e}")
        return None
    finally:
        if connection:
            close_db_connection(connection)

def get_all_targets_for_user():
    """
    Get all the most recent target values for the authenticated user.
    
    Returns:
        dict: Dictionary mapping metric names to target values
    """
    client_id = get_jwt_identity()
    
    sql_query = """
        SELECT DISTINCT ON (metric_name) metric_name, target_value
        FROM core.metric_targets
        WHERE client_id = %s
        ORDER BY metric_name, created_at DESC
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id,))
        results = cursor.fetchall()
        cursor.close()
        
        targets = {}
        for row in results:
            if row[1] is not None:
                targets[row[0]] = float(row[1])
        
        return targets
        
    except Exception as e:
        print(f"Error getting all targets: {e}")
        return {}
    finally:
        if connection:
            close_db_connection(connection)

def update_metric_target_for_user(metric_name, target_value):
    """
    Create a new target value for a specific metric for the authenticated user.
    This function always inserts a new row to preserve target history.
    
    Args:
        metric_name (str): The name of the metric
        target_value (float): The target value
        
    Returns:
        bool: True if successful, False otherwise
    """
    client_id = get_jwt_identity()
    
    sql_query = """
        INSERT INTO core.metric_targets (client_id, metric_name, target_value, created_at)
        VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id, metric_name, target_value))
        connection.commit()
        cursor.close()
        
        return True
        
    except Exception as e:
        print(f"Error updating target for {metric_name}: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if connection:
            close_db_connection(connection)

def update_multiple_targets_for_user(targets_dict):
    """
    Create multiple new target values in a single transaction.
    This function only inserts targets that have changed from their current values.
    
    Args:
        targets_dict (dict): Dictionary mapping metric names to target values
        
    Returns:
        bool: True if all successful, False otherwise
    """
    client_id = get_jwt_identity()
    
    # First, get current targets to compare with new values
    current_targets = get_all_targets_for_user()
    
    # Only insert targets that have changed or are new
    changed_targets = {}
    for metric_name, target_value in targets_dict.items():
        current_value = current_targets.get(metric_name)
        
        # Insert if it's a new target or if the value has changed
        if current_value is None or current_value != target_value:
            changed_targets[metric_name] = target_value
    
    # If no targets have changed, return success
    if not changed_targets:
        print("No targets have changed, skipping database update")
        return True
    
    sql_query = """
        INSERT INTO core.metric_targets (client_id, metric_name, target_value, created_at)
        VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Execute each target update individually using execute_batch
        from psycopg2.extras import execute_batch
        values = [
            (client_id, metric_name, target_value)
            for metric_name, target_value in changed_targets.items()
        ]
        
        print(f"Inserting {len(values)} changed targets: {[v[1] for v in values]}")
        execute_batch(cursor, sql_query, values)
        connection.commit()
        cursor.close()
        
        return True
        
    except Exception as e:
        print(f"Error updating multiple targets: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if connection:
            close_db_connection(connection)

def delete_metric_target_for_user(metric_name):
    """
    Delete the most recent target value for a specific metric for the authenticated user.
    
    Args:
        metric_name (str): The name of the metric
        
    Returns:
        bool: True if successful, False otherwise
    """
    client_id = get_jwt_identity()
    
    sql_query = """
        DELETE FROM core.metric_targets
        WHERE client_id = %s AND metric_name = %s
        AND created_at = (
            SELECT MAX(created_at)
            FROM core.metric_targets
            WHERE client_id = %s AND metric_name = %s
        )
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id, metric_name, client_id, metric_name))
        connection.commit()
        rows_affected = cursor.rowcount
        cursor.close()
        
        print(f"Deleted {rows_affected} target(s) for {metric_name}")
        return rows_affected > 0
        
    except Exception as e:
        print(f"Error deleting target for {metric_name}: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if connection:
            close_db_connection(connection)

def delete_all_targets_for_user():
    """
    Delete all target values for the authenticated user.
    
    Returns:
        bool: True if successful, False otherwise
    """
    client_id = get_jwt_identity()
    
    sql_query = """
        DELETE FROM core.metric_targets
        WHERE client_id = %s
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id,))
        connection.commit()
        cursor.close()
        
        return True
        
    except Exception as e:
        print(f"Error deleting all targets: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if connection:
            close_db_connection(connection)

def compare_values(actual_value, target_value):
    """
    Compare actual value with target value and return status.
    
    Args:
        actual_value (float): The actual metric value
        target_value (float): The target value
        
    Returns:
        str: 'above', 'below', 'equal', or 'no_target'
    """
    if target_value is None or actual_value is None:
        return 'no_target'
    
    if actual_value > target_value:
        return 'above'
    elif actual_value < target_value:
        return 'below'
    else:
        return 'equal'

def calculate_target_percentage(actual_value, target_value):
    """
    Calculate the percentage difference between actual and target values.
    
    Args:
        actual_value (float): The actual metric value
        target_value (float): The target value
        
    Returns:
        dict: Dictionary containing percentage, status, and formatted display text
    """
    if target_value is None or actual_value is None:
        return {
            'percentage': 0,
            'status': 'no_target',
            'display_text': 'No target'
        }
    
    # Calculate percentage difference
    percentage = ((actual_value / target_value - 1) * 100)
    
    # Determine status
    status = compare_values(actual_value, target_value)
    
    # Format display text
    if status == 'above':
        display_text = f"↑ {abs(percentage):.1f}%"
    elif status == 'below':
        display_text = f"↓ {abs(percentage):.1f}%"
    else:  # equal
        display_text = "✓"
    
    return {
        'percentage': percentage,
        'status': status,
        'display_text': display_text
    }

def get_metric_with_target(metric_name, actual_value):
    """
    Get metric data including target information for API responses.
    
    Args:
        metric_name (str): The name of the metric
        actual_value (float): The actual metric value
        
    Returns:
        dict: Dictionary containing metric value, target, and comparison data
    """
    # Get target value
    target_value = get_metric_target_for_user(metric_name)
    
    # Calculate target comparison
    target_comparison = calculate_target_percentage(actual_value, target_value)
    
    return {
        'value': actual_value,
        'target': target_value,
        'target_status': target_comparison['status'],
        'target_percentage': target_comparison['percentage'],
        'target_display_text': target_comparison['display_text']
    }

def get_client_name_for_user():
    """
    Get the client name for the authenticated user.
    
    Returns:
        str: The client name or None if not found
    """
    client_id = get_jwt_identity()
    return get_client_name_by_id(client_id)

def get_client_name_by_id(client_id):
    """
    Get the client name for a specific client ID.
    
    Args:
        client_id (int): The client ID
        
    Returns:
        str: The client name or None if not found
    """
    sql_query = """
        SELECT client_name
        FROM core.clients
        WHERE client_id = %s;
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id,))
        result = cursor.fetchone()
        cursor.close()
        
        if result and result[0]:
            return result[0]
        return None
        
    except Exception as e:
        print(f"Error getting client name for id {client_id}: {e}")
        return None
    finally:
        if connection:
            close_db_connection(connection)

# Metric Detail Modal Functions

def get_metric_details(metric_name):
    """
    Get metric details including formula, description, and tables used.
    
    Args:
        metric_name (str): The name of the metric (e.g., 'net-worth')
        
    Returns:
        dict: Metric details or None if not found
    """
    # Map metric functions to their details
    metric_details_map = {
        'net-worth': {
            'title': 'Net Worth',
            'category': 'assets',
            'formula': 'Total Assets - Total Liabilities',
            'description': 'Net worth represents the difference between what you own (assets) and what you owe (liabilities). It\'s a measure of your overall financial health.',
            'tables': ['holdings', 'real_estate_assets', 'liability_note_accounts', 'investment_deposit_accounts', 'personal_property_accounts', 'businesses']
        },
        'portfolio-value': {
            'title': 'Portfolio Value',
            'category': 'assets',
            'formula': 'Sum of Investment Holdings + Investment Deposit Accounts',
            'description': 'Portfolio value represents the total value of your investment accounts including holdings and deposit accounts.',
            'tables': ['holdings', 'investment_deposit_accounts']
        },
        'real-estate-value': {
            'title': 'Real Estate Value',
            'category': 'assets',
            'formula': 'Sum of Real Estate Assets',
            'description': 'Real estate value represents the total value of all property investments you own.',
            'tables': ['real_estate_assets']
        },
        'debt': {
            'title': 'Debt',
            'category': 'assets',
            'formula': 'Sum of All Liabilities',
            'description': 'Debt represents the total amount of money you owe across all loans and liabilities.',
            'tables': ['liability_note_accounts']
        },
        'equity': {
            'title': 'Equity',
            'category': 'assets',
            'formula': 'Equity Holdings + Investment Equity',
            'description': 'Equity represents your ownership in stocks and equity-based investments.',
            'tables': ['holdings', 'investment_deposit_accounts']
        },
        'fixed-income': {
            'title': 'Fixed Income',
            'category': 'assets',
            'formula': 'Sum of Fixed Income Holdings',
            'description': 'Fixed income represents investments that pay a fixed return, such as bonds.',
            'tables': ['holdings']
        },
        'cash': {
            'title': 'Cash',
            'category': 'assets',
            'formula': 'Cash from Holdings + Cash from Investments',
            'description': 'Cash represents your liquid assets and cash equivalents.',
            'tables': ['holdings', 'investment_deposit_accounts']
        },
        'earned-income': {
            'title': 'Earned Income',
            'category': 'income',
            'formula': 'Sum of Salary Income',
            'description': 'Earned income represents income from employment and active work.',
            'tables': ['incomes']
        },
        'social-security-income': {
            'title': 'Social Security Income',
            'category': 'income',
            'formula': 'Sum of Social Security Benefits',
            'description': 'Social security income represents government benefits received.',
            'tables': ['incomes']
        },
        'pension-income': {
            'title': 'Pension Income',
            'category': 'income',
            'formula': 'Sum of Pension Benefits',
            'description': 'Pension income represents retirement benefits from former employers.',
            'tables': ['incomes']
        },
        'real-estate-income': {
            'title': 'Real Estate Income',
            'category': 'income',
            'formula': 'Sum of Rental and Property Income',
            'description': 'Real estate income represents income generated from property investments.',
            'tables': ['incomes']
        },
        'business-income': {
            'title': 'Business Income',
            'category': 'income',
            'formula': 'Sum of Business Income',
            'description': 'Business income represents income from self-employment and business ownership.',
            'tables': ['incomes']
        },
        'total-income': {
            'title': 'Total Income',
            'category': 'income',
            'formula': 'Sum of All Income Sources',
            'description': 'Total income represents all income sources combined.',
            'tables': ['incomes']
        },
        'current-year-giving': {
            'title': 'Current Year Giving',
            'category': 'expenses',
            'formula': 'Sum of Philanthropic Giving',
            'description': 'Current year giving represents donations and charitable contributions.',
            'tables': ['expenses']
        },
        'current-year-savings': {
            'title': 'Current Year Savings',
            'category': 'expenses',
            'formula': 'Sum of Active Savings Plans',
            'description': 'Current year savings represents contributions to savings and investment accounts.',
            'tables': ['savings']
        },
        'current-year-debt': {
            'title': 'Current Year Debt',
            'category': 'expenses',
            'formula': 'Sum of Annual Debt Payments',
            'description': 'Current year debt represents annual payments on outstanding loans.',
            'tables': ['liability_note_accounts']
        },
        'current-year-taxes': {
            'title': 'Current Year Taxes',
            'category': 'expenses',
            'formula': 'Estimated Tax Payments (15% of Income)',
            'description': 'Current year taxes represents estimated tax obligations.',
            'tables': ['incomes']
        },
        'current-year-living-expenses': {
            'title': 'Current Year Living Expenses',
            'category': 'expenses',
            'formula': 'Sum of Living Expenses',
            'description': 'Current year living expenses represents day-to-day living costs.',
            'tables': ['expenses']
        },
        'total-expenses': {
            'title': 'Total Expenses',
            'category': 'expenses',
            'formula': 'Sum of All Expense Categories',
            'description': 'Total expenses represents all expenses combined.',
            'tables': ['expenses', 'savings', 'liability_note_accounts', 'incomes']
        },
        'margin': {
            'title': 'Margin',
            'category': 'expenses',
            'formula': 'Total Income - Total Expenses',
            'description': 'Margin represents the difference between income and expenses.',
            'tables': ['incomes', 'expenses', 'savings', 'liability_note_accounts']
        },
        'life-insurance': {
            'title': 'Life Insurance',
            'category': 'insurance',
            'formula': 'Sum of Life Insurance Death Benefits',
            'description': 'Life insurance represents the total death benefit coverage.',
            'tables': ['life_insurance_annuity_accounts']
        },
        'disability': {
            'title': 'Disability',
            'category': 'insurance',
            'formula': 'Sum of Disability Benefits',
            'description': 'Disability represents disability insurance coverage.',
            'tables': ['disability_ltc_insurance_accounts']
        },
        'ltc': {
            'title': 'LTC',
            'category': 'insurance',
            'formula': 'Sum of Long-Term Care Benefits',
            'description': 'LTC represents long-term care insurance coverage.',
            'tables': ['disability_ltc_insurance_accounts']
        },
        'umbrella': {
            'title': 'Umbrella',
            'category': 'insurance',
            'formula': 'Sum of Umbrella Insurance Coverage',
            'description': 'Umbrella represents excess liability insurance coverage.',
            'tables': ['property_casualty_insurance_accounts']
        },
        'business-insurance': {
            'title': 'Business Insurance',
            'category': 'insurance',
            'formula': 'Sum of Business Insurance Benefits',
            'description': 'Business insurance represents business-related insurance coverage.',
            'tables': ['disability_ltc_insurance_accounts']
        },
        'flood-insurance': {
            'title': 'Flood Insurance',
            'category': 'insurance',
            'formula': 'Sum of Flood Insurance Coverage',
            'description': 'Flood insurance represents flood damage insurance coverage.',
            'tables': ['property_casualty_insurance_accounts']
        },
        'at-risk': {
            'title': 'At Risk',
            'category': 'insurance',
            'formula': 'Taxable Investments - Umbrella Coverage',
            'description': 'At risk represents exposure to liability beyond insurance coverage.',
            'tables': ['investment_deposit_accounts', 'property_casualty_insurance_accounts']
        },
        'retirement-ratio': {
            'title': 'Retirement Ratio',
            'category': 'planning',
            'formula': '(Future Income PV + Current Assets + Retirement Savings) / (Future Expenses PV + Current Liabilities)',
            'description': 'Retirement ratio measures your readiness for retirement based on projected income and expenses.',
            'tables': ['incomes', 'expenses', 'investment_deposit_accounts', 'real_estate_assets', 'personal_property_accounts', 'savings', 'liability_note_accounts']
        },
        'survivor-ratio': {
            'title': 'Survivor Ratio',
            'category': 'planning',
            'formula': '(Life Insurance + Future Income PV + Current Assets) / (Future Expenses PV + Current Liabilities)',
            'description': 'Survivor ratio measures financial protection for surviving family members.',
            'tables': ['incomes', 'expenses', 'investment_deposit_accounts', 'real_estate_assets', 'personal_property_accounts', 'life_insurance_annuity_accounts', 'liability_note_accounts']
        },
        'education-ratio': {
            'title': 'Education Ratio',
            'category': 'planning',
            'formula': '(Education Savings PV + Education Account Balances) / Education Expenses PV',
            'description': 'Education ratio measures preparation for education expenses.',
            'tables': ['savings', 'investment_deposit_accounts', 'personal_property_accounts', 'expenses']
        },
        'new-cars-ratio': {
            'title': 'New Cars Ratio',
            'category': 'planning',
            'formula': '(Taxable Accounts + Taxable Savings PV) / Car Expenses PV',
            'description': 'New cars ratio measures preparation for vehicle purchases.',
            'tables': ['investment_deposit_accounts', 'savings', 'expenses']
        },
        'ltc-ratio': {
            'title': 'LTC Ratio',
            'category': 'planning',
            'formula': '(Future Income PV + Current Assets) / (Future Expenses PV + LTC Expenses PV)',
            'description': 'LTC ratio measures preparation for long-term care expenses.',
            'tables': ['incomes', 'investment_deposit_accounts', 'real_estate_assets', 'personal_property_accounts', 'expenses', 'disability_ltc_insurance_accounts']
        },
        'ltd-ratio': {
            'title': 'LTD Ratio',
            'category': 'planning',
            'formula': 'LTD Value / Earned Income',
            'description': 'LTD ratio measures disability insurance coverage relative to income.',
            'tables': ['disability_ltc_insurance_accounts', 'incomes']
        },
        'savings-ratio': {
            'title': 'Savings Ratio',
            'category': 'wisdom-index',
            'formula': 'Current Year Savings / Total Income',
            'description': 'Savings ratio measures the proportion of income directed to active savings for the current year.',
            'tables': ['savings', 'incomes']
        },
        'giving-ratio': {
            'title': 'Giving Ratio',
            'category': 'wisdom-index',
            'formula': 'Current Year Giving / Total Income',
            'description': 'Giving ratio measures charitable giving as a share of total income for the current year.',
            'tables': ['expenses', 'incomes']
        },
        'reserves-ratio': {
            'title': 'Reserves Ratio',
            'category': 'wisdom-index',
            'formula': '(Cash Holdings + Cash in Investments) / Living Expenses * 0.5',
            'description': 'Reserves ratio measures short-term cash coverage of living expenses (half-year target).',
            'tables': ['holdings', 'investment_deposit_accounts', 'expenses']
        },
        'debt-ratio': {
            'title': 'Debt Ratio',
            'category': 'wisdom-index',
            'formula': 'Home Equity / Home Value',
            'description': 'Debt ratio measures mortgage leverage on the primary residence based on equity to value.',
            'tables': ['real_estate_assets', 'liability_note_accounts']
        },
        'diversification-ratio': {
            'title': 'Diversification Ratio',
            'category': 'wisdom-index',
            'formula': '1 - (Largest Holding / Total Portfolio)',
            'description': 'Diversification ratio measures concentration risk by comparing the largest position to the total portfolio.',
            'tables': ['holdings', 'investment_deposit_accounts']
        }
    }
    
    return metric_details_map.get(metric_name)

def get_table_data_for_user(table_name, page=1, limit=50, metric_name=None):
    """
    Get raw table data for authenticated user with pagination.
    
    Args:
        table_name (str): The name of the table
        page (int): Page number for pagination
        limit (int): Number of records per page
        metric_name (str): Optional metric name for filtering
        
    Returns:
        dict: Table data with pagination info
    """
    client_id = get_jwt_identity()
    
    # Whitelist of allowed tables with their configurations
    table_configs = {
        'holdings': {
            'table_name': 'core.holdings',
            'display_name': 'Investment Holdings',
            'columns': [
                {'name': 'ticker', 'type': 'text', 'display_name': 'Ticker'},
                {'name': 'description', 'type': 'text', 'display_name': 'Description'},
                {'name': 'asset_class', 'type': 'text', 'display_name': 'Asset Class'},
                {'name': 'value', 'type': 'currency', 'display_name': 'Value'},
                {'name': 'units', 'type': 'number', 'display_name': 'Units'}
            ],
            'order_by': 'asset_class, ticker'
        },
        'incomes': {
            'table_name': 'core.incomes',
            'display_name': 'Income Streams',
            'columns': [
                {'name': 'income_name', 'type': 'text', 'display_name': 'Income Name'},
                {'name': 'income_type', 'type': 'text', 'display_name': 'Income Type'},
                {'name': 'current_year_amount', 'type': 'currency', 'display_name': 'Current Year Amount'},
                {'name': 'annual_amount', 'type': 'currency', 'display_name': 'Annual Amount'},
                {'name': 'owner_type', 'type': 'text', 'display_name': 'Owner'}
            ],
            'order_by': 'income_type, income_name'
        },
        'expenses': {
            'table_name': 'core.expenses',
            'display_name': 'Expenses',
            'columns': [
                {'name': 'type', 'type': 'text', 'display_name': 'Type'},
                {'name': 'sub_type', 'type': 'text', 'display_name': 'Sub Type'},
                {'name': 'expense_item', 'type': 'text', 'display_name': 'Expense Item'},
                {'name': 'annual_amount', 'type': 'currency', 'display_name': 'Annual Amount'}
            ],
            'order_by': 'type, sub_type'
        },
        'real_estate_assets': {
            'table_name': 'core.real_estate_assets',
            'display_name': 'Real Estate Assets',
            'columns': [
                {'name': 'account_name', 'type': 'text', 'display_name': 'Property Name'},
                {'name': 'sub_type', 'type': 'text', 'display_name': 'Property Type'},
                {'name': 'total_value', 'type': 'currency', 'display_name': 'Total Value'},
                {'name': 'address1', 'type': 'text', 'display_name': 'Address'},
                {'name': 'city', 'type': 'text', 'display_name': 'City'},
                {'name': 'state', 'type': 'text', 'display_name': 'State'}
            ],
            'order_by': 'account_name'
        },
        'liability_note_accounts': {
            'table_name': 'core.liability_note_accounts',
            'display_name': 'Loans & Debts',
            'columns': [
                {'name': 'account_name', 'type': 'text', 'display_name': 'Account Name'},
                {'name': 'total_value', 'type': 'currency', 'display_name': 'Balance'},
                {'name': 'interest_rate', 'type': 'percentage', 'display_name': 'Interest Rate'},
                {'name': 'repayment_type', 'type': 'text', 'display_name': 'Payment Type'}
            ],
            'order_by': 'account_name'
        },
        'investment_deposit_accounts': {
            'table_name': 'core.investment_deposit_accounts',
            'display_name': 'Investment Accounts',
            'columns': [
                {'name': 'fact_type_name', 'type': 'text', 'display_name': 'Account Type'},
                {'name': 'total_value', 'type': 'currency', 'display_name': 'Total Value'},
                {'name': 'holdings_value', 'type': 'currency', 'display_name': 'Holdings Value'},
                {'name': 'cash_balance', 'type': 'currency', 'display_name': 'Cash Balance'}
            ],
            'order_by': 'fact_type_name'
        },
        'life_insurance_annuity_accounts': {
            'table_name': 'core.life_insurance_annuity_accounts',
            'display_name': 'Life Insurance',
            'columns': [
                {'name': 'fact_type_name', 'type': 'text', 'display_name': 'Policy Type'},
                {'name': 'death_benefit', 'type': 'currency', 'display_name': 'Death Benefit'},
                {'name': 'account_name', 'type': 'text', 'display_name': 'Policy Name'}
            ],
            'order_by': 'fact_type_name'
        },
        'disability_ltc_insurance_accounts': {
            'table_name': 'core.disability_ltc_insurance_accounts',
            'display_name': 'Disability & LTC Insurance',
            'columns': [
                {'name': 'fact_type_name', 'type': 'text', 'display_name': 'Policy Type'},
                {'name': 'sub_type', 'type': 'text', 'display_name': 'Sub Type'},
                {'name': 'benefit_amount', 'type': 'currency', 'display_name': 'Benefit Amount'},
                {'name': 'annual_premium', 'type': 'currency', 'display_name': 'Annual Premium'}
            ],
            'order_by': 'fact_type_name'
        },
        'property_casualty_insurance_accounts': {
            'table_name': 'core.property_casualty_insurance_accounts',
            'display_name': 'Property Insurance',
            'columns': [
                {'name': 'sub_type', 'type': 'text', 'display_name': 'Insurance Type'},
                {'name': 'maximum_annual_benefit', 'type': 'currency', 'display_name': 'Max Annual Benefit'},
                {'name': 'account_name', 'type': 'text', 'display_name': 'Policy Name'}
            ],
            'order_by': 'sub_type'
        },
        'savings': {
            'table_name': 'core.savings',
            'display_name': 'Savings Plans',
            'columns': [
                {'name': 'destination', 'type': 'text', 'display_name': 'Destination'},
                {'name': 'calculated_annual_amount_usd', 'type': 'currency', 'display_name': 'Annual Amount'},
                {'name': 'start_type', 'type': 'text', 'display_name': 'Status'},
                {'name': 'account_id', 'type': 'text', 'display_name': 'Account ID'}
            ],
            'order_by': 'destination'
        },
        'businesses': {
            'table_name': 'core.businesses',
            'display_name': 'Business Assets',
            'columns': [
                {'name': 'name', 'type': 'text', 'display_name': 'Business Name'},
                {'name': 'fact_type_name', 'type': 'text', 'display_name': 'Type'},
                {'name': 'sub_type', 'type': 'text', 'display_name': 'Sub Type'},
                {'name': 'amount', 'type': 'currency', 'display_name': 'Amount'}
            ],
            'order_by': 'name'
        },
        'personal_property_accounts': {
            'table_name': 'core.personal_property_accounts',
            'display_name': 'Personal Property',
            'columns': [
                {'name': 'account_name', 'type': 'text', 'display_name': 'Property Name'},
                {'name': 'fact_type_name', 'type': 'text', 'display_name': 'Type'},
                {'name': 'total_value', 'type': 'currency', 'display_name': 'Total Value'}
            ],
            'order_by': 'account_name'
        }
    }
    
    if table_name not in table_configs:
        raise ValueError(f"Table '{table_name}' not allowed")
    
    config = table_configs[table_name]
    offset = (page - 1) * limit
    
    # Get total count for pagination
    count_query = f"""
        SELECT COUNT(*) FROM {config['table_name']}
        WHERE client_id = %s
    """
    
    # Get data with pagination
    data_query = f"""
        SELECT {', '.join([col['name'] for col in config['columns']])}
        FROM {config['table_name']}
        WHERE client_id = %s
        ORDER BY {config['order_by']}
        LIMIT %s OFFSET %s
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Get total count
        cursor.execute(count_query, (client_id,))
        total_count = cursor.fetchone()[0]
        
        # Get data
        cursor.execute(data_query, (client_id, limit, offset))
        results = cursor.fetchall()
        cursor.close()
        
        # Format results
        data = []
        for row in results:
            row_dict = {}
            for i, value in enumerate(row):
                col_config = config['columns'][i]
                formatted_value = value
                
                # Format based on column type
                if value is not None:
                    if col_config['type'] == 'currency' and isinstance(value, (int, float)):
                        formatted_value = float(value)
                    elif col_config['type'] == 'percentage' and isinstance(value, (int, float)):
                        formatted_value = float(value)
                
                row_dict[col_config['name']] = formatted_value
            data.append(row_dict)
        
        total_pages = (total_count + limit - 1) // limit
        
        return {
            'table_name': table_name,
            'display_name': config['display_name'],
            'columns': config['columns'],
            'data': data,
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total_count,
                'total_pages': total_pages
            }
        }
        
    except Exception as e:
        print(f"Error getting table data for {table_name}: {e}")
        return None
    finally:
        if connection:
            close_db_connection(connection)

# Admin-specific functions for client management

def get_all_metrics_for_client(client_id):
    """
    Get all 37+ metrics for a specific client in a SINGLE optimized query (admin function).
    Reduces 37 sequential queries to just 1 comprehensive query.
    
    Args:
        client_id (int): The client ID to get metrics for
        
    Returns:
        dict: All metrics organized by category
    """
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Single comprehensive query that calculates ALL 37 metrics at once
        sql_query = """
        WITH
        -- Client info for age-based calculations
        client_info AS (
            SELECT
                client_id,
                EXTRACT(YEAR FROM AGE(CURRENT_DATE, hh_date_of_birth)) AS current_age,
                65 AS retirement_age
            FROM core.clients
            WHERE client_id = %s
        ),
        
        -- ASSETS & LIABILITIES (7 metrics)
        holdings_agg AS (
            SELECT
                SUM(value) AS total_holdings,
                SUM(CASE WHEN asset_class IN ('largecap', 'smallcap', 'largevalue', 'smallvalue', 'internat', 'emerging', 'ips') THEN value ELSE 0 END) AS equity_holdings,
                SUM(CASE WHEN asset_class IN ('highyldbond', 'inttermmun', 'investbond', 'shortermbond', 'shortermmun') THEN value ELSE 0 END) AS fixed_income,
                SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) AS cash_holdings
            FROM core.holdings
            WHERE client_id = %s
        ),
        investment_accounts_agg AS (
            SELECT
                SUM(total_value) AS total_investments,
                SUM(CASE WHEN fact_type_name IN ('Taxable Investment', 'Roth IRA', 'Qualified Retirement') THEN holdings_value ELSE 0 END) AS equity_investments,
                SUM(CASE WHEN fact_type_name = 'Cash Alternative' THEN cash_balance ELSE 0 END) AS cash_investments,
                SUM(CASE WHEN fact_type_name = 'Taxable Investment' THEN total_value ELSE 0 END) AS taxable_investments
            FROM core.investment_deposit_accounts
            WHERE client_id = %s
        ),
        real_estate_agg AS (
            SELECT SUM(total_value) AS total_real_estate
            FROM core.real_estate_assets
            WHERE client_id = %s
        ),
        businesses_agg AS (
            SELECT SUM(amount) AS total_businesses
            FROM core.businesses
            WHERE client_id = %s
        ),
        personal_property_agg AS (
            SELECT SUM(total_value) AS total_personal_property
            FROM core.personal_property_accounts
            WHERE client_id = %s
        ),
        liabilities_agg AS (
            SELECT SUM(ABS(total_value)) AS total_liabilities
            FROM core.liability_note_accounts
            WHERE client_id = %s
        ),
        
        -- INCOME ANALYSIS (6 metrics)
        income_breakdown AS (
            SELECT
                SUM(CASE WHEN income_type = 'Salary' THEN current_year_amount ELSE 0 END) AS earned_income,
                SUM(CASE WHEN income_type = 'SocialSecurity' THEN current_year_amount ELSE 0 END) AS social_security,
                SUM(CASE WHEN income_type = 'Pension' THEN current_year_amount ELSE 0 END) AS pension,
                SUM(CASE WHEN income_type = 'Real Estate' THEN current_year_amount ELSE 0 END) AS real_estate_income,
                SUM(CASE WHEN income_type = 'Business' THEN current_year_amount ELSE 0 END) AS business_income,
                SUM(current_year_amount) AS total_income
            FROM core.incomes
            WHERE client_id = %s AND current_year_amount IS NOT NULL
        ),
        
        -- EXPENSE TRACKING (7 metrics)
        giving_expense AS (
            SELECT SUM(annual_amount) AS giving
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Spending' AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        ),
        savings_expense AS (
            SELECT SUM(calculated_annual_amount_usd) AS savings
            FROM core.savings
            WHERE client_id = %s AND start_type = 'Active'
        ),
        debt_payments AS (
            SELECT SUM(
                CASE
                    WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL
                    THEN ABS(total_value) * (interest_rate / 12) / (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                    ELSE ABS(total_value) / 12
                END
            ) AS debt_expense
            FROM core.liability_note_accounts
            WHERE client_id = %s
                AND total_value < 0
                AND repayment_type = 'PrincipalAndInterest'
                AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (loan_term_in_years IS NULL OR EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))
        ),
        living_expense AS (
            SELECT SUM(annual_amount) AS living
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Living' AND annual_amount > 0
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
        ),
        
        -- INSURANCE COVERAGE (7 metrics)
        life_insurance_agg AS (
            SELECT SUM(death_benefit) AS life_insurance
            FROM core.life_insurance_annuity_accounts
            WHERE client_id = %s AND fact_type_name = 'Life Insurance'
        ),
        disability_agg AS (
            SELECT
                SUM(CASE WHEN fact_type_name IN ('Disability Policy', 'Business Disability Policy') THEN benefit_amount ELSE 0 END) AS disability,
                SUM(CASE WHEN sub_type = 'PersonalLT' THEN benefit_amount ELSE 0 END) AS ltc,
                SUM(CASE WHEN sub_type = 'BusinessReducingTerm' THEN benefit_amount ELSE 0 END) AS business_insurance
            FROM core.disability_ltc_insurance_accounts
            WHERE client_id = %s
        ),
        property_insurance_agg AS (
            SELECT
                SUM(CASE WHEN sub_type = 'Umbrella' THEN maximum_annual_benefit ELSE 0 END) AS umbrella,
                SUM(CASE WHEN sub_type = 'Flood' THEN maximum_annual_benefit ELSE 0 END) AS flood
            FROM core.property_casualty_insurance_accounts
            WHERE client_id = %s
        ),
        
        -- FUTURE PLANNING - Helper CTEs for ratio calculations
        future_income_pv AS (
            SELECT SUM(
                CASE
                    WHEN (65 - ci.current_age) > 0 AND
                         (i.end_type IS NULL OR i.end_type != 'Age' OR (i.end_type = 'Age' AND i.end_value > 65))
                    THEN i.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, 65 - ci.current_age))) / 0.04
                    ELSE 0
                END
            ) AS pv_income
            FROM core.incomes i
            JOIN client_info ci ON i.client_id = ci.client_id
            WHERE (i.deleted IS NULL OR i.deleted = false)
        ),
        future_expenses_pv AS (
            SELECT SUM(
                CASE
                    WHEN (65 - ci.current_age) > 0 AND
                         (e.end_type IS NULL OR e.end_type != 'Age' OR
                          (e.end_type = 'Age' AND EXTRACT(YEAR FROM AGE(e.end_actual_date, e.start_actual_date)) > (65 - ci.current_age)))
                    THEN e.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, 65 - ci.current_age))) / 0.04
                    ELSE 0
                END
            ) AS pv_expenses
            FROM core.expenses e
            JOIN client_info ci ON e.client_id = ci.client_id
        ),
        survivor_income_pv AS (
            SELECT SUM(
                CASE
                    WHEN (i.end_type = 'SpousesDeath' OR i.owner_type = 'Spouse') AND
                         (i.end_value IS NULL OR i.end_value > EXTRACT(YEAR FROM CURRENT_DATE))
                    THEN i.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
                    ELSE 0
                END
            ) AS survivor_income
            FROM core.incomes i
            WHERE i.client_id = %s AND (i.deleted IS NULL OR i.deleted = false)
        ),
        survivor_expenses_pv AS (
            SELECT SUM(
                CASE
                    WHEN e.end_type != 'AtSecondDeath' AND
                         (e.end_actual_date IS NULL OR e.end_actual_date > CURRENT_DATE)
                    THEN e.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
                    ELSE 0
                END
            ) AS survivor_expenses
            FROM core.expenses e
            WHERE e.client_id = %s
        ),
        retirement_savings AS (
            SELECT SUM(COALESCE(calculated_annual_amount_usd, fixed_amount_usd)) AS retirement_savings
            FROM core.savings
            WHERE client_id = %s
                AND (destination ~* 'retirement|401k|ira' OR account_id ~* 'retirement|401k|ira')
        ),
        retirement_assets AS (
            SELECT SUM(total_value) AS current_assets
            FROM (
                SELECT total_value FROM core.investment_deposit_accounts WHERE client_id = %s
                UNION ALL
                SELECT total_value FROM core.real_estate_assets WHERE client_id = %s
                UNION ALL
                SELECT total_value FROM core.personal_property_accounts WHERE client_id = %s
            ) assets
        ),
        retirement_liabilities AS (
            SELECT SUM(total_value) AS current_liabilities
            FROM core.liability_note_accounts
            WHERE client_id = %s
        ),
        survivor_life_insurance AS (
            SELECT SUM(death_benefit) AS life_insurance_value
            FROM core.life_insurance_annuity_accounts
            WHERE client_id = %s AND death_benefit IS NOT NULL AND death_benefit > 0
        ),
        survivor_liabilities AS (
            SELECT SUM(ABS(total_value)) AS current_liabilities
            FROM core.liability_note_accounts
            WHERE client_id = %s
        ),
        ltc_future_income_pv AS (
            SELECT SUM(
                CASE
                    WHEN (i.deleted IS NULL OR i.deleted = false)
                    THEN i.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
                    ELSE 0
                END
            ) AS pv_future_income
            FROM core.incomes i
            WHERE i.client_id = %s
        ),
        education_savings_pv AS (
            SELECT SUM(
                CASE WHEN destination ~* 'education'
                THEN COALESCE(calculated_annual_amount_usd, fixed_amount_usd) * (1 - POWER(1.0/1.04, 10)) / 0.04
                ELSE 0 END
            ) AS education_savings
            FROM core.savings WHERE client_id = %s
        ),
        education_accounts AS (
            SELECT SUM(total_value) AS education_balance
            FROM (
                SELECT total_value FROM core.investment_deposit_accounts WHERE client_id = %s AND sub_type ~* 'education'
                UNION ALL
                SELECT total_value FROM core.personal_property_accounts WHERE client_id = %s
            ) edu
        ),
        education_expenses_pv AS (
            SELECT SUM(
                CASE WHEN type ~* 'education' OR sub_type ~* 'education' OR expense_item ~* 'education'
                THEN annual_amount * (1 - POWER(1.0/1.04, 10)) / 0.04
                ELSE 0 END
            ) AS education_expenses
            FROM core.expenses WHERE client_id = %s
        ),
        taxable_savings_pv AS (
            SELECT SUM(
                CASE WHEN NOT (destination ~* 'retirement|education')
                THEN COALESCE(calculated_annual_amount_usd, fixed_amount_usd) * (1 - POWER(1.0/1.04, 5)) / 0.04
                ELSE 0 END
            ) AS taxable_savings
            FROM core.savings WHERE client_id = %s
        ),
        car_expenses_pv AS (
            SELECT SUM(
                CASE WHEN expense_item ~* 'car|vehicle|auto' OR type ~* 'car|vehicle|auto' OR sub_type ~* 'car|vehicle|auto'
                THEN annual_amount * (1 - POWER(1.0/1.04, 5)) / 0.04
                ELSE 0 END
            ) AS car_expenses
            FROM core.expenses WHERE client_id = %s
        ),
        ltc_expenses_pv AS (
            SELECT SUM(
                CASE WHEN sub_type ~* 'ltc' OR fact_type_name ~* 'long term care'
                THEN COALESCE(annual_premium, 0) * (1 - POWER(1.0/1.04, 20)) / 0.04
                ELSE 0 END
            ) AS ltc_expenses
            FROM core.disability_ltc_insurance_accounts WHERE client_id = %s
        ),
        future_regular_expenses_pv AS (
            SELECT SUM(
                CASE WHEN NOT (type ~* 'ltc' OR expense_item ~* 'long term care')
                THEN annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
                ELSE 0 END
            ) AS regular_expenses
            FROM core.expenses WHERE client_id = %s
        ),
        ltd_value AS (
            SELECT SUM(benefit_amount) AS ltd_benefit
            FROM core.disability_ltc_insurance_accounts
            WHERE client_id = %s AND fact_type_name ~* 'disability'
        ),
         taxable_accounts AS (
             SELECT SUM(total_value) AS taxable_balance
             FROM core.investment_deposit_accounts
             WHERE client_id = %s
                 AND (sub_type ~* 'taxable' OR account_name ~* 'taxable|brokerage')
         ),
         
         -- WISDOM INDEX RATIOS CTEs
         savings_ratio_calc AS (
             SELECT
                 ROUND(COALESCE(savings.current_year_savings, 0) / NULLIF(income.total_income, 0), 2) as savings_ratio
             FROM (
                 SELECT
                     COALESCE(SUM(calculated_annual_amount_usd), 0) as current_year_savings
                 FROM core.savings
                 WHERE start_type = 'Active'
                   AND client_id = %s
             ) savings
             CROSS JOIN (
                 SELECT
                     COALESCE(SUM(current_year_amount), 0) as total_income
                 FROM core.incomes
                 WHERE client_id = %s
                   AND current_year_amount IS NOT NULL
             ) income
         ),
         giving_ratio_calc AS (
             SELECT
                 ROUND(COALESCE(giving.current_year_giving, 0) / NULLIF(income.total_income, 0), 2) as giving_ratio
             FROM (
                 SELECT
                     COALESCE(SUM(annual_amount), 0) AS current_year_giving
                 FROM core.expenses
                 WHERE client_id = %s
                     AND type = 'Spending'
                     AND sub_type = 'GivingAndPhilanthropy'
                     AND annual_amount > 0
                     -- Check if expense overlaps with current year
                     AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                     AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
             ) giving
             CROSS JOIN (
                 SELECT
                     COALESCE(SUM(current_year_amount), 0) as total_income
                 FROM core.incomes
                 WHERE client_id = %s
                   AND current_year_amount IS NOT NULL
             ) income
         ),
         reserves_ratio_calc AS (
             SELECT
                 ROUND((COALESCE(cash.total_cash, 0) / NULLIF(expenses.current_year_living_expenses, 0)) * 0.5, 2) as reserves
             FROM (
                 WITH holdings_cash AS (
                     SELECT
                         client_id,
                         SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) as cash_from_holdings
                     FROM core.holdings
                     WHERE asset_class = 'cash' AND value IS NOT NULL
                     AND client_id = %s
                     GROUP BY client_id
                 ),
                 investment_cash AS (
                     SELECT
                         client_id,
                         SUM(COALESCE(cash_balance, 0)) as cash_from_investments
                     FROM core.investment_deposit_accounts
                     WHERE fact_type_name = 'Cash Alternative'
                         AND cash_balance IS NOT NULL
                         AND client_id = %s
                     GROUP BY client_id
                 )
         
                 SELECT
                     COALESCE(h.cash_from_holdings, 0) + COALESCE(i.cash_from_investments, 0) as total_cash
                 FROM holdings_cash h
                 FULL OUTER JOIN investment_cash i ON h.client_id = i.client_id
             ) cash
             CROSS JOIN (
                 SELECT
                     COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
                 FROM core.expenses
                 WHERE client_id = %s
                     AND type = 'Living'
                     AND annual_amount > 0
                     -- Check if expense overlaps with current year
                     AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                     AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                     -- Ensure logical date ranges
                     AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
             ) expenses
         ),
         debt_ratio_calc AS (
             SELECT
                 ROUND(house_equity / NULLIF(house_value, 0), 2) as debt
             FROM (
                 SELECT
                     COALESCE(SUM(rea.total_value), 0) as house_value,
                     COALESCE(SUM(rea.total_value), 0) - COALESCE(SUM(CASE WHEN lna.sub_type = 'Mortgage' THEN ABS(lna.total_value) ELSE 0 END), 0) as house_equity
                 FROM core.real_estate_assets rea
                 LEFT JOIN core.liability_note_accounts lna ON rea.account_id = lna.real_estate_id
                 WHERE rea.sub_type = 'Residence'
                   AND rea.client_id = %s
                   AND (lna.sub_type = 'Mortgage' OR lna.sub_type IS NULL)
                 GROUP BY rea.client_id
             ) house_data
         ),
         diversification_ratio_calc AS (
             SELECT
                 ROUND(1 - (largest_holding / NULLIF(total_portfolio, 0)), 2) as diversification
             FROM (
                 SELECT
                     (SELECT MAX(value)
                      FROM core.holdings
                      WHERE client_id = %s
                        AND value IS NOT NULL
                        AND value > 0) as largest_holding,
                     (SELECT SUM(portfolio_value)
                      FROM (
                          SELECT COALESCE(SUM(value), 0) AS portfolio_value
                          FROM core.holdings
                          WHERE client_id = %s
                          
                          UNION ALL
                          
                          SELECT COALESCE(SUM(total_value), 0)
                          FROM core.investment_deposit_accounts
                          WHERE client_id = %s
                      ) portfolio_components) as total_portfolio
             ) calculations
         )
         
         -- FINAL SELECT: Calculate all metrics in one query
         SELECT
            -- Assets & Liabilities
            COALESCE(h.total_holdings, 0) + COALESCE(rea.total_real_estate, 0) + COALESCE(b.total_businesses, 0) +
            COALESCE(ia.total_investments, 0) + COALESCE(pp.total_personal_property, 0) - COALESCE(l.total_liabilities, 0) AS net_worth,
            
            COALESCE(h.total_holdings, 0) + COALESCE(ia.total_investments, 0) AS portfolio_value,
            COALESCE(rea.total_real_estate, 0) AS real_estate_value,
            COALESCE(l.total_liabilities, 0) AS debt,
            COALESCE(h.equity_holdings, 0) + COALESCE(ia.equity_investments, 0) AS equity,
            COALESCE(h.fixed_income, 0) AS fixed_income,
            COALESCE(h.cash_holdings, 0) + COALESCE(ia.cash_investments, 0) AS cash,
            
            -- Income Analysis
            COALESCE(ib.earned_income, 0) AS earned_income,
            COALESCE(ib.social_security, 0) AS social_security_income,
            COALESCE(ib.pension, 0) AS pension_income,
            COALESCE(ib.real_estate_income, 0) AS real_estate_income,
            COALESCE(ib.business_income, 0) AS business_income,
            COALESCE(ib.total_income, 0) AS total_income,
            
            -- Expense Tracking
            COALESCE(ge.giving, 0) AS current_year_giving,
            COALESCE(se.savings, 0) AS current_year_savings,
            ROUND(COALESCE(dp.debt_expense, 0), 2) AS current_year_debt,
            ROUND(COALESCE(ib.total_income, 0) * 0.15, 2) AS current_year_taxes,
            COALESCE(le.living, 0) AS current_year_living_expenses,
            ROUND(COALESCE(ge.giving, 0) + COALESCE(se.savings, 0) + COALESCE(dp.debt_expense, 0) +
                  ROUND(COALESCE(ib.total_income, 0) * 0.15, 2) + COALESCE(le.living, 0), 2) AS total_expenses,
            ROUND(COALESCE(ib.total_income, 0) - (COALESCE(ge.giving, 0) + COALESCE(se.savings, 0) +
                  COALESCE(dp.debt_expense, 0) + ROUND(COALESCE(ib.total_income, 0) * 0.15, 2) + COALESCE(le.living, 0)), 2) AS margin,
            
            -- Insurance Coverage
            COALESCE(li.life_insurance, 0) AS life_insurance,
            COALESCE(da.disability, 0) AS disability,
            COALESCE(da.ltc, 0) AS ltc,
            COALESCE(pi.umbrella, 0) AS umbrella,
            COALESCE(da.business_insurance, 0) AS business_insurance,
            COALESCE(pi.flood, 0) AS flood_insurance,
            COALESCE(ia.taxable_investments, 0) - COALESCE(pi.umbrella, 0) AS at_risk,
            
            -- Future Planning Ratios
            CASE WHEN ci.current_age < 65 THEN
                ROUND((COALESCE(fipv.pv_income, 0) + COALESCE(ra.current_assets, 0) +
                       COALESCE(rs.retirement_savings, 0)) /
                      NULLIF(COALESCE(fepv.pv_expenses, 0) + COALESCE(rl.current_liabilities, 0), 0), 2)
            ELSE NULL END AS retirement_ratio,
            
            ROUND((COALESCE(sli.life_insurance_value, 0) + COALESCE(sipv.survivor_income, 0) +
                   COALESCE(ra.current_assets, 0)) /
                  NULLIF(COALESCE(sepv.survivor_expenses, 0) + COALESCE(sl.current_liabilities, 0), 0), 2) AS survivor_ratio,
            
            ROUND((COALESCE(espv.education_savings, 0) + COALESCE(ea.education_balance, 0)) /
                  NULLIF(COALESCE(eepv.education_expenses, 0), 0), 2) AS education_ratio,
            
            ROUND((COALESCE(ta.taxable_balance, 0) + COALESCE(tspv.taxable_savings, 0)) /
                  NULLIF(COALESCE(cepv.car_expenses, 0), 0), 2) AS new_cars_ratio,
            
             ROUND((COALESCE(lfi.pv_future_income, 0) + COALESCE(ra.current_assets, 0)) /
                   NULLIF(COALESCE(frepv.regular_expenses, 0) + COALESCE(ltcepv.ltc_expenses, 0), 0), 2) AS ltc_ratio,
             
             ROUND(COALESCE(ltdv.ltd_benefit, 0) / NULLIF(COALESCE(ib.earned_income, 0), 0), 2) AS ltd_ratio,
             
             -- Wisdom Index Ratios
             COALESCE(src.savings_ratio, 0) AS savings_ratio,
             COALESCE(grc.giving_ratio, 0) AS giving_ratio,
             COALESCE(rrc.reserves, 0) AS reserves_ratio,
             COALESCE(drc.debt, 0) AS debt_ratio,
             COALESCE(dirc.diversification, 0) AS diversification_ratio
             
         FROM client_info ci
        LEFT JOIN holdings_agg h ON true
        LEFT JOIN investment_accounts_agg ia ON true
        LEFT JOIN real_estate_agg rea ON true
        LEFT JOIN businesses_agg b ON true
        LEFT JOIN personal_property_agg pp ON true
        LEFT JOIN liabilities_agg l ON true
        LEFT JOIN income_breakdown ib ON true
        LEFT JOIN giving_expense ge ON true
        LEFT JOIN savings_expense se ON true
        LEFT JOIN debt_payments dp ON true
        LEFT JOIN living_expense le ON true
        LEFT JOIN life_insurance_agg li ON true
        LEFT JOIN disability_agg da ON true
        LEFT JOIN property_insurance_agg pi ON true
        LEFT JOIN future_income_pv fipv ON true
        LEFT JOIN future_expenses_pv fepv ON true
        LEFT JOIN survivor_income_pv sipv ON true
        LEFT JOIN survivor_expenses_pv sepv ON true
        LEFT JOIN survivor_life_insurance sli ON true
        LEFT JOIN survivor_liabilities sl ON true
        LEFT JOIN ltc_future_income_pv lfi ON true
        LEFT JOIN retirement_savings rs ON true
        LEFT JOIN retirement_assets ra ON true
        LEFT JOIN retirement_liabilities rl ON true
        LEFT JOIN education_savings_pv espv ON true
        LEFT JOIN education_accounts ea ON true
        LEFT JOIN education_expenses_pv eepv ON true
        LEFT JOIN taxable_accounts ta ON true
        LEFT JOIN taxable_savings_pv tspv ON true
        LEFT JOIN car_expenses_pv cepv ON true
         LEFT JOIN ltc_expenses_pv ltcepv ON true
         LEFT JOIN future_regular_expenses_pv frepv ON true
         LEFT JOIN ltd_value ltdv ON true
         LEFT JOIN savings_ratio_calc src ON true
         LEFT JOIN giving_ratio_calc grc ON true
         LEFT JOIN reserves_ratio_calc rrc ON true
         LEFT JOIN debt_ratio_calc drc ON true
         LEFT JOIN diversification_ratio_calc dirc ON true;
        """
        
        # Count %s placeholders and provide matching number of client_id parameters
        import re
        param_count = len(re.findall(r'(?<!%)%s(?!%)', sql_query))
        params = tuple(client_id for _ in range(param_count))
        
        cursor.execute(sql_query, params)
        result = cursor.fetchone()
        cursor.close()
        
        if not result:
            raise Exception("No data returned from batch query")
        
        # Map result columns to metrics
        return {
            'assets_and_liabilities': {
                'net_worth': float(result[0]) if result[0] is not None else 0,
                'portfolio_value': float(result[1]) if result[1] is not None else 0,
                'real_estate_value': float(result[2]) if result[2] is not None else 0,
                'debt': float(result[3]) if result[3] is not None else 0,
                'equity': float(result[4]) if result[4] is not None else 0,
                'fixed_income': float(result[5]) if result[5] is not None else 0,
                'cash': float(result[6]) if result[6] is not None else 0
            },
            'income_analysis': {
                'earned_income': float(result[7]) if result[7] is not None else 0,
                'social_security_income': float(result[8]) if result[8] is not None else 0,
                'pension_income': float(result[9]) if result[9] is not None else 0,
                'real_estate_income': float(result[10]) if result[10] is not None else 0,
                'business_income': float(result[11]) if result[11] is not None else 0,
                'total_income': float(result[12]) if result[12] is not None else 0
            },
            'expense_tracking': {
                'current_year_giving': float(result[13]) if result[13] is not None else 0,
                'current_year_savings': float(result[14]) if result[14] is not None else 0,
                'current_year_debt': float(result[15]) if result[15] is not None else 0,
                'current_year_taxes': float(result[16]) if result[16] is not None else 0,
                'current_year_living_expenses': float(result[17]) if result[17] is not None else 0,
                'total_expenses': float(result[18]) if result[18] is not None else 0,
                'margin': float(result[19]) if result[19] is not None else 0
            },
            'insurance_coverage': {
                'life_insurance': float(result[20]) if result[20] is not None else 0,
                'disability': float(result[21]) if result[21] is not None else 0,
                'ltc': float(result[22]) if result[22] is not None else 0,
                'umbrella': float(result[23]) if result[23] is not None else 0,
                'business_insurance': float(result[24]) if result[24] is not None else 0,
                'flood_insurance': float(result[25]) if result[25] is not None else 0,
                'at_risk': float(result[26]) if result[26] is not None else 0
            },
            'future_planning_ratios': {
                'retirement_ratio': float(result[27]) if result[27] is not None else None,
                'survivor_ratio': float(result[28]) if result[28] is not None else None,
                'education_ratio': float(result[29]) if result[29] is not None else None,
                'new_cars_ratio': float(result[30]) if result[30] is not None else None,
                'ltc_ratio': float(result[31]) if result[31] is not None else None,
                'ltd_ratio': float(result[32]) if result[32] is not None else None
            },
             'wisdom_index_ratios': {
                 'savings_ratio': float(result[33]) if result[33] is not None else 0,
                 'giving_ratio': float(result[34]) if result[34] is not None else 0,
                 'reserves_ratio': float(result[35]) if result[35] is not None else 0,
                 'debt_ratio': float(result[36]) if result[36] is not None else 0,
                 'diversification_ratio': float(result[37]) if result[37] is not None else 0
             }
        }
        
    except Exception as e:
        print(f"Error getting all metrics for client {client_id}: {e}")
        # Return empty dict with all keys to prevent frontend errors
        return {
            'assets_and_liabilities': {
                'net_worth': 0, 'portfolio_value': 0, 'real_estate_value': 0,
                'debt': 0, 'equity': 0, 'fixed_income': 0, 'cash': 0
            },
            'income_analysis': {
                'earned_income': 0, 'social_security_income': 0, 'pension_income': 0,
                'real_estate_income': 0, 'business_income': 0, 'total_income': 0
            },
            'expense_tracking': {
                'current_year_giving': 0, 'current_year_savings': 0, 'current_year_debt': 0,
                'current_year_taxes': 0, 'current_year_living_expenses': 0, 'total_expenses': 0, 'margin': 0
            },
            'insurance_coverage': {
                'life_insurance': 0, 'disability': 0, 'ltc': 0,
                'umbrella': 0, 'business_insurance': 0, 'flood_insurance': 0, 'at_risk': 0
            },
            'future_planning_ratios': {
                'retirement_ratio': None, 'survivor_ratio': None, 'education_ratio': None,
                'new_cars_ratio': None, 'ltc_ratio': None, 'ltd_ratio': None
            },
             'wisdom_index_ratios': {
                 'savings_ratio': 0, 'giving_ratio': 0, 'reserves_ratio': 0, 'debt_ratio': 0, 'diversification_ratio': 0
             }
        }
    finally:
        if connection:
            close_db_connection(connection)

def get_key_metrics_for_client(client_id):
    """
    Get key metrics for client summary view.
    
    Args:
        client_id (int): The client ID
        
    Returns:
        dict: Key metrics summary
    """
    try:
        return {
            'net_worth': calculate_net_worth_for_client(client_id),
            'portfolio_value': calculate_portfolio_value_for_client(client_id),
            'total_income': calculate_total_income_for_client(client_id),
            'total_expenses': calculate_total_expenses_for_client(client_id),
            'margin': calculate_margin_for_client(client_id),
            'life_insurance': calculate_life_insurance_for_client(client_id),
            'retirement_ratio': calculate_retirement_ratio_for_client(client_id)
        }
    except Exception as e:
        print(f"Error getting key metrics for client {client_id}: {e}")
        # Return empty dict with all keys to prevent frontend errors
        return {
            'net_worth': 0,
            'portfolio_value': 0,
            'total_income': 0,
            'total_expenses': 0,
            'margin': 0,
            'life_insurance': 0,
            'retirement_ratio': 0
        }

def get_key_metrics_for_all_clients_batch():
    """
    Get key metrics for ALL clients in a single optimized query.
    This batch function reduces 42 queries (7 metrics × 6 clients) to just 1 query.
    
    Returns:
        dict: Dictionary mapping client_id to their key metrics
    """
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Single comprehensive query that calculates all 7 key metrics for all clients
        sql_query = """
        WITH
        -- Net Worth calculation
        assets AS (
            SELECT client_id, SUM(value) AS total_value FROM core.holdings GROUP BY client_id
            UNION ALL
            SELECT client_id, SUM(total_value) FROM core.real_estate_assets GROUP BY client_id
            UNION ALL
            SELECT client_id, SUM(amount) FROM core.businesses GROUP BY client_id
            UNION ALL
            SELECT client_id, SUM(total_value) FROM core.investment_deposit_accounts GROUP BY client_id
            UNION ALL
            SELECT client_id, SUM(total_value) FROM core.personal_property_accounts GROUP BY client_id
        ),
        total_assets AS (
            SELECT client_id, COALESCE(SUM(total_value), 0) AS total_assets
            FROM assets
            GROUP BY client_id
        ),
        liabilities AS (
            SELECT client_id, COALESCE(SUM(ABS(total_value)), 0) AS total_liabilities
            FROM core.liability_note_accounts
            GROUP BY client_id
        ),
        net_worth AS (
            SELECT
                COALESCE(a.client_id, l.client_id) AS client_id,
                COALESCE(a.total_assets, 0) - COALESCE(l.total_liabilities, 0) AS net_worth
            FROM total_assets a
            FULL OUTER JOIN liabilities l ON a.client_id = l.client_id
        ),
        
        -- Portfolio Value calculation
        portfolio AS (
            SELECT client_id, SUM(value) AS portfolio_value
            FROM core.holdings
            GROUP BY client_id
            UNION ALL
            SELECT client_id, SUM(total_value)
            FROM core.investment_deposit_accounts
            GROUP BY client_id
        ),
        total_portfolio AS (
            SELECT client_id, COALESCE(SUM(portfolio_value), 0) AS portfolio_value
            FROM portfolio
            GROUP BY client_id
        ),
        
        -- Total Income calculation
        total_income AS (
            SELECT client_id, COALESCE(SUM(current_year_amount), 0) AS total_income
            FROM core.incomes
            WHERE current_year_amount IS NOT NULL
            GROUP BY client_id
        ),
        
        -- Total Expenses calculation components
        giving_expense AS (
            SELECT client_id, COALESCE(SUM(annual_amount), 0) AS current_year_giving
            FROM core.expenses
            WHERE type = 'Spending'
                AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
            GROUP BY client_id
        ),
        savings_expense AS (
            SELECT client_id, COALESCE(SUM(calculated_annual_amount_usd), 0) AS current_year_savings
            FROM core.savings
            WHERE start_type = 'Active'
            GROUP BY client_id
        ),
        active_debts AS (
            SELECT
                client_id,
                CASE
                    WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                        ABS(total_value) * (interest_rate / 12) / (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                    ELSE ABS(total_value) / 12
                END AS annual_payment
            FROM core.liability_note_accounts
            WHERE total_value < 0
                AND repayment_type = 'PrincipalAndInterest'
                AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (loan_term_in_years IS NULL OR EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))
        ),
        debt_expense AS (
            SELECT client_id, ROUND(COALESCE(SUM(annual_payment), 0), 2) AS current_year_debt
            FROM active_debts
            GROUP BY client_id
        ),
        tax_expense AS (
            SELECT client_id, ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) AS current_year_taxes
            FROM core.incomes
            WHERE current_year_amount IS NOT NULL
            GROUP BY client_id
        ),
        living_expense AS (
            SELECT client_id, COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
            FROM core.expenses
            WHERE type = 'Living'
                AND annual_amount > 0
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
            GROUP BY client_id
        ),
        total_expenses AS (
            SELECT
                COALESCE(g.client_id, s.client_id, d.client_id, t.client_id, l.client_id) AS client_id,
                ROUND(
                    COALESCE(g.current_year_giving, 0) +
                    COALESCE(s.current_year_savings, 0) +
                    COALESCE(d.current_year_debt, 0) +
                    COALESCE(t.current_year_taxes, 0) +
                    COALESCE(l.current_year_living_expenses, 0), 2
                ) AS total_expenses
            FROM giving_expense g
            FULL OUTER JOIN savings_expense s USING (client_id)
            FULL OUTER JOIN debt_expense d USING (client_id)
            FULL OUTER JOIN tax_expense t USING (client_id)
            FULL OUTER JOIN living_expense l USING (client_id)
        ),
        
        -- Life Insurance calculation
        life_insurance AS (
            SELECT client_id, COALESCE(SUM(death_benefit), 0) AS life_insurance
            FROM core.life_insurance_annuity_accounts
            WHERE fact_type_name = 'Life Insurance'
            GROUP BY client_id
        ),
        
        -- Retirement Ratio calculation
        client_age AS (
            SELECT client_id, EXTRACT(YEAR FROM AGE(CURRENT_DATE, hh_date_of_birth)) AS current_age
            FROM core.clients
        ),
        retirement_eligible AS (
            SELECT client_id FROM client_age WHERE current_age < 65
        ),
        future_income_pv AS (
            SELECT i.client_id,
                SUM(
                    CASE
                        WHEN (65 - ca.current_age) > 0 AND
                             (i.end_type IS NULL OR i.end_type != 'Age' OR (i.end_type = 'Age' AND i.end_value > 65))
                        THEN i.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, 65 - ca.current_age))) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_income
            FROM core.incomes i
            JOIN client_age ca ON i.client_id = ca.client_id
            WHERE (i.deleted IS NULL OR i.deleted = false)
            GROUP BY i.client_id
        ),
        future_expenses_pv AS (
            SELECT e.client_id,
                SUM(
                    CASE
                        WHEN (65 - ca.current_age) > 0 AND
                             (e.end_type IS NULL OR e.end_type != 'Age' OR
                              (e.end_type = 'Age' AND EXTRACT(YEAR FROM AGE(e.end_actual_date, e.start_actual_date)) > (65 - ca.current_age)))
                        THEN e.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, 65 - ca.current_age))) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_expenses
            FROM core.expenses e
            JOIN client_age ca ON e.client_id = ca.client_id
            GROUP BY e.client_id
        ),
        current_assets_detail AS (
            SELECT client_id, total_value FROM core.investment_deposit_accounts
            UNION ALL
            SELECT client_id, total_value FROM core.real_estate_assets
            UNION ALL
            SELECT client_id, total_value FROM core.personal_property_accounts
        ),
        current_assets_sum AS (
            SELECT client_id, SUM(total_value) AS current_assets
            FROM current_assets_detail
            GROUP BY client_id
        ),
        retirement_savings AS (
            SELECT client_id, SUM(COALESCE(calculated_annual_amount_usd, fixed_amount_usd)) AS retirement_savings
            FROM core.savings
            WHERE destination ~* 'retirement|401k|ira' OR account_id ~* 'retirement|401k|ira'
            GROUP BY client_id
        ),
        current_liabilities_sum AS (
            SELECT client_id, SUM(total_value) AS current_liabilities
            FROM core.liability_note_accounts
            GROUP BY client_id
        ),
        retirement_ratio AS (
            SELECT
                re.client_id,
                ROUND(
                    (COALESCE(fi.pv_future_income, 0) + COALESCE(ca.current_assets, 0) + COALESCE(rs.retirement_savings, 0)) /
                    NULLIF((COALESCE(fe.pv_future_expenses, 0) + COALESCE(cl.current_liabilities, 0)), 0),
                    2
                ) AS retirement_ratio
            FROM retirement_eligible re
            LEFT JOIN future_income_pv fi ON re.client_id = fi.client_id
            LEFT JOIN future_expenses_pv fe ON re.client_id = fe.client_id
            LEFT JOIN current_assets_sum ca ON re.client_id = ca.client_id
            LEFT JOIN retirement_savings rs ON re.client_id = rs.client_id
            LEFT JOIN current_liabilities_sum cl ON re.client_id = cl.client_id
        ),
        
        -- Get all clients
        all_clients AS (
            SELECT DISTINCT client_id FROM core.clients
        )
        
        -- Final result combining all metrics
        SELECT
            ac.client_id,
            COALESCE(nw.net_worth, 0) AS net_worth,
            COALESCE(pv.portfolio_value, 0) AS portfolio_value,
            COALESCE(ti.total_income, 0) AS total_income,
            COALESCE(te.total_expenses, 0) AS total_expenses,
            COALESCE(ti.total_income, 0) - COALESCE(te.total_expenses, 0) AS margin,
            COALESCE(li.life_insurance, 0) AS life_insurance,
            rr.retirement_ratio
        FROM all_clients ac
        LEFT JOIN net_worth nw ON ac.client_id = nw.client_id
        LEFT JOIN total_portfolio pv ON ac.client_id = pv.client_id
        LEFT JOIN total_income ti ON ac.client_id = ti.client_id
        LEFT JOIN total_expenses te ON ac.client_id = te.client_id
        LEFT JOIN life_insurance li ON ac.client_id = li.client_id
        LEFT JOIN retirement_ratio rr ON ac.client_id = rr.client_id
        ORDER BY ac.client_id;
        """
        
        cursor.execute(sql_query)
        results = cursor.fetchall()
        cursor.close()
        
        # Format results as dictionary keyed by client_id
        metrics_by_client = {}
        for row in results:
            client_id = row[0]
            metrics_by_client[client_id] = {
                'net_worth': float(row[1]) if row[1] is not None else 0,
                'portfolio_value': float(row[2]) if row[2] is not None else 0,
                'total_income': float(row[3]) if row[3] is not None else 0,
                'total_expenses': float(row[4]) if row[4] is not None else 0,
                'margin': float(row[5]) if row[5] is not None else 0,
                'life_insurance': float(row[6]) if row[6] is not None else 0,
                'retirement_ratio': float(row[7]) if row[7] is not None else None
            }
        
        return metrics_by_client
        
    except Exception as e:
        print(f"Error getting batch key metrics: {e}")
        return {}
    finally:
        if connection:
            close_db_connection(connection)

def update_targets_for_client(client_id, targets_dict):
    """
    Update targets for a specific client (admin function).
    This function inserts new target records only if the value has changed,
    maintaining a complete history of target values without duplicates.
    
    Args:
        client_id (int): The client ID
        targets_dict (dict): Dictionary of targets to update
        
    Returns:
        bool: True if successful, False otherwise
    """
    # Get current targets for this client to check for duplicates
    get_current_sql = """
        SELECT DISTINCT ON (metric_name) metric_name, target_value, created_at
        FROM core.metric_targets
        WHERE client_id = %s
        ORDER BY metric_name, created_at DESC
    """
    
    # Insert new targets without deleting existing ones to maintain history
    insert_sql = """
        INSERT INTO core.metric_targets (client_id, metric_name, target_value, created_at)
        VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Get current targets
        cursor.execute(get_current_sql, (client_id,))
        current_targets_data = cursor.fetchall()
        current_targets = {row[0]: float(row[1]) for row in current_targets_data if row[1] is not None}
        
        # Only insert targets that have changed or are new
        changed_targets = []
        for metric_name, target_value in targets_dict.items():
            current_value = current_targets.get(metric_name)
            # Insert if it's a new target or if the value has changed
            if current_value is None or current_value != target_value:
                changed_targets.append((client_id, metric_name, target_value))
        
        # Insert only the changed targets
        if changed_targets:
            from psycopg2.extras import execute_batch
            execute_batch(cursor, insert_sql, changed_targets)
            connection.commit()
        
        cursor.close()
        
        return True
        
    except Exception as e:
        print(f"Error updating targets for client {client_id}: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if connection:
            close_db_connection(connection)

def get_all_targets_for_client(client_id):
    """
    Get all the most recent target values for a specific client (admin function).
    Returns the most recent target for each metric based on created_at timestamp.
    
    Args:
        client_id (int): The client ID to get targets for
        
    Returns:
        dict: Dictionary mapping metric names to target values
    """
    sql_query = """
        SELECT DISTINCT ON (metric_name) metric_name, target_value
        FROM core.metric_targets
        WHERE client_id = %s
        ORDER BY metric_name, created_at DESC
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id,))
        results = cursor.fetchall()
        cursor.close()
        
        targets = {}
        for row in results:
            if row[1] is not None:
                targets[row[0]] = float(row[1])
        
        return targets
        
    except Exception as e:
        print(f"Error getting all targets for client {client_id}: {e}")
        return {}
    finally:
        if connection:
            close_db_connection(connection)

def delete_target_for_client(client_id, metric_name):
    """
    Delete the most recent target value for a specific metric for a specific client (admin function).
    
    Args:
        client_id (int): The client ID
        metric_name (str): The name of the metric
        
    Returns:
        bool: True if successful, False otherwise
    """
    sql_query = """
        DELETE FROM core.metric_targets
        WHERE client_id = %s AND metric_name = %s
        AND created_at = (
            SELECT MAX(created_at)
            FROM core.metric_targets
            WHERE client_id = %s AND metric_name = %s
        )
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id, metric_name, client_id, metric_name))
        connection.commit()
        rows_affected = cursor.rowcount
        cursor.close()
        
        print(f"Deleted {rows_affected} target(s) for {metric_name} for client {client_id}")
        return rows_affected > 0
        
    except Exception as e:
        print(f"Error deleting target for {metric_name} for client {client_id}: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if connection:
            close_db_connection(connection)

def delete_all_targets_for_client(client_id):
    """
    Delete all target values for a specific client (admin function).
    
    Args:
        client_id (int): The client ID
        
    Returns:
        bool: True if successful, False otherwise
    """
    sql_query = """
        DELETE FROM core.metric_targets
        WHERE client_id = %s
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id,))
        connection.commit()
        rows_affected = cursor.rowcount
        cursor.close()
        
        print(f"Deleted {rows_affected} target(s) for client {client_id}")
        return True
        
    except Exception as e:
        print(f"Error deleting all targets for client {client_id}: {e}")
        if connection:
            connection.rollback()
        return False
    finally:
        if connection:
            close_db_connection(connection)

# Client-specific metric calculation functions

def calculate_net_worth_for_client(client_id):
    """Calculate net worth for specific client."""
    sql_query = """
        WITH assets AS (
            SELECT COALESCE(SUM(value), 0) AS total_value FROM core.holdings WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(total_value), 0) FROM core.real_estate_assets WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(amount), 0) FROM core.businesses WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(total_value), 0) FROM core.investment_deposit_accounts WHERE client_id = %s
            UNION ALL
            SELECT COALESCE(SUM(total_value), 0) FROM core.personal_property_accounts WHERE client_id = %s
        ),
        liabilities AS (
            SELECT COALESCE(SUM(ABS(total_value)), 0) AS total_liabilities
            FROM core.liability_note_accounts
            WHERE client_id = %s
        ),
        asset_summary AS (
            SELECT SUM(total_value) AS total_assets FROM assets
        )
        SELECT total_assets - total_liabilities AS net_worth FROM asset_summary, liabilities;
    """
    
    return execute_metric_query(sql_query, client_id, "net_worth_for_client")

def calculate_portfolio_value_for_client(client_id):
    """Calculate portfolio value for specific client."""
    sql_query = """
        WITH portfolio_components AS (
            -- Holdings
            SELECT COALESCE(SUM(value), 0) AS portfolio_value
            FROM core.holdings
            WHERE client_id = %s
            
            UNION ALL
            
            -- Investment deposit accounts
            SELECT COALESCE(SUM(total_value), 0)
            FROM core.investment_deposit_accounts
            WHERE client_id = %s
        )
        SELECT SUM(portfolio_value) AS total_portfolio_value
        FROM portfolio_components;
    """
    return execute_metric_query(sql_query, client_id, "portfolio_value_for_client")

def calculate_real_estate_value_for_client(client_id):
    """Calculate real estate value for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(total_value),0) as real_estate_value
        FROM core.real_estate_assets
        WHERE client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "real_estate_value_for_client")

def calculate_debt_for_client(client_id):
    """Calculate debt for specific client."""
    sql_query = """
        SELECT
        ABS(COALESCE(SUM(total_value),0))
        FROM core.liability_note_accounts
        WHERE client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "debt_for_client")

def calculate_equity_for_client(client_id):
    """Calculate equity for specific client."""
    sql_query = """
        WITH equity_holdings AS (
            SELECT
                client_id,
                COALESCE(SUM(value),0) as equity_holdings_value
            FROM core.holdings
            WHERE asset_class IN ('largecap', 'smallcap', 'largevalue', 'smallvalue', 'internat', 'emerging', 'ips')
                AND client_id = %s
                AND value IS NOT NULL
            GROUP BY client_id
        ),
        investment_equity AS (
            SELECT
                client_id,
                COALESCE(SUM(holdings_value),0) as investment_equity_value
            FROM core.investment_deposit_accounts
            WHERE fact_type_name IN ('Taxable Investment', 'Roth IRA', 'Qualified Retirement')
                AND client_id = %s
                AND holdings_value IS NOT NULL
            GROUP BY client_id
        )
        SELECT
            (COALESCE(eh.equity_holdings_value, 0) + COALESCE(ie.investment_equity_value, 0)) as total_equity
        FROM equity_holdings eh
        FULL OUTER JOIN investment_equity ie ON eh.client_id = ie.client_id;
    """
    return execute_metric_query(sql_query, client_id, "equity_for_client")

def calculate_fixed_income_for_client(client_id):
    """Calculate fixed income for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(value),0) as fixed_income_total
        FROM core.holdings
        WHERE asset_class IN ('highyldbond', 'inttermmun', 'investbond', 'shortermbond', 'shortermmun')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "fixed_income_for_client")

def calculate_cash_for_client(client_id):
    """Calculate cash for specific client."""
    sql_query = """
        WITH holdings_cash AS (
            SELECT
                client_id,
                SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) as cash_from_holdings
            FROM core.holdings
            WHERE asset_class = 'cash' AND value IS NOT NULL
            AND client_id = %s
            GROUP BY client_id
        ),
        investment_cash AS (
            SELECT
                client_id,
                SUM(COALESCE(cash_balance, 0)) as cash_from_investments
            FROM core.investment_deposit_accounts
            WHERE fact_type_name = 'Cash Alternative'
                AND cash_balance IS NOT NULL
                AND client_id = %s
            GROUP BY client_id
        )
        SELECT
            COALESCE(h.cash_from_holdings, 0) + COALESCE(i.cash_from_investments, 0) as total_cash
        FROM holdings_cash h
        FULL OUTER JOIN investment_cash i ON h.client_id = i.client_id;
    """
    return execute_metric_query(sql_query, client_id, "cash_for_client")

def calculate_earned_income_for_client(client_id):
    """Calculate earned income for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as earned_income
        FROM core.incomes
        WHERE income_type IN ('Salary')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "earned_income_for_client")

def calculate_social_security_income_for_client(client_id):
    """Calculate social security income for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as social_income
        FROM core.incomes
        WHERE income_type IN ('SocialSecurity')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "social_security_income_for_client")

def calculate_pension_income_for_client(client_id):
    """Calculate pension income for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as pension_income
        FROM core.incomes
        WHERE income_type IN ('Pension')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "pension_income_for_client")

def calculate_real_estate_income_for_client(client_id):
    """Calculate real estate income for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as real_estate_income
        FROM core.incomes
        WHERE income_type IN ('Real Estate')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "real_estate_income_for_client")

def calculate_business_income_for_client(client_id):
    """Calculate business income for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(current_year_amount),0) as business_income
        FROM core.incomes
        WHERE income_type IN ('Business')
            AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "business_income_for_client")

def calculate_total_income_for_client(client_id):
    """Calculate total income for specific client."""
    sql_query = """
        WITH income_breakdown AS (
            SELECT
                client_id,
                -- Earned Income
                COALESCE(SUM(CASE WHEN income_type = 'Salary' THEN current_year_amount ELSE 0 END), 0) as earned_income,
                -- Social Income
                COALESCE(SUM(CASE WHEN income_type = 'SocialSecurity' THEN current_year_amount ELSE 0 END), 0) as social_income,
                -- Pension Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Pension' THEN current_year_amount ELSE 0 END), 0) as pension_income,
                -- Real Estate Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Real Estate' THEN current_year_amount ELSE 0 END), 0) as real_estate_income,
                -- Business Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Business' THEN current_year_amount ELSE 0 END), 0) as business_income,
                -- Other Income (exists as 'Other' in database)
                COALESCE(SUM(CASE WHEN income_type = 'Other' THEN current_year_amount ELSE 0 END), 0) as other_income,
                -- Total Income = Sum of all components
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s  -- Filter for specific client
              AND current_year_amount IS NOT NULL
            GROUP BY client_id
        )
        SELECT
            total_income
        FROM income_breakdown;
    """
    return execute_metric_query(sql_query, client_id, "total_income_for_client")

def calculate_current_year_giving_for_client(client_id):
    """Calculate current year giving for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(annual_amount), 0) AS current_year_giving
        FROM core.expenses
        WHERE client_id = %s
            AND type = 'Spending'
            AND sub_type = 'GivingAndPhilanthropy'
            AND annual_amount > 0
            AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
            AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE));
    """
    return execute_metric_query(sql_query, client_id, "current_year_giving_for_client")

def calculate_current_year_savings_for_client(client_id):
    """Calculate current year savings for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(calculated_annual_amount_usd),0) as current_year_savings
        FROM core.savings
        WHERE start_type = 'Active'
          AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "current_year_savings_for_client")

def calculate_current_year_debt_for_client(client_id):
    """Calculate current year debt for specific client."""
    sql_query = """
        WITH active_debts AS (
            SELECT
                client_id,
                CASE
                  WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                    ABS(total_value) * (interest_rate / 12) /
                    (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                  ELSE
                    ABS(total_value) / 12
                END as annual_payment
            FROM core.liability_note_accounts
            WHERE client_id = %s
              AND total_value < 0
              AND repayment_type = 'PrincipalAndInterest'
              AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
              AND (loan_term_in_years IS NULL OR
                   EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))
        )
        SELECT ROUND(SUM(annual_payment),2) as current_year_debt
        FROM active_debts;
    """
    return execute_metric_query(sql_query, client_id, "current_year_debt_for_client")

def calculate_current_year_taxes_for_client(client_id):
    """Calculate current year taxes for specific client."""
    sql_query = """
        SELECT
            ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) as current_year_taxes
        FROM core.incomes
        WHERE client_id = %s
          AND current_year_amount IS NOT NULL;
    """
    return execute_metric_query(sql_query, client_id, "current_year_taxes_for_client")

def calculate_current_year_living_expenses_for_client(client_id):
    """Calculate current year living expenses for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
        FROM core.expenses
        WHERE client_id = %s
            AND type = 'Living'
            AND annual_amount > 0
            AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
            AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
            AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date);
    """
    return execute_metric_query(sql_query, client_id, "current_year_living_expenses_for_client")

def calculate_total_expenses_for_client(client_id):
    """Calculate total expenses for specific client."""
    sql_query = """
        WITH
        -- Giving Expense
        giving_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_giving
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Spending'
                AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        ),

        -- Savings Expense
        savings_expense AS (
            SELECT
                COALESCE(SUM(calculated_annual_amount_usd), 0) as current_year_savings
            FROM core.savings
            WHERE start_type = 'Active'
              AND client_id = %s
        ),

        -- Debt Expense
        debt_expense AS (
            WITH active_debts AS (
                SELECT
                    client_id,
                    account_name,
                    total_value,
                    interest_rate,
                    loan_term_in_years,
                    payment_frequency,
                    loan_date,
                    -- Calculate annual payment using amortization formula or simplified estimate
                    CASE
                      WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                        -- Standard amortization: monthly payment * 12
                        ABS(total_value) * (interest_rate / 12) /
                        (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                      ELSE
                        -- For loans missing rate/term, assume 12-month repayment
                        ABS(total_value) / 12
                    END as annual_payment
                FROM core.liability_note_accounts
                WHERE client_id = %s -- Filter for specific client
                  AND total_value < 0  -- Only include actual debt (negative values)
                  AND repayment_type = 'PrincipalAndInterest'  -- Only active debt being serviced
                  -- DYNAMIC CURRENT YEAR LOGIC:
                  AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)  -- Loan originated before or in current year
                  AND (loan_term_in_years IS NULL OR
                       EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))  -- Still active in current year
            )
            SELECT
                ROUND(SUM(annual_payment),2) as current_year_debt
            FROM active_debts
        ),

        -- Tax Expense
        tax_expense AS (
            SELECT
                ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) as current_year_taxes
            FROM core.incomes
            WHERE client_id = %s -- Filter for specific client
              AND current_year_amount IS NOT NULL
        ),

        -- Living Expense
        living_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Living'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                -- Ensure logical date ranges
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
        )

        -- Final Total Expense Calculation
        SELECT
            -- Total Expense = Sum of all components
            ROUND((COALESCE(g.current_year_giving, 0) +
             COALESCE(s.current_year_savings, 0) +
             COALESCE(d.current_year_debt, 0) +
             COALESCE(t.current_year_taxes, 0) +
             COALESCE(l.current_year_living_expenses, 0)),2) as total_expense
        FROM giving_expense g, savings_expense s, debt_expense d, tax_expense t, living_expense l;
    """
    return execute_metric_query(sql_query, client_id, "total_expenses_for_client")

def calculate_margin_for_client(client_id):
    """Calculate margin for specific client."""
    sql_query = """
        WITH
        -- Total Income Calculation
        income_breakdown AS (
            SELECT
                client_id,
                -- Earned Income
                COALESCE(SUM(CASE WHEN income_type = 'Salary' THEN current_year_amount ELSE 0 END), 0) as earned_income,
                -- Social Income
                COALESCE(SUM(CASE WHEN income_type = 'SocialSecurity' THEN current_year_amount ELSE 0 END), 0) as social_income,
                -- Pension Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Pension' THEN current_year_amount ELSE 0 END), 0) as pension_income,
                -- Real Estate Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Real Estate' THEN current_year_amount ELSE 0 END), 0) as real_estate_income,
                -- Business Income (no data exists but included for completeness)
                COALESCE(SUM(CASE WHEN income_type = 'Business' THEN current_year_amount ELSE 0 END), 0) as business_income,
                -- Other Income (exists as 'Other Income' in database)
                COALESCE(SUM(CASE WHEN income_type = 'Other' THEN current_year_amount ELSE 0 END), 0) as other_income,
                -- Total Income = Sum of all components
                COALESCE(SUM(current_year_amount), 0) as total_income
            FROM core.incomes
            WHERE client_id = %s  -- Filter for specific client
              AND current_year_amount IS NOT NULL
            GROUP BY client_id
        ),

        -- Giving Expense
        giving_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_giving
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Spending'
                AND sub_type = 'GivingAndPhilanthropy'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        ),

        -- Savings Expense
        savings_expense AS (
            SELECT
                COALESCE(SUM(calculated_annual_amount_usd), 0) as current_year_savings
            FROM core.savings
            WHERE start_type = 'Active'
              AND client_id = %s
        ),

        -- Debt Expense
        debt_expense AS (
            WITH active_debts AS (
                SELECT
                    client_id,
                    account_name,
                    total_value,
                    interest_rate,
                    loan_term_in_years,
                    payment_frequency,
                    loan_date,
                    -- Calculate annual payment using amortization formula or simplified estimate
                    CASE
                      WHEN interest_rate IS NOT NULL AND loan_term_in_years IS NOT NULL THEN
                        -- Standard amortization: monthly payment * 12
                        ABS(total_value) * (interest_rate / 12) /
                        (1 - POWER(1 + (interest_rate / 12), -loan_term_in_years * 12)) * 12
                      ELSE
                        -- For loans missing rate/term, assume 12-month repayment
                        ABS(total_value) / 12
                    END as annual_payment
                FROM core.liability_note_accounts
                WHERE client_id = %s -- Filter for specific client
                  AND total_value < 0  -- Only include actual debt (negative values)
                  AND repayment_type = 'PrincipalAndInterest'  -- Only active debt being serviced
                  -- DYNAMIC CURRENT YEAR LOGIC:
                  AND EXTRACT(YEAR FROM loan_date) <= EXTRACT(YEAR FROM CURRENT_DATE)  -- Loan originated before or in current year
                  AND (loan_term_in_years IS NULL OR
                       EXTRACT(YEAR FROM loan_date) + loan_term_in_years >= EXTRACT(YEAR FROM CURRENT_DATE))  -- Still active in current year
            )
            SELECT
                ROUND(SUM(annual_payment),2) as current_year_debt
            FROM active_debts
        ),

        -- Tax Expense
        tax_expense AS (
            SELECT
                ROUND(COALESCE(SUM(current_year_amount), 0) * 0.15, 2) as current_year_taxes
            FROM core.incomes
            WHERE client_id = %s -- Filter for specific client
              AND current_year_amount IS NOT NULL
        ),

        -- Living Expense
        living_expense AS (
            SELECT
                COALESCE(SUM(annual_amount), 0) AS current_year_living_expenses
            FROM core.expenses
            WHERE client_id = %s
                AND type = 'Living'
                AND annual_amount > 0
                -- Check if expense overlaps with current year
                AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
                AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
                -- Ensure logical date ranges
                AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
        ),

        -- Total Expense Calculation
        total_expense_calc AS (
            SELECT
                ROUND((COALESCE(g.current_year_giving, 0) +
                       COALESCE(s.current_year_savings, 0) +
                       COALESCE(d.current_year_debt, 0) +
                       COALESCE(t.current_year_taxes, 0) +
                       COALESCE(l.current_year_living_expenses, 0)), 2) as total_expense
            FROM giving_expense g, savings_expense s, debt_expense d, tax_expense t, living_expense l
        )

        -- Final Margin Calculation (Total Income - Total Expenses)
        SELECT
            -- Margin = Total Income - Total Expenses, rounded to 2 decimal places
            ROUND((COALESCE(i.total_income, 0) - COALESCE(e.total_expense, 0)), 2) as margin
        FROM income_breakdown i, total_expense_calc e;
    """
    return execute_metric_query(sql_query, client_id, "margin_for_client")

def calculate_life_insurance_for_client(client_id):
    """Calculate life insurance for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(death_benefit), 0) as life_insurance_metric
        FROM core.life_insurance_annuity_accounts
        WHERE fact_type_name = 'Life Insurance'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "life_insurance_for_client")

def calculate_disability_for_client(client_id):
    """Calculate disability for specific client."""
    sql_query = """
        SELECT COALESCE(SUM(benefit_amount), 0) as disability_metric
        FROM core.disability_ltc_insurance_accounts
        WHERE fact_type_name IN ('Disability Policy', 'Business Disability Policy')
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "disability_for_client")

def calculate_ltc_for_client(client_id):
    """Calculate LTC for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(benefit_amount),0) as ltc_metric
        FROM core.disability_ltc_insurance_accounts
        WHERE sub_type = 'PersonalLT'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "ltc_for_client")

def calculate_umbrella_for_client(client_id):
    """Calculate umbrella for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(maximum_annual_benefit),0) as umbrella_metric
        FROM core.property_casualty_insurance_accounts
        WHERE sub_type = 'Umbrella'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "umbrella_for_client")

def calculate_business_insurance_for_client(client_id):
    """Calculate business insurance for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(benefit_amount), 0) AS business_insurance
        FROM core.disability_ltc_insurance_accounts
        WHERE sub_type = 'BusinessReducingTerm'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "business_insurance_for_client")

def calculate_flood_insurance_for_client(client_id):
    """Calculate flood insurance for specific client."""
    sql_query = """
        SELECT
            COALESCE(SUM(maximum_annual_benefit), 0) as flood_insurance_metric
        FROM core.property_casualty_insurance_accounts
        WHERE sub_type = 'Flood'
        AND client_id = %s;
    """
    return execute_metric_query(sql_query, client_id, "flood_insurance_for_client")

def calculate_at_risk_for_client(client_id):
    """Calculate at risk for specific client."""
    sql_query = """
        WITH taxable AS (
            SELECT COALESCE(SUM(total_value), 0) AS taxable_investments_usd
            FROM core.investment_deposit_accounts
            WHERE client_id = %s
              AND fact_type_name = 'Taxable Investment'
        ),
        umbrella AS (
            SELECT COALESCE(SUM(maximum_annual_benefit), 0) AS umbrella_coverage_usd
            FROM core.property_casualty_insurance_accounts
            WHERE client_id = %s
              AND sub_type = 'Umbrella'
        )
        SELECT (taxable.taxable_investments_usd - umbrella.umbrella_coverage_usd) AS at_risk_usd
        FROM taxable, umbrella;
    """
    return execute_metric_query(sql_query, client_id, "at_risk_for_client")

def calculate_retirement_ratio_for_client(client_id):
    """Calculate retirement ratio for specific client."""
    sql_query = """
        WITH
        client_info AS (
            SELECT
                c.client_id,
                EXTRACT(YEAR FROM AGE(CURRENT_DATE, c.hh_date_of_birth)) AS current_age,
                65 AS retirement_age
            FROM core.clients c
            WHERE c.client_id = %s
        ),
        future_income_pv AS (
            SELECT
                i.client_id,
                SUM(
                    CASE
                        WHEN (ci.retirement_age - ci.current_age) > 0 AND
                             (i.end_type IS NULL OR
                              i.end_type != 'Age' OR
                              (i.end_type = 'Age' AND i.end_value > ci.retirement_age))
                        THEN i.annual_amount *
                             (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_income
            FROM core.incomes i
            JOIN client_info ci ON i.client_id = ci.client_id
            WHERE i.deleted IS NULL OR i.deleted = false
            GROUP BY i.client_id
        ),
        future_expenses_pv AS (
            SELECT
                e.client_id,
                SUM(
                    CASE
                        WHEN (ci.retirement_age - ci.current_age) > 0 AND
                             (e.end_type IS NULL OR
                              e.end_type != 'Age' OR
                              (e.end_type = 'Age' AND
                               EXTRACT(YEAR FROM AGE(e.end_actual_date, e.start_actual_date)) > (ci.retirement_age - ci.current_age)))
                        THEN e.annual_amount *
                             (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_expenses
            FROM core.expenses e
            JOIN client_info ci ON e.client_id = ci.client_id
            GROUP BY e.client_id
        ),
        current_assets AS (
            SELECT
                client_id,
                SUM(total_value) AS current_assets
            FROM (
                SELECT client_id, total_value FROM core.investment_deposit_accounts
                UNION ALL
                SELECT client_id, total_value FROM core.real_estate_assets
                UNION ALL
                SELECT client_id, total_value FROM core.personal_property_accounts
            ) all_assets
            GROUP BY client_id
        ),
        retirement_savings AS (
            SELECT
                client_id,
                SUM(COALESCE(calculated_annual_amount_usd, fixed_amount_usd)) AS retirement_savings
            FROM core.savings
            WHERE destination ~* 'retirement|401k|ira' OR
                  account_id ~* 'retirement|401k|ira'
            GROUP BY client_id
        ),
        current_liabilities AS (
            SELECT
                client_id,
                SUM(total_value) AS current_liabilities
            FROM core.liability_note_accounts
            GROUP BY client_id
        )
        SELECT
            ROUND(
                (
                    COALESCE(fi.pv_future_income, 0) +
                    COALESCE(ca.current_assets, 0) +
                    COALESCE(rs.retirement_savings, 0)
                ) /
                NULLIF(
                    (COALESCE(fe.pv_future_expenses, 0) + COALESCE(cl.current_liabilities, 0)),
                    0
                ),
                2
            ) AS retirement_ratio
        FROM client_info ci
        LEFT JOIN future_income_pv fi ON ci.client_id = fi.client_id
        LEFT JOIN future_expenses_pv fe ON ci.client_id = fe.client_id
        LEFT JOIN current_assets ca ON ci.client_id = ca.client_id
        LEFT JOIN retirement_savings rs ON ci.client_id = rs.client_id
        LEFT JOIN current_liabilities cl ON ci.client_id = cl.client_id
        WHERE ci.current_age < ci.retirement_age;
    """
    return execute_metric_query(sql_query, client_id, "retirement_ratio_for_client")

def calculate_survivor_ratio_for_client(client_id):
    """Calculate survivor ratio for specific client."""
    sql_query = """
        WITH
        client_info AS (
            SELECT
                c.client_id,
                EXTRACT(YEAR FROM AGE(CURRENT_DATE, c.hh_date_of_birth)) AS current_age
            FROM core.clients c
            WHERE c.client_id = %s
        ),
        future_income_pv AS (
            SELECT
                i.client_id,
                SUM(
                    CASE
                        WHEN (i.end_type = 'SpousesDeath' OR i.owner_type = 'Spouse') AND
                             (i.end_value IS NULL OR i.end_value > EXTRACT(YEAR FROM CURRENT_DATE))
                        THEN i.annual_amount *
                             (1 - POWER(1.0/1.04, 20)) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_income
            FROM core.incomes i
            JOIN client_info ci ON i.client_id = ci.client_id
            WHERE (i.deleted IS NULL OR i.deleted = false)
            GROUP BY i.client_id
        ),
        future_expenses_pv AS (
            SELECT
                e.client_id,
                SUM(
                    CASE
                        WHEN e.end_type != 'AtSecondDeath' AND
                             (e.end_actual_date IS NULL OR e.end_actual_date > CURRENT_DATE)
                        THEN e.annual_amount *
                             (1 - POWER(1.0/1.04, 20)) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_expenses
            FROM core.expenses e
            JOIN client_info ci ON e.client_id = ci.client_id
            GROUP BY e.client_id
        ),
        current_assets AS (
            SELECT
                client_id,
                SUM(total_value) AS current_assets
            FROM (
                SELECT client_id, total_value FROM core.investment_deposit_accounts
                UNION ALL
                SELECT client_id, total_value FROM core.real_estate_assets
                UNION ALL
                SELECT client_id, total_value FROM core.personal_property_accounts
            ) all_assets
            GROUP BY client_id
        ),
        life_insurance AS (
            SELECT
                client_id,
                SUM(death_benefit) AS life_insurance_value
            FROM core.life_insurance_annuity_accounts
            WHERE death_benefit IS NOT NULL AND death_benefit > 0
            GROUP BY client_id
        ),
        current_liabilities AS (
            SELECT
                client_id,
                ABS(SUM(total_value)) AS current_liabilities
            FROM core.liability_note_accounts
            GROUP BY client_id
        )
        SELECT
            ROUND(
                (
                    COALESCE(li.life_insurance_value, 0) +
                    COALESCE(fi.pv_future_income, 0) +
                    COALESCE(ca.current_assets, 0)
                ) /
                NULLIF(
                    (COALESCE(fe.pv_future_expenses, 0) + COALESCE(cl.current_liabilities, 0)),
                    0
                ),
                2
            ) AS survivor_ratio
        FROM client_info ci
        LEFT JOIN future_income_pv fi ON ci.client_id = fi.client_id
        LEFT JOIN future_expenses_pv fe ON ci.client_id = fe.client_id
        LEFT JOIN current_assets ca ON ci.client_id = ca.client_id
        LEFT JOIN life_insurance li ON ci.client_id = li.client_id
        LEFT JOIN current_liabilities cl ON ci.client_id = cl.client_id;
    """
    return execute_metric_query(sql_query, client_id, "survivor_ratio_for_client")

def calculate_education_ratio_for_client(client_id):
    """Calculate education ratio for specific client."""
    sql_query = """
        WITH
        client_info AS (
            SELECT
                c.client_id
            FROM core.clients c
            WHERE c.client_id = %s
        ),
        education_savings_pv AS (
            SELECT
                s.client_id,
                SUM(
                    CASE
                        WHEN s.destination ~* 'education'
                        THEN COALESCE(s.calculated_annual_amount_usd, s.fixed_amount_usd) *
                             (1 - POWER(1.0/1.04, 10)) / 0.04
                        ELSE 0
                    END
                ) AS pv_education_savings
            FROM core.savings s
            JOIN client_info ci ON s.client_id = ci.client_id
            GROUP BY s.client_id
        ),
        education_accounts AS (
            SELECT
                client_id,
                SUM(total_value) AS education_account_balances
            FROM (
                SELECT client_id, total_value
                FROM core.investment_deposit_accounts
                WHERE sub_type ~* 'education'
                UNION ALL
                SELECT client_id, total_value
                FROM core.personal_property_accounts
            ) edu_accounts
            GROUP BY client_id
        ),
        education_expenses_pv AS (
            SELECT
                e.client_id,
                SUM(
                    CASE
                        WHEN e.type ~* 'education' OR
                             e.sub_type ~* 'education' OR
                             e.expense_item ~* 'education'
                        THEN e.annual_amount *
                             (1 - POWER(1.0/1.04, 10)) / 0.04
                        ELSE 0
                    END
                ) AS pv_education_expenses
            FROM core.expenses e
            JOIN client_info ci ON e.client_id = ci.client_id
            GROUP BY e.client_id
        )
        SELECT
            ROUND(
                (
                    COALESCE(es.pv_education_savings, 0) +
                    COALESCE(ea.education_account_balances, 0)
                ) /
                NULLIF(COALESCE(ee.pv_education_expenses, 0), 0),
                2
            ) AS education_ratio
        FROM client_info ci
        LEFT JOIN education_savings_pv es ON ci.client_id = es.client_id
        LEFT JOIN education_accounts ea ON ci.client_id = ea.client_id
        LEFT JOIN education_expenses_pv ee ON ci.client_id = ee.client_id;
    """
    return execute_metric_query(sql_query, client_id, "education_ratio_for_client")

def calculate_new_cars_ratio_for_client(client_id):
    """Calculate new cars ratio for specific client."""
    sql_query = """
        WITH
        client_info AS (
            SELECT
                c.client_id
            FROM core.clients c
            WHERE c.client_id = %s
        ),
        taxable_accounts AS (
            SELECT
                client_id,
                SUM(total_value) AS taxable_account_value
            FROM core.investment_deposit_accounts
            WHERE sub_type ~* 'taxable'
               OR account_name ~* 'taxable|brokerage'
            GROUP BY client_id
        ),
        taxable_savings_pv AS (
            SELECT
                s.client_id,
                SUM(
                    CASE
                        WHEN NOT (s.destination ~* 'retirement|education')
                        THEN COALESCE(s.calculated_annual_amount_usd, s.fixed_amount_usd) *
                             (1 - POWER(1.0/1.04, 5)) / 0.04
                        ELSE 0
                    END
                ) AS pv_taxable_savings
            FROM core.savings s
            JOIN client_info ci ON s.client_id = ci.client_id
            GROUP BY s.client_id
        ),
        car_expenses_pv AS (
            SELECT
                e.client_id,
                SUM(
                    CASE
                        WHEN e.expense_item ~* 'car|vehicle|auto' OR
                             e.type ~* 'car|vehicle|auto' OR
                             e.sub_type ~* 'car|vehicle|auto'
                        THEN e.annual_amount *
                             (1 - POWER(1.0/1.04, 5)) / 0.04
                        ELSE 0
                    END
                ) AS pv_car_expenses
            FROM core.expenses e
            JOIN client_info ci ON e.client_id = ci.client_id
            GROUP BY e.client_id
        )
        SELECT
            ROUND(
                (
                    COALESCE(ta.taxable_account_value, 0) +
                    COALESCE(ts.pv_taxable_savings, 0)
                ) /
                NULLIF(COALESCE(ce.pv_car_expenses, 0), 0),
                2
            ) AS new_cars_ratio
        FROM client_info ci
        LEFT JOIN taxable_accounts ta ON ci.client_id = ta.client_id
        LEFT JOIN taxable_savings_pv ts ON ci.client_id = ts.client_id
        LEFT JOIN car_expenses_pv ce ON ci.client_id = ce.client_id;
    """
    return execute_metric_query(sql_query, client_id, "new_cars_ratio_for_client")

def calculate_ltc_ratio_for_client(client_id):
    """Calculate LTC ratio for specific client."""
    sql_query = """
        WITH
        client_info AS (
            SELECT
                c.client_id
            FROM core.clients c
            WHERE c.client_id = %s
        ),
        future_income_pv AS (
            SELECT
                i.client_id,
                SUM(
                    CASE
                        WHEN (i.deleted IS NULL OR i.deleted = false)
                        THEN i.annual_amount *
                             (1 - POWER(1.0/1.04, 20)) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_income
            FROM core.incomes i
            JOIN client_info ci ON i.client_id = ci.client_id
            GROUP BY i.client_id
        ),
        current_assets AS (
            SELECT
                client_id,
                SUM(total_value) AS total_assets
            FROM (
                SELECT client_id, total_value FROM core.investment_deposit_accounts
                UNION ALL
                SELECT client_id, total_value FROM core.real_estate_assets
                UNION ALL
                SELECT client_id, total_value FROM core.personal_property_accounts
            ) all_assets
            GROUP BY client_id
        ),
        future_expenses_pv AS (
            SELECT
                e.client_id,
                SUM(
                    CASE
                        WHEN NOT (e.type ~* 'ltc' OR e.expense_item ~* 'long term care')
                        THEN e.annual_amount *
                             (1 - POWER(1.0/1.04, 20)) / 0.04
                        ELSE 0
                    END
                ) AS pv_future_expenses
            FROM core.expenses e
            JOIN client_info ci ON e.client_id = ci.client_id
            GROUP BY e.client_id
        ),
        ltc_expenses_pv AS (
            SELECT
                l.client_id,
                SUM(
                    CASE
                        WHEN l.sub_type ~* 'ltc' OR l.fact_type_name ~* 'long term care'
                        THEN COALESCE(l.annual_premium, 0) *
                             (1 - POWER(1.0/1.04, 20)) / 0.04
                        ELSE 0
                    END
                ) AS pv_ltc_expenses
            FROM core.disability_ltc_insurance_accounts l
            JOIN client_info ci ON l.client_id = ci.client_id
            GROUP BY l.client_id
        )
        SELECT
            ROUND(
                (
                    COALESCE(fi.pv_future_income, 0) +
                    COALESCE(ca.total_assets, 0)
                ) /
                NULLIF(
                    (COALESCE(fe.pv_future_expenses, 0) + COALESCE(le.pv_ltc_expenses, 0)),
                    0
                ),
                2
            ) AS ltc_ratio
        FROM client_info ci
        LEFT JOIN future_income_pv fi ON ci.client_id = fi.client_id
        LEFT JOIN current_assets ca ON ci.client_id = ca.client_id
        LEFT JOIN future_expenses_pv fe ON ci.client_id = fe.client_id
        LEFT JOIN ltc_expenses_pv le ON ci.client_id = le.client_id;
    """
    return execute_metric_query(sql_query, client_id, "ltc_ratio_for_client")

def calculate_ltd_ratio_for_client(client_id):
    """Calculate LTD ratio for specific client."""
    sql_query = """
        WITH
        client_info AS (
            SELECT
                c.client_id
            FROM core.clients c
            WHERE c.client_id = %s
        ),
        ltd_value AS (
            SELECT
                client_id,
                SUM(COALESCE(benefit_amount, 0)) AS ltd_value
            FROM core.disability_ltc_insurance_accounts
            WHERE fact_type_name ~* 'disability'
            GROUP BY client_id
        ),
        earned_income AS (
            SELECT
                client_id,
                SUM(COALESCE(current_year_amount, 0)) AS earned_income
            FROM core.incomes
            WHERE income_type = 'Salary'
              AND (deleted IS NULL OR deleted = false)
            GROUP BY client_id
        )
        SELECT
            ROUND(
                COALESCE(l.ltd_value, 0) / NULLIF(COALESCE(e.earned_income, 0), 0),
                2
            ) AS ltd_ratio
        FROM client_info ci
        LEFT JOIN ltd_value l ON ci.client_id = l.client_id
        LEFT JOIN earned_income e ON ci.client_id = e.client_id;
    """
    return execute_metric_query(sql_query, client_id, "ltd_ratio_for_client")

# Account History Functions

def _get_accounts_for_client(client_id):
    """
    Shared helper to fetch account metadata for the provided client.
    """
    sql_query = """
        WITH account_info AS (
            SELECT
                account_id,
                MIN(as_of_date) as first_date,
                MAX(as_of_date) as last_date,
                COUNT(*) as record_count
            FROM core.account_history
            WHERE client_id = %s
            GROUP BY account_id
        ),
        account_values AS (
            SELECT
                ah.account_id,
                ah.value as current_value
            FROM core.account_history ah
            INNER JOIN (
                SELECT account_id, MAX(as_of_date) as max_date
                FROM core.account_history
                WHERE client_id = %s
                GROUP BY account_id
            ) latest ON ah.account_id = latest.account_id AND ah.as_of_date = latest.max_date
            WHERE ah.client_id = %s
        )
        SELECT
            ai.account_id,
            COALESCE(f.sub_type, 'Investment Account') || ' (' || ai.account_id || ')' as account_name,
            COALESCE(f.fact_type_name, 'Account Type') as account_type,
            COALESCE(av.current_value, 0) as current_value,
            TO_CHAR(ai.first_date, 'YYYY-MM-DD') as start_date,
            TO_CHAR(ai.last_date, 'YYYY-MM-DD') as end_date,
            ai.record_count as total_records
        FROM account_info ai
        LEFT JOIN account_values av ON ai.account_id = av.account_id
        LEFT JOIN core.facts f ON ai.account_id = f.fact_id AND f.client_id = %s
        ORDER BY ai.account_id;
    """

    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        import re
        param_count = len(re.findall(r'(?<!%)%s(?!%)', sql_query))
        params = tuple(client_id for _ in range(param_count))

        cursor.execute(sql_query, params)
        results = cursor.fetchall()
        cursor.close()

        accounts = []
        for row in results:
            accounts.append({
                'account_id': row[0],
                'account_name': row[1],
                'account_type': row[2],
                'current_value': float(row[3]) if row[3] is not None else 0,
                'date_range': {
                    'start_date': row[4],
                    'end_date': row[5]
                },
                'total_records': row[6]
            })

        return accounts

    except Exception as e:
        print(f"Error getting accounts for client {client_id}: {e}")
        return []
    finally:
        if connection:
            close_db_connection(connection)


def get_user_accounts():
    """
    Get all accounts for the authenticated user.
    """
    client_id = get_jwt_identity()
    return _get_accounts_for_client(client_id)


def get_user_accounts_for_admin(client_id):
    """
    Admin helper to reuse user account logic for a specified client.
    """
    return _get_accounts_for_client(client_id)


def _get_account_history_for_client(client_id, account_id, start_date=None, end_date=None, limit=100, offset=0):
    """
    Shared helper that returns account history for a specific client/account pair.
    """
    sql_query = """
        SELECT
            as_of_date,
            value
        FROM core.account_history
        WHERE client_id = %s AND account_id = %s
    """

    params = [client_id, account_id]
    if start_date:
        sql_query += " AND as_of_date >= %s"
        params.append(start_date)
    if end_date:
        sql_query += " AND as_of_date <= %s"
        params.append(end_date)

    sql_query += " ORDER BY as_of_date DESC LIMIT %s OFFSET %s"
    params.extend([limit, offset])

    count_query = """
        SELECT COUNT(*)
        FROM core.account_history
        WHERE client_id = %s AND account_id = %s
    """

    count_params = [client_id, account_id]
    if start_date:
        count_query += " AND as_of_date >= %s"
        count_params.append(start_date)
    if end_date:
        count_query += " AND as_of_date <= %s"
        count_params.append(end_date)

    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        cursor.execute(count_query, count_params)
        total_count = cursor.fetchone()[0]

        cursor.execute(sql_query, params)
        results = cursor.fetchall()
        cursor.close()

        history = []
        for row in results:
            history.append({
                'as_of_date': row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0]),
                'value': float(row[1]) if row[1] is not None else 0
            })

        has_more = offset + len(history) < total_count

        return {
            'history': history,
            'pagination': {
                'limit': limit,
                'offset': offset,
                'total': total_count,
                'has_more': has_more
            }
        }

    except Exception as e:
        print(f"Error getting account history for client {client_id}, account {account_id}: {e}")
        return {
            'history': [],
            'pagination': {
                'limit': limit,
                'offset': offset,
                'total': 0,
                'has_more': False
            }
        }
    finally:
        if connection:
            close_db_connection(connection)


def get_account_history_for_user(account_id, start_date=None, end_date=None, limit=100, offset=0):
    """
    Get historical data for a specific account for the authenticated user.
    """
    client_id = get_jwt_identity()
    return _get_account_history_for_client(client_id, account_id, start_date, end_date, limit, offset)


def get_account_history_for_admin(client_id, account_id, start_date=None, end_date=None, limit=100, offset=0):
    """
    Admin helper that returns account history for a specified client.
    """
    return _get_account_history_for_client(client_id, account_id, start_date, end_date, limit, offset)

def get_multiple_account_history(account_ids, start_date=None, end_date=None):
    """
    Get historical data for multiple accounts for the authenticated user.
    
    Args:
        account_ids (list): List of account IDs to get history for
        start_date (str, optional): Filter from date
        end_date (str, optional): Filter to date
        
    Returns:
        dict: Dictionary mapping account IDs to their history data
    """
    client_id = get_jwt_identity()
    
    if not account_ids:
        return {}
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        histories = {}
        
        # Get history for each account
        for account_id in account_ids:
            # Build the query for each account
            sql_query = """
                SELECT
                    as_of_date,
                    value
                FROM core.account_history
                WHERE client_id = %s AND account_id = %s
            """
            
            params = [client_id, account_id]
            
            # Add date filters if provided
            if start_date:
                sql_query += " AND as_of_date >= %s"
                params.append(start_date)
            if end_date:
                sql_query += " AND as_of_date <= %s"
                params.append(end_date)
            
            sql_query += " ORDER BY as_of_date DESC"
            
            cursor.execute(sql_query, params)
            results = cursor.fetchall()
            
            # Format results
            history = []
            for row in results:
                history.append({
                    'as_of_date': row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0]),
                    'value': float(row[1]) if row[1] is not None else 0
                })
            
            histories[account_id] = history
        
        cursor.close()
        return histories
        
    except Exception as e:
        print(f"Error getting multiple account histories: {e}")
        return {}
    finally:
        if connection:
            close_db_connection(connection)

def get_account_summary_for_user(account_id):
    """
    Get summary statistics for a specific account for the authenticated user.
    
    Args:
        account_id (str): The account ID to get summary for
        
    Returns:
        dict: Dictionary with account summary statistics
    """
    client_id = get_jwt_identity()
    
    sql_query = """
        SELECT
            MIN(as_of_date) as first_date,
            MAX(as_of_date) as last_date,
            MIN(value) as min_value,
            MAX(value) as max_value,
            COUNT(*) as total_records,
            AVG(value) as average_value,
            (SELECT value FROM core.account_history
             WHERE client_id = %s AND account_id = %s
             ORDER BY as_of_date DESC LIMIT 1) as current_value
        FROM core.account_history
        WHERE client_id = %s AND account_id = %s;
    """
    
    connection = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute(sql_query, (client_id, account_id, client_id, account_id))
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            return {
                'first_date': result[0].isoformat() if hasattr(result[0], 'isoformat') else str(result[0]),
                'last_date': result[1].isoformat() if hasattr(result[1], 'isoformat') else str(result[1]),
                'min_value': float(result[2]) if result[2] is not None else 0,
                'max_value': float(result[3]) if result[3] is not None else 0,
                'total_records': result[4],
                'average_value': float(result[5]) if result[5] is not None else 0,
                'current_value': float(result[6]) if result[6] is not None else 0
            }
        
        return None
        
    except Exception as e:
        print(f"Error getting account summary for {account_id}: {e}")
        return None
    finally:
        if connection:
            close_db_connection(connection)

def get_treemap_data_for_user():
    """
    Get treemap data for Equity, Cash, Real Estate, and Fixed Income for the authenticated user.
    
    Returns:
        list: List of dictionaries with category and value for treemap visualization
    """
    client_id = get_jwt_identity()
    
    sql_query = """
    -- Optimized Treemap Query
    SELECT
        'Equity' as category,
        COALESCE(h.equity_holdings, 0) + COALESCE(i.investment_equity, 0) as value
    FROM
        (SELECT client_id, COALESCE(SUM(value),0) as equity_holdings
         FROM core.holdings
         WHERE asset_class IN ('largecap', 'smallcap', 'largevalue', 'smallvalue', 'internat', 'emerging', 'ips')
             AND client_id = %s AND value IS NOT NULL
         GROUP BY client_id) h
    FULL OUTER JOIN
        (SELECT client_id, COALESCE(SUM(holdings_value),0) as investment_equity
         FROM core.investment_deposit_accounts
         WHERE fact_type_name IN ('Taxable Investment', 'Roth IRA', 'Qualified Retirement')
             AND client_id = %s AND holdings_value IS NOT NULL
         GROUP BY client_id) i ON h.client_id = i.client_id
    WHERE COALESCE(h.client_id, i.client_id) = %s

    UNION ALL

    SELECT
        'Cash',
        COALESCE(h.cash_holdings, 0) + COALESCE(i.cash_alternatives, 0)
    FROM
        (SELECT client_id, SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) as cash_holdings
         FROM core.holdings
         WHERE asset_class = 'cash' AND client_id = %s
         GROUP BY client_id) h
    FULL OUTER JOIN
        (SELECT client_id, SUM(COALESCE(cash_balance, 0)) as cash_alternatives
         FROM core.investment_deposit_accounts
         WHERE fact_type_name = 'Cash Alternative' AND client_id = %s
         GROUP BY client_id) i ON h.client_id = i.client_id
    WHERE COALESCE(h.client_id, i.client_id) = %s

    UNION ALL

    SELECT 'Real Estate', COALESCE(SUM(total_value),0)
    FROM core.real_estate_assets
    WHERE client_id = %s
    GROUP BY client_id

    UNION ALL

    SELECT 'Fixed Income', COALESCE(SUM(value),0)
    FROM core.holdings
    WHERE asset_class IN ('highyldbond', 'inttermmun', 'investbond', 'shortermbond', 'shortermmun')
        AND client_id = %s
    GROUP BY client_id

    ORDER BY category;
    """
    
    return execute_chart_query(sql_query, client_id, "treemap_data")

def get_bar_chart_data_for_user():
    """
    Get bar chart data for 9 financial ratios for the authenticated user.
    
    Returns:
        list: List of dictionaries with metric names and values for bar chart visualization
    """
    client_id = get_jwt_identity()
    
    sql_query = """
-- Unified Bar Chart Query for 9 Financial Ratios
-- Returns data in format: metric_name, metric_value

WITH
-- Client information for age-based calculations
client_info AS (
    SELECT
        client_id,
        EXTRACT(YEAR FROM AGE(CURRENT_DATE, hh_date_of_birth)) AS current_age,
        65 AS retirement_age
    FROM core.clients
    WHERE client_id = %s
),

-- Common data aggregations
income_data AS (
    SELECT
        SUM(CASE WHEN income_type = 'Salary' THEN current_year_amount ELSE 0 END) AS earned_income,
        SUM(current_year_amount) AS total_income
    FROM core.incomes
    WHERE client_id = %s AND current_year_amount IS NOT NULL
),

savings_data AS (
    SELECT SUM(calculated_annual_amount_usd) AS current_year_savings
    FROM core.savings
    WHERE client_id = %s AND start_type = 'Active'
),

giving_data AS (
    SELECT SUM(annual_amount) AS current_year_giving
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Spending'
        AND sub_type = 'GivingAndPhilanthropy'
        AND annual_amount > 0
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
),

living_expenses_data AS (
    SELECT SUM(annual_amount) AS current_year_living_expenses
    FROM core.expenses
    WHERE client_id = %s
        AND type = 'Living'
        AND annual_amount > 0
        AND EXTRACT(YEAR FROM start_actual_date) <= EXTRACT(YEAR FROM CURRENT_DATE)
        AND (end_actual_date IS NULL OR EXTRACT(YEAR FROM end_actual_date) >= EXTRACT(YEAR FROM CURRENT_DATE))
        AND (end_actual_date IS NULL OR end_actual_date >= start_actual_date)
),

cash_data AS (
    SELECT
        SUM(CASE WHEN asset_class = 'cash' THEN value ELSE 0 END) AS cash_from_holdings,
        SUM(value) AS total_holdings,
        MAX(value) AS largest_holding
    FROM core.holdings
    WHERE client_id = %s AND value IS NOT NULL AND value > 0
),

investment_cash_data AS (
    SELECT
        SUM(COALESCE(cash_balance, 0)) AS cash_from_investments,
        SUM(total_value) AS total_investments
    FROM core.investment_deposit_accounts
    WHERE client_id = %s
),

real_estate_data AS (
    SELECT
        SUM(rea.total_value) AS house_value,
        SUM(rea.total_value) - COALESCE(SUM(CASE WHEN lna.sub_type = 'Mortgage' THEN ABS(lna.total_value) ELSE 0 END), 0) AS house_equity
    FROM core.real_estate_assets rea
    LEFT JOIN core.liability_note_accounts lna ON rea.account_id = lna.real_estate_id
    WHERE rea.sub_type = 'Residence'
        AND rea.client_id = %s
        AND (lna.sub_type = 'Mortgage' OR lna.sub_type IS NULL)
    GROUP BY rea.client_id
),

life_insurance_data AS (
    SELECT SUM(death_benefit) AS life_insurance_value
    FROM core.life_insurance_annuity_accounts
    WHERE client_id = %s
        AND death_benefit IS NOT NULL 
        AND death_benefit > 0
),

disability_data AS (
    SELECT SUM(COALESCE(benefit_amount, 0)) AS ltd_value
    FROM core.disability_ltc_insurance_accounts
    WHERE client_id = %s
        AND fact_type_name ~* 'disability'
),

current_assets_data AS (
    SELECT SUM(total_value) AS current_assets
    FROM (
        SELECT total_value FROM core.investment_deposit_accounts WHERE client_id = %s
        UNION ALL
        SELECT total_value FROM core.real_estate_assets WHERE client_id = %s
        UNION ALL
        SELECT total_value FROM core.personal_property_accounts WHERE client_id = %s
    ) all_assets
),

current_liabilities_data AS (
    SELECT ABS(SUM(total_value)) AS current_liabilities
    FROM core.liability_note_accounts
    WHERE client_id = %s
),

-- Present value calculations for retirement ratio
retirement_future_income_pv AS (
    SELECT SUM(
        CASE
            WHEN (ci.retirement_age - ci.current_age) > 0 AND
                 (i.end_type IS NULL OR i.end_type != 'Age' OR (i.end_type = 'Age' AND i.end_value > ci.retirement_age))
            THEN i.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
            ELSE 0
        END
    ) AS pv_retirement_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    WHERE (i.deleted IS NULL OR i.deleted = false)
    GROUP BY ci.client_id
),

retirement_future_expenses_pv AS (
    SELECT SUM(
        CASE
            WHEN (ci.retirement_age - ci.current_age) > 0 AND
                 (e.end_type IS NULL OR e.end_type != 'Age' OR
                  (e.end_type = 'Age' AND EXTRACT(YEAR FROM AGE(e.end_actual_date, e.start_actual_date)) > (ci.retirement_age - ci.current_age)))
            THEN e.annual_amount * (1 - POWER(1.0/1.04, GREATEST(0, ci.retirement_age - ci.current_age))) / 0.04
            ELSE 0
        END
    ) AS pv_retirement_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
    GROUP BY ci.client_id
),

retirement_savings_data AS (
    SELECT SUM(COALESCE(calculated_annual_amount_usd, fixed_amount_usd)) AS retirement_savings
    FROM core.savings
    WHERE client_id = %s
        AND (destination ~* 'retirement|401k|ira' OR account_id ~* 'retirement|401k|ira')
),

-- Present value calculations for survivor ratio
future_income_survivor_pv AS (
    SELECT SUM(
        CASE
            WHEN (i.end_type = 'SpousesDeath' OR i.owner_type = 'Spouse') AND
                 (i.end_value IS NULL OR i.end_value > EXTRACT(YEAR FROM CURRENT_DATE))
            THEN i.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_survivor_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    WHERE (i.deleted IS NULL OR i.deleted = false)
),

future_expenses_survivor_pv AS (
    SELECT SUM(
        CASE
            WHEN e.end_type != 'AtSecondDeath' AND
                 (e.end_actual_date IS NULL OR e.end_actual_date > CURRENT_DATE)
            THEN e.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_survivor_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
),

-- Present value calculations for LTC ratio
future_income_ltc_pv AS (
    SELECT SUM(
        CASE
            WHEN (i.deleted IS NULL OR i.deleted = false)
            THEN i.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_ltc_income
    FROM core.incomes i
    JOIN client_info ci ON i.client_id = ci.client_id
    WHERE (i.deleted IS NULL OR i.deleted = false)
),

future_expenses_ltc_pv AS (
    SELECT SUM(
        CASE
            WHEN NOT (e.type ~* 'ltc' OR e.expense_item ~* 'long term care')
            THEN e.annual_amount * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_ltc_regular_expenses
    FROM core.expenses e
    JOIN client_info ci ON e.client_id = ci.client_id
),

ltc_expenses_data AS (
    SELECT SUM(
        CASE
            WHEN sub_type ~* 'ltc' OR fact_type_name ~* 'long term care'
            THEN COALESCE(annual_premium, 0) * (1 - POWER(1.0/1.04, 20)) / 0.04
            ELSE 0
        END
    ) AS pv_ltc_expenses
    FROM core.disability_ltc_insurance_accounts
    WHERE client_id = %s
)

-- Final ratios calculation
SELECT 'Savings Ratio' AS metric_name,
       ROUND(COALESCE(sd.current_year_savings, 0) / NULLIF(id.total_income, 0), 2) AS metric_value
FROM savings_data sd, income_data id

UNION ALL

SELECT 'Giving Ratio' AS metric_name,
       ROUND(COALESCE(gd.current_year_giving, 0) / NULLIF(id.total_income, 0), 2) AS metric_value
FROM giving_data gd, income_data id

UNION ALL

SELECT 'Reserves Ratio' AS metric_name,
       ROUND((COALESCE(cd.cash_from_holdings, 0) + COALESCE(icd.cash_from_investments, 0)) / NULLIF(led.current_year_living_expenses, 0) * 0.5, 2) AS metric_value
FROM cash_data cd, investment_cash_data icd, living_expenses_data led

UNION ALL

SELECT 'Debt Ratio' AS metric_name,
       ROUND(COALESCE(red.house_equity, 0) / NULLIF(red.house_value, 0), 2) AS metric_value
FROM real_estate_data red

UNION ALL

SELECT 'Diversification Ratio' AS metric_name,
       ROUND(1 - (COALESCE(cd.largest_holding, 0) / NULLIF((COALESCE(cd.total_holdings, 0) + COALESCE(icd.total_investments, 0)), 0)), 2) AS metric_value
FROM cash_data cd, investment_cash_data icd

UNION ALL

SELECT 'Survivor Ratio' AS metric_name,
       ROUND(
           (COALESCE(lid.life_insurance_value, 0) + COALESCE(fipv.pv_survivor_income, 0) + COALESCE(cad.current_assets, 0)) /
           NULLIF((COALESCE(fepv.pv_survivor_expenses, 0) + COALESCE(cld.current_liabilities, 0)), 0),
           2
       ) AS metric_value
FROM life_insurance_data lid, future_income_survivor_pv fipv, current_assets_data cad, future_expenses_survivor_pv fepv, current_liabilities_data cld

UNION ALL

SELECT 'Retirement Ratio' AS metric_name,
    CASE
        WHEN ci.current_age < ci.retirement_age THEN
            ROUND(
                (COALESCE(ripv.pv_retirement_income, 0) + COALESCE(cad.current_assets, 0) + COALESCE(rs.retirement_savings, 0)) /
                NULLIF((COALESCE(repv.pv_retirement_expenses, 0) + COALESCE(cld.current_liabilities, 0)), 0),
                2
            )
        ELSE NULL
    END AS metric_value
FROM client_info ci, retirement_future_income_pv ripv, current_assets_data cad, retirement_savings_data rs, retirement_future_expenses_pv repv, current_liabilities_data cld

UNION ALL

SELECT 'LTD Ratio' AS metric_name,
       ROUND(COALESCE(dd.ltd_value, 0) / NULLIF(id.earned_income, 0), 2) AS metric_value
FROM disability_data dd, income_data id

UNION ALL

SELECT 'LTC Ratio' AS metric_name,
       ROUND(
           (COALESCE(filpv.pv_ltc_income, 0) + COALESCE(cad.current_assets, 0)) /
           NULLIF((COALESCE(feipv.pv_ltc_regular_expenses, 0) + COALESCE(ltc.pv_ltc_expenses, 0)), 0),
           2
       ) AS metric_value
FROM future_income_ltc_pv filpv, current_assets_data cad, future_expenses_ltc_pv feipv, ltc_expenses_data ltc

ORDER BY metric_name;
    """
    
    return execute_chart_query(sql_query, client_id, "bar_chart_data")
