from flask import Flask, jsonify, request, current_app, Response, stream_with_context
from flask_cors import CORS
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity
from auth import login_user, register_user, register_admin_user, admin_required, superuser_required, get_admin_security_code, update_admin_security_code
from metrics import (
    calculate_net_worth_for_user,
    calculate_portfolio_value_for_user,
    calculate_real_estate_value_for_user,
    calculate_debt_for_user,
    calculate_equity_for_user,
    calculate_fixed_income_for_user,
    calculate_cash_for_user,
    calculate_earned_income_for_user,
    calculate_social_security_income_for_user,
    calculate_pension_income_for_user,
    calculate_real_estate_income_for_user,
    calculate_business_income_for_user,
    calculate_total_income_for_user,
    calculate_current_year_giving_for_user,
    calculate_current_year_savings_for_user,
    calculate_current_year_debt_for_user,
    calculate_current_year_taxes_for_user,
    calculate_current_year_living_expenses_for_user,
    calculate_total_expenses_for_user,
    calculate_margin_for_user,
    calculate_life_insurance_for_user,
    calculate_disability_for_user,
    calculate_ltc_for_user,
    calculate_umbrella_for_user,
    calculate_business_insurance_for_user,
    calculate_flood_insurance_for_user,
    calculate_at_risk_for_user,
    calculate_retirement_ratio_for_user,
    calculate_survivor_ratio_for_user,
    calculate_education_ratio_for_user,
    calculate_new_cars_ratio_for_user,
    calculate_ltc_ratio_for_user,
    calculate_ltd_ratio_for_user,
    calculate_savings_ratio_for_user,
    calculate_giving_ratio_for_user,
    calculate_reserves_ratio_for_user,
    calculate_debt_ratio_for_user,
    calculate_diversification_ratio_for_user,
    get_income_chart_data_for_user,
    get_expense_chart_data_for_user,
    get_treemap_data_for_user,
    get_bar_chart_data_for_user,
    get_client_profile,
    update_client_profile,
    get_all_user_metrics_for_export,
    get_chart_data_for_export,
    create_metrics_sheet,
    create_chart_sheet,
    get_metric_target_for_user,
    get_all_targets_for_user,
    get_all_targets_for_client,
    update_metric_target_for_user,
    update_multiple_targets_for_user,
    delete_metric_target_for_user,
    delete_all_targets_for_user,
    delete_target_for_client,
    delete_all_targets_for_client,
    compare_values,
    get_metric_with_target,
    get_metric_details,
    get_table_data_for_user,
    get_key_metrics_for_all_clients_batch,
    get_all_metrics_for_client,
    get_user_accounts,
    get_user_accounts_for_admin,
    get_account_history_for_user,
    get_account_history_for_admin,
    get_multiple_account_history,
    get_account_summary_for_user
)

# Import client name function
from metrics import get_client_name_for_user, get_client_name_by_id
from insights import get_all_user_metrics, generate_ai_insights, generate_financial_summary, generate_ai_insights_stream
import os
import json
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Initialize Flask application
app = Flask(__name__)

# Configure JWT
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'your-super-secret-jwt-key-change-this-in-production')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = 3600  # 1 hour

# Initialize JWT manager
jwt = JWTManager(app)

# Enable CORS for cross-origin requests
# Configure CORS for specific origins in production
if os.getenv('ENVIRONMENT') == 'production':
    # In production, allow your Vercel frontend domain
    frontend_url = os.getenv('FRONTEND_URL', 'https://wisdom-index-frontend.vercel.app')
    # Remove trailing slash if present
    frontend_url = frontend_url.rstrip('/') if frontend_url else frontend_url
    CORS(app, origins=[frontend_url, 'https://vercel.com'])
else:
    # In development, allow localhost
    CORS(app, origins=['http://localhost:3000', 'http://127.0.0.1:3000'])

@app.route('/auth/login', methods=['POST'])
def login():
    """
    Authenticate user and return JWT token.
    
    Request Body:
        {
            "username": "string",
            "password": "string"
        }
        
    Returns:
        JSON response with access token or error message
    """
    try:
        # Extract credentials from request
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        # Validate required parameters
        if not username or not password:
            return jsonify({"error": "Username and password are required"}), 400
        
        # Attempt to login user
        response_data, status_code = login_user(username, password)
        
        if status_code == 200:
            # Check if user is admin and add to response
            user_info = response_data.get('user', {})
            user_id = user_info.get('user_id')  # Get the actual user_id
            
            if user_id:
                from auth import is_admin_user
                is_admin = is_admin_user(int(user_id))
                response_data['user']['isAdmin'] = is_admin
        
        return jsonify(response_data), status_code
        
    except Exception as e:
        print(f"Error in /auth/login: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/auth/register', methods=['POST'])
def register():
    """
    Register a new user if they exist in the core.clients table.
    
    Request Body:
        {
            "first_name": "string",
            "last_name": "string",
            "email": "string",
            "username": "string",
            "password": "string"
        }
        
    Returns:
        JSON response with success or error message
    """
    try:
        # Extract registration data from request
        data = request.get_json()
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        email = data.get('email')
        username = data.get('username')
        password = data.get('password')
        
        # Validate required parameters
        if not all([first_name, last_name, email, username, password]):
            return jsonify({"error": "All fields are required: first_name, last_name, email, username, password"}), 400
        
        # Attempt to register user
        response_data, status_code = register_user(first_name, last_name, email, username, password)
        return jsonify(response_data), status_code
        
    except Exception as e:
        print(f"Error in /auth/register: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/auth/admin/register', methods=['POST'])
def admin_register():
    """
    Register a new admin user with security code validation.
    
    Request Body:
        {
            "first_name": "Admin",
            "last_name": "User",
            "email": "admin@wisdomindex.com",
            "password": "securePassword123!",
            "security_code": "WisdomAdmin2025!"
        }
    """
    try:
        data = request.get_json()
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        email = data.get('email')
        password = data.get('password')
        security_code = data.get('security_code')
        
        # Validate required fields
        if not all([first_name, last_name, email, password, security_code]):
            return jsonify({"error": "All fields are required"}), 400
        
        # Attempt to register admin user
        response_data, status_code = register_admin_user(first_name, last_name, email, password, security_code)
        return jsonify(response_data), status_code
        
    except Exception as e:
        print(f"Error in admin registration: {e}")
        return jsonify({"error": "Registration failed"}), 500

@app.route('/api/auth/admin/check-registration', methods=['GET'])
def check_admin_registration():
    """
    Check if admin registration is currently allowed.
    This can be used to enable/disable admin registration.
    """
    # In production, you might want to check if any admins exist
    # or have other logic to control when registration is allowed
    
    return jsonify({
        "admin_registration_allowed": True,
        "message": "Admin registration is currently enabled"
    })

@app.route('/auth/verify', methods=['GET'])
@jwt_required()
def verify_auth():
    """
    Verify JWT token and return user information including admin status.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with user information or error message
    """
    try:
        # Get user ID from JWT token
        user_id = get_jwt_identity()
        
        if not user_id:
            return jsonify({"error": "Invalid token"}), 401
        
        # Get user information from database
        from database import get_db_connection, close_db_connection
        from auth import is_admin_user
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        cursor.execute("""
            SELECT user_id, client_id, username, email, "isAdmin"
            FROM core.users
            WHERE user_id = %s
        """, (user_id,))
        
        user_data = cursor.fetchone()
        cursor.close()
        
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        # Check admin status
        is_admin = is_admin_user(int(user_id))
        
        return jsonify({
            "user": {
                "user_id": user_data[0],
                "client_id": user_data[1],
                "username": user_data[2],
                "email": user_data[3],
                "isAdmin": is_admin
            },
            "valid": True
        })
        
    except Exception as e:
        print(f"Error in /auth/verify: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/metrics/net-worth', methods=['GET'])
@jwt_required()
def get_net_worth():
    """
    API endpoint to fetch the Net Worth metric with target comparison for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with net_worth value, target, and status or error message
    """
    try:
        # Calculate net worth for the authenticated user
        # The client_id is extracted from the JWT token automatically
        net_worth = calculate_net_worth_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('net-worth', net_worth)
        
        # Return successful response with target information
        return jsonify({
            "metric": "net_worth",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/metrics/net-worth: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/portfolio-value', methods=['GET'])
@jwt_required()
def get_portfolio_value():
    """
    API endpoint to fetch the Portfolio Value metric for the authenticated user.
    """
    try:
        portfolio_value = calculate_portfolio_value_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('portfolio-value', portfolio_value)
        
        return jsonify({
            "metric": "portfolio_value",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/portfolio-value: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/real-estate-value', methods=['GET'])
@jwt_required()
def get_real_estate_value():
    """
    API endpoint to fetch the Real Estate Value metric for the authenticated user.
    """
    try:
        real_estate_value = calculate_real_estate_value_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('real-estate-value', real_estate_value)
        
        return jsonify({
            "metric": "real_estate_value",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/real-estate-value: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/debt', methods=['GET'])
@jwt_required()
def get_debt():
    """
    API endpoint to fetch the Debt metric for the authenticated user.
    """
    try:
        debt = calculate_debt_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('debt', debt)
        
        return jsonify({
            "metric": "debt",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/debt: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/equity', methods=['GET'])
@jwt_required()
def get_equity():
    """
    API endpoint to fetch the Equity metric for the authenticated user.
    """
    try:
        equity = calculate_equity_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('equity', equity)
        
        return jsonify({
            "metric": "equity",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/equity: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/fixed-income', methods=['GET'])
@jwt_required()
def get_fixed_income():
    """
    API endpoint to fetch the Fixed Income metric for the authenticated user.
    """
    try:
        fixed_income = calculate_fixed_income_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('fixed-income', fixed_income)
        
        return jsonify({
            "metric": "fixed_income",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/fixed-income: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/cash', methods=['GET'])
@jwt_required()
def get_cash():
    """
    API endpoint to fetch the Cash metric for the authenticated user.
    """
    try:
        cash = calculate_cash_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('cash', cash)
        
        return jsonify({
            "metric": "cash",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/cash: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/earned-income', methods=['GET'])
@jwt_required()
def get_earned_income():
    """
    API endpoint to fetch the Earned Income metric for the authenticated user.
    """
    try:
        earned_income = calculate_earned_income_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('earned-income', earned_income)
        
        return jsonify({
            "metric": "earned_income",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/earned-income: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/social-security-income', methods=['GET'])
@jwt_required()
def get_social_security_income():
    """
    API endpoint to fetch the Social Security Income metric for the authenticated user.
    """
    try:
        social_security_income = calculate_social_security_income_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('social-security-income', social_security_income)
        
        return jsonify({
            "metric": "social_security_income",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/social-security-income: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/pension-income', methods=['GET'])
@jwt_required()
def get_pension_income():
    """
    API endpoint to fetch the Pension Income metric for the authenticated user.
    """
    try:
        pension_income = calculate_pension_income_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('pension-income', pension_income)
        
        return jsonify({
            "metric": "pension_income",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/pension-income: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/real-estate-income', methods=['GET'])
@jwt_required()
def get_real_estate_income():
    """
    API endpoint to fetch the Real Estate Income metric for the authenticated user.
    """
    try:
        real_estate_income = calculate_real_estate_income_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('real-estate-income', real_estate_income)
        
        return jsonify({
            "metric": "real_estate_income",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/real-estate-income: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/business-income', methods=['GET'])
@jwt_required()
def get_business_income():
    """
    API endpoint to fetch the Business Income metric for the authenticated user.
    """
    try:
        business_income = calculate_business_income_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('business-income', business_income)
        
        return jsonify({
            "metric": "business_income",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/business-income: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/total-income', methods=['GET'])
@jwt_required()
def get_total_income():
    """
    API endpoint to fetch the Total Income metric for the authenticated user.
    """
    try:
        total_income = calculate_total_income_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('total-income', total_income)
        
        return jsonify({
            "metric": "total_income",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/total-income: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/current-year-giving', methods=['GET'])
@jwt_required()
def get_current_year_giving():
    """
    API endpoint to fetch the Current Year Giving metric for the authenticated user.
    """
    try:
        current_year_giving = calculate_current_year_giving_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('current-year-giving', current_year_giving)
        
        return jsonify({
            "metric": "current_year_giving",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/current-year-giving: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/current-year-savings', methods=['GET'])
@jwt_required()
def get_current_year_savings():
    """
    API endpoint to fetch the Current Year Savings metric for the authenticated user.
    """
    try:
        current_year_savings = calculate_current_year_savings_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('current-year-savings', current_year_savings)
        
        return jsonify({
            "metric": "current_year_savings",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/current-year-savings: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/current-year-debt', methods=['GET'])
@jwt_required()
def get_current_year_debt():
    """
    API endpoint to fetch the Current Year Debt metric for the authenticated user.
    """
    try:
        current_year_debt = calculate_current_year_debt_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('current-year-debt', current_year_debt)
        
        return jsonify({
            "metric": "current_year_debt",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/current-year-debt: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/current-year-taxes', methods=['GET'])
@jwt_required()
def get_current_year_taxes():
    """
    API endpoint to fetch the Current Year Taxes metric for the authenticated user.
    """
    try:
        current_year_taxes = calculate_current_year_taxes_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('current-year-taxes', current_year_taxes)
        
        return jsonify({
            "metric": "current_year_taxes",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/current-year-taxes: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/current-year-living-expenses', methods=['GET'])
@jwt_required()
def get_current_year_living_expenses():
    """
    API endpoint to fetch the Current Year Living Expenses metric for the authenticated user.
    """
    try:
        current_year_living_expenses = calculate_current_year_living_expenses_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('current-year-living-expenses', current_year_living_expenses)
        
        return jsonify({
            "metric": "current_year_living_expenses",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/current-year-living-expenses: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/total-expenses', methods=['GET'])
@jwt_required()
def get_total_expenses():
    """
    API endpoint to fetch the Total Expenses metric for the authenticated user.
    """
    try:
        total_expenses = calculate_total_expenses_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('total-expenses', total_expenses)
        
        return jsonify({
            "metric": "total_expenses",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/total-expenses: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/margin', methods=['GET'])
@jwt_required()
def get_margin():
    """
    API endpoint to fetch the Margin metric for the authenticated user.
    """
    try:
        margin = calculate_margin_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('margin', margin)
        
        return jsonify({
            "metric": "margin",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/margin: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/life-insurance', methods=['GET'])
@jwt_required()
def get_life_insurance():
    """
    API endpoint to fetch the Life Insurance metric for the authenticated user.
    """
    try:
        life_insurance = calculate_life_insurance_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('life-insurance', life_insurance)
        
        return jsonify({
            "metric": "life_insurance",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/life-insurance: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/disability', methods=['GET'])
@jwt_required()
def get_disability():
    """
    API endpoint to fetch the Disability metric for the authenticated user.
    """
    try:
        disability = calculate_disability_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('disability', disability)
        
        return jsonify({
            "metric": "disability",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/disability: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/ltc', methods=['GET'])
@jwt_required()
def get_ltc():
    """
    API endpoint to fetch the LTC (Long Term Care) metric for the authenticated user.
    """
    try:
        ltc = calculate_ltc_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('ltc', ltc)
        
        return jsonify({
            "metric": "ltc",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/ltc: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/umbrella', methods=['GET'])
@jwt_required()
def get_umbrella():
    """
    API endpoint to fetch the Umbrella metric for the authenticated user.
    """
    try:
        umbrella = calculate_umbrella_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('umbrella', umbrella)
        
        return jsonify({
            "metric": "umbrella",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/umbrella: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/business-insurance', methods=['GET'])
@jwt_required()
def get_business_insurance():
    """
    API endpoint to fetch the Business Insurance metric for the authenticated user.
    """
    try:
        business_insurance = calculate_business_insurance_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('business-insurance', business_insurance)
        
        return jsonify({
            "metric": "business_insurance",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/business-insurance: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/flood-insurance', methods=['GET'])
@jwt_required()
def get_flood_insurance():
    """
    API endpoint to fetch the Flood Insurance metric for the authenticated user.
    """
    try:
        flood_insurance = calculate_flood_insurance_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('flood-insurance', flood_insurance)
        
        return jsonify({
            "metric": "flood_insurance",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/flood-insurance: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/at-risk', methods=['GET'])
@jwt_required()
def get_at_risk():
    """
    API endpoint to fetch the At Risk metric for the authenticated user.
    """
    try:
        at_risk = calculate_at_risk_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('at-risk', at_risk)
        
        return jsonify({
            "metric": "at_risk",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/at-risk: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/retirement-ratio', methods=['GET'])
@jwt_required()
def get_retirement_ratio():
    """
    API endpoint to fetch the Retirement Ratio metric for the authenticated user.
    """
    try:
        retirement_ratio = calculate_retirement_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('retirement-ratio', retirement_ratio)
        
        return jsonify({
            "metric": "retirement_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/retirement-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/survivor-ratio', methods=['GET'])
@jwt_required()
def get_survivor_ratio():
    """
    API endpoint to fetch the Survivor Ratio metric for the authenticated user.
    """
    try:
        survivor_ratio = calculate_survivor_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('survivor-ratio', survivor_ratio)
        
        return jsonify({
            "metric": "survivor_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/survivor-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/education-ratio', methods=['GET'])
@jwt_required()
def get_education_ratio():
    """
    API endpoint to fetch the Education Ratio metric for the authenticated user.
    """
    try:
        education_ratio = calculate_education_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('education-ratio', education_ratio)
        
        return jsonify({
            "metric": "education_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/education-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/new-cars-ratio', methods=['GET'])
@jwt_required()
def get_new_cars_ratio():
    """
    API endpoint to fetch the New Cars Ratio metric for the authenticated user.
    """
    try:
        new_cars_ratio = calculate_new_cars_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('new-cars-ratio', new_cars_ratio)
        
        return jsonify({
            "metric": "new_cars_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/new-cars-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/ltc-ratio', methods=['GET'])
@jwt_required()
def get_ltc_ratio():
    """
    API endpoint to fetch the LTC Ratio metric for the authenticated user.
    """
    try:
        ltc_ratio = calculate_ltc_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('ltc-ratio', ltc_ratio)
        
        return jsonify({
            "metric": "ltc_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/ltc-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/ltd-ratio', methods=['GET'])
@jwt_required()
def get_ltd_ratio():
    """
    API endpoint to fetch the LTD Ratio metric for the authenticated user.
    """
    try:
        ltd_ratio = calculate_ltd_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('ltd-ratio', ltd_ratio)
        
        return jsonify({
            "metric": "ltd_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/ltd-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/savings-ratio', methods=['GET'])
@jwt_required()
def get_savings_ratio():
    """
    API endpoint to fetch the Savings Ratio metric for the authenticated user.
    """
    try:
        savings_ratio = calculate_savings_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('savings-ratio', savings_ratio)
        
        return jsonify({
            "metric": "savings_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/savings-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/giving-ratio', methods=['GET'])
@jwt_required()
def get_giving_ratio():
    """
    API endpoint to fetch the Giving Ratio metric for the authenticated user.
    """
    try:
        giving_ratio = calculate_giving_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('giving-ratio', giving_ratio)
        
        return jsonify({
            "metric": "giving_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/giving-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/reserves-ratio', methods=['GET'])
@jwt_required()
def get_reserves_ratio():
    """
    API endpoint to fetch the Reserves Ratio metric for the authenticated user.
    """
    try:
        reserves_ratio = calculate_reserves_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('reserves-ratio', reserves_ratio)
        
        return jsonify({
            "metric": "reserves_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/reserves-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/debt-ratio', methods=['GET'])
@jwt_required()
def get_debt_ratio():
    """
    API endpoint to fetch the Debt Ratio metric for the authenticated user.
    """
    try:
        debt_ratio = calculate_debt_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('debt-ratio', debt_ratio)
        
        return jsonify({
            "metric": "debt_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/debt-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/metrics/diversification-ratio', methods=['GET'])
@jwt_required()
def get_diversification_ratio():
    """
    API endpoint to fetch the Diversification Ratio metric for the authenticated user.
    """
    try:
        diversification_ratio = calculate_diversification_ratio_for_user()
        user_id = get_jwt_identity()
        
        # Get metric data with target information
        metric_data = get_metric_with_target('diversification-ratio', diversification_ratio)
        
        return jsonify({
            "metric": "diversification_ratio",
            "value": metric_data['value'],
            "target": metric_data['target'],
            "status": metric_data['target_status'],
            "target_percentage": metric_data['target_percentage'],
            "target_display_text": metric_data['target_display_text'],
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/metrics/diversification-ratio: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/charts/income-bar-chart', methods=['GET'])
@jwt_required()
def get_income_bar_chart():
    """
    API endpoint to fetch income data for bar chart visualization for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with income categories and amounts or error message
    """
    try:
        # Get income chart data for the authenticated user
        # The client_id is extracted from the JWT token automatically
        income_data = get_income_chart_data_for_user()
        
        # Get the user ID from the JWT token for response
        user_id = get_jwt_identity()
        
        # Return successful response
        return jsonify({
            "chart_type": "income_bar_chart",
            "data": income_data,
            "user_id": user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/charts/income-bar-chart: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/charts/expense-pie-chart', methods=['GET'])
@jwt_required()
def get_expense_pie_chart():
    """
    API endpoint to fetch expense data for pie chart visualization for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with expense categories and amounts or error message
    """
    try:
        # Get expense chart data for the authenticated user
        # The client_id is extracted from the JWT token automatically
        expense_data = get_expense_chart_data_for_user()
        
        # Get the user ID from the JWT token for response
        user_id = get_jwt_identity()
        
        # Return successful response
        return jsonify({
            "chart_type": "expense_pie_chart",
            "data": expense_data,
            "user_id": user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/charts/expense-pie-chart: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/charts/treemap', methods=['GET'])
@jwt_required()
def get_treemap_chart():
    """
    API endpoint to fetch treemap data for visualization for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with treemap data (categories and values) or error message
    """
    try:
        # Get treemap data for the authenticated user
        # The client_id is extracted from the JWT token automatically
        treemap_data = get_treemap_data_for_user()
        
        # Get the user ID from the JWT token for response
        user_id = get_jwt_identity()
        
        response_data = {
            "chart_type": "treemap",
            "data": treemap_data,
            "user_id": user_id
        }
        
        # Return successful response
        return jsonify(response_data)
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/charts/treemap: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/charts/bar-chart', methods=['GET'])
@jwt_required()
def get_bar_chart_data():
    """
    API endpoint to fetch bar chart data for 9 financial ratios for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with financial ratios and values for bar chart visualization
    """
    try:
        # Get bar chart data for the authenticated user
        # The client_id is extracted from the JWT token automatically
        bar_chart_data = get_bar_chart_data_for_user()
        
        # Get the user ID from the JWT token for response
        user_id = get_jwt_identity()
        
        # Return successful response
        return jsonify({
            "chart_type": "bar_chart",
            "data": bar_chart_data,
            "user_id": user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/charts/bar-chart: {e}")
        return jsonify({"error": "Internal server error"}), 500

# Admin Chart Data API Endpoints

@app.route('/api/admin/client/<int:client_id>/charts/income-bar-chart', methods=['GET'])
@jwt_required()
@admin_required
def get_income_bar_chart_admin(client_id):
    """
    API endpoint to fetch income data for bar chart visualization for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to get chart data for
        
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with income categories and amounts or error message
    """
    try:
        from metrics import get_income_chart_data_for_client
        
        # Get income chart data for the specified client
        income_data = get_income_chart_data_for_client(client_id)
        
        # Get the admin user ID from the JWT token for audit logging
        admin_user_id = get_jwt_identity()
        
        # Return successful response
        return jsonify({
            "chart_type": "income_bar_chart",
            "data": income_data,
            "client_id": client_id,
            "accessed_by_admin": admin_user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/admin/client/{client_id}/charts/income-bar-chart: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/client/<int:client_id>/charts/expense-pie-chart', methods=['GET'])
@jwt_required()
@admin_required
def get_expense_pie_chart_admin(client_id):
    """
    API endpoint to fetch expense data for pie chart visualization for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to get chart data for
        
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with expense categories and amounts or error message
    """
    try:
        from metrics import get_expense_chart_data_for_client
        
        # Get expense chart data for the specified client
        expense_data = get_expense_chart_data_for_client(client_id)
        
        # Get the admin user ID from the JWT token for audit logging
        admin_user_id = get_jwt_identity()
        
        # Return successful response
        return jsonify({
            "chart_type": "expense_pie_chart",
            "data": expense_data,
            "client_id": client_id,
            "accessed_by_admin": admin_user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/admin/client/{client_id}/charts/expense-pie-chart: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/client/<int:client_id>/charts/treemap', methods=['GET'])
@jwt_required()
@admin_required
def get_treemap_chart_admin(client_id):
    """
    API endpoint to fetch treemap data for visualization for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to get chart data for
        
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with treemap data (categories and values) or error message
    """
    try:
        from metrics import get_treemap_chart_data_for_client

        # Get treemap data for the specified client
        treemap_data = get_treemap_chart_data_for_client(client_id)
        
        # Get the admin user ID from the JWT token for audit logging
        admin_user_id = get_jwt_identity()
        
        response_data = {
            "chart_type": "treemap",
            "data": treemap_data,
            "client_id": client_id,
            "accessed_by_admin": admin_user_id
        }
        
        # Return successful response
        return jsonify(response_data)
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/admin/client/{client_id}/charts/treemap: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/client/<int:client_id>/charts/bar-chart', methods=['GET'])
@jwt_required()
@admin_required
def get_bar_chart_data_admin(client_id):
    """
    API endpoint to fetch bar chart data for 9 financial ratios for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to get chart data for
        
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with financial ratios and values for bar chart visualization
    """
    try:
        from metrics import get_wisdom_index_chart_data_for_client

        # Get bar chart data for the specified client
        bar_chart_data = get_wisdom_index_chart_data_for_client(client_id)
        
        # Get the admin user ID from the JWT token for audit logging
        admin_user_id = get_jwt_identity()
        
        # Return successful response
        return jsonify({
            "chart_type": "bar_chart",
            "data": bar_chart_data,
            "client_id": client_id,
            "accessed_by_admin": admin_user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/admin/client/{client_id}/charts/bar-chart: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/clients/compare/charts/treemap', methods=['GET'])
@jwt_required()
@admin_required
def get_treemap_chart_comparison():
    """
    API endpoint to fetch and compare treemap data for two clients (admin access).
    """
    try:
        client1_id = request.args.get('client1_id')
        client2_id = request.args.get('client2_id')
        
        if not client1_id or not client2_id:
            return jsonify({"error": "Both client1_id and client2_id are required"}), 400
            
        from metrics import get_treemap_chart_data_for_client
        
        data1 = get_treemap_chart_data_for_client(client1_id)
        data2 = get_treemap_chart_data_for_client(client2_id)
        
        name1 = get_client_name_by_id(client1_id)
        name2 = get_client_name_by_id(client2_id)
        
        return jsonify({
            "chart_type": "treemap_comparison",
            "client1": {
                "client_id": client1_id,
                "client_name": name1,
                "data": data1
            },
            "client2": {
                "client_id": client2_id,
                "client_name": name2,
                "data": data2
            }
        })
    except Exception as e:
        print(f"Error in treemap comparison: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/clients/compare/charts/bar-chart', methods=['GET'])
@jwt_required()
@admin_required
def get_bar_chart_comparison():
    """
    API endpoint to fetch and compare Wisdom Index bar chart data for two clients.
    """
    try:
        client1_id = request.args.get('client1_id')
        client2_id = request.args.get('client2_id')
        
        if not client1_id or not client2_id:
            return jsonify({"error": "Both client1_id and client2_id are required"}), 400
            
        from metrics import get_wisdom_index_chart_data_for_client
        
        data1 = get_wisdom_index_chart_data_for_client(client1_id)
        data2 = get_wisdom_index_chart_data_for_client(client2_id)
        
        name1 = get_client_name_by_id(client1_id)
        name2 = get_client_name_by_id(client2_id)
        
        return jsonify({
            "chart_type": "bar_chart_comparison",
            "client1": {
                "client_id": client1_id,
                "client_name": name1,
                "data": data1
            },
            "client2": {
                "client_id": client2_id,
                "client_name": name2,
                "data": data2
            }
        })
    except Exception as e:
        print(f"Error in bar chart comparison: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/clients/compare/charts/income-bar-chart', methods=['GET'])
@jwt_required()
@admin_required
def get_income_chart_comparison():
    """
    API endpoint to fetch and compare income bar chart data for two clients.
    """
    try:
        client1_id = request.args.get('client1_id')
        client2_id = request.args.get('client2_id')
        
        if not client1_id or not client2_id:
            return jsonify({"error": "Both client1_id and client2_id are required"}), 400
            
        from metrics import get_income_chart_data_for_client
        
        data1 = get_income_chart_data_for_client(client1_id)
        data2 = get_income_chart_data_for_client(client2_id)
        
        name1 = get_client_name_by_id(client1_id)
        name2 = get_client_name_by_id(client2_id)
        
        return jsonify({
            "chart_type": "income_bar_chart_comparison",
            "client1": {
                "client_id": client1_id,
                "client_name": name1,
                "data": data1
            },
            "client2": {
                "client_id": client2_id,
                "client_name": name2,
                "data": data2
            }
        })
    except Exception as e:
        print(f"Error in income chart comparison: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/clients/compare/charts/expense-pie-chart', methods=['GET'])
@jwt_required()
@admin_required
def get_expense_chart_comparison():
    """
    API endpoint to fetch and compare expense pie chart data for two clients.
    """
    try:
        client1_id = request.args.get('client1_id')
        client2_id = request.args.get('client2_id')
        
        if not client1_id or not client2_id:
            return jsonify({"error": "Both client1_id and client2_id are required"}), 400
            
        from metrics import get_expense_chart_data_for_client
        
        data1 = get_expense_chart_data_for_client(client1_id)
        data2 = get_expense_chart_data_for_client(client2_id)
        
        name1 = get_client_name_by_id(client1_id)
        name2 = get_client_name_by_id(client2_id)
        
        return jsonify({
            "chart_type": "expense_pie_chart_comparison",
            "client1": {
                "client_id": client1_id,
                "client_name": name1,
                "data": data1
            },
            "client2": {
                "client_id": client2_id,
                "client_name": name2,
                "data": data2
            }
        })
    except Exception as e:
        print(f"Error in expense chart comparison: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/profile', methods=['GET'])
@jwt_required()
def get_profile():
    """
    API endpoint to fetch client profile information for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with client profile data or error message
    """
    try:
        # Get client profile data for the authenticated user
        profile_data = get_client_profile()
        
        # Get the user ID from the JWT token for response
        user_id = get_jwt_identity()
        
        # Return successful response
        return jsonify({
            "profile": profile_data,
            "user_id": user_id
        })
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/profile: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/profile', methods=['PUT'])
@jwt_required()
def update_profile():
    """
    API endpoint to update client profile information for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Request Body:
        {
            "first_name": "John",
            "last_name": "Doe",
            "date_of_birth": "1980-01-01",
            "gender": "Male",
            "marital_status": "Married",
            "citizenship": "US",
            "spouse_first_name": "Jane",
            "spouse_last_name": "Doe",
            "spouse_date_of_birth": "1982-05-15",
            "address1": "123 Main St",
            "city": "Anytown",
            "state": "CA",
            "postal_code": "12345",
            "home_phone": "555-123-4567",
            "business_phone": "555-987-6543",
            "cell_phone": "555-555-5555",
            "employer_name": "Acme Corp",
            "job_title": "Manager",
            "years_employed": 10
        }
        
    Returns:
        JSON response with updated profile data or error message
    """
    try:
        # Get the profile data from the request
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "No profile data provided"}), 400
        
        # Update client profile data for the authenticated user
        updated_profile = update_client_profile(data)
        
        if updated_profile:
            # Get the user ID from the JWT token for response
            user_id = get_jwt_identity()
            
            # Return successful response with updated profile
            return jsonify({
                "profile": updated_profile,
                "message": "Profile updated successfully",
                "user_id": user_id
            })
        else:
            return jsonify({"error": "Failed to update profile"}), 500
        
    except ValueError as e:
        # Handle validation errors
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/profile PUT: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/client-name', methods=['GET'])
@jwt_required()
def get_client_name():
    """
    API endpoint to fetch the client name for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Returns:
        JSON response with client name or error message
    """
    try:
        # Get client name for the authenticated user
        client_name = get_client_name_for_user()
        user_id = get_jwt_identity()
        
        if client_name:
            # Return successful response
            return jsonify({
                "client_name": client_name,
                "user_id": user_id
            })
        else:
            return jsonify({"error": "Client name not found"}), 404
        
    except Exception as e:
        # Handle database and other errors
        print(f"Error in /api/client-name: {e}")
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/dashboard/summary', methods=['GET'])
@jwt_required()
def get_dashboard_summary():
    """
    Return a fully-hydrated dashboard payload in a single request.
    Bundles client info, metrics, targets, and charts to eliminate 30+ round trips.
    """
    try:
        client_id = get_jwt_identity()
        client_name = get_client_name_for_user() or "Client"
        metrics_by_category = get_all_metrics_for_client(client_id) or {}
        targets = get_all_targets_for_user()
        income_chart = get_income_chart_data_for_user()
        expense_chart = get_expense_chart_data_for_user()

        def snake_to_kebab(value: str) -> str:
            return value.replace('_', '-') if isinstance(value, str) else value

        flat_metrics = {}
        for category_data in metrics_by_category.values():
            if not isinstance(category_data, dict):
                continue
            for metric_key, metric_value in category_data.items():
                flat_metrics[snake_to_kebab(metric_key)] = metric_value

        return jsonify({
            "client_name": client_name,
            "metrics": flat_metrics,
            "metrics_by_category": metrics_by_category,
            "targets": targets,
            "charts": {
                "income": income_chart,
                "expense": expense_chart
            }
        })
    except Exception as e:
        print(f"Error in /api/dashboard/summary: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/insights/generate', methods=['POST'])
@jwt_required()
def generate_insights():
    """
    API endpoint to generate AI-powered financial insights for the authenticated user.
    
    Headers:
        Authorization: Bearer <jwt_token>
        
    Request Body (optional):
        {
            "include_summary": true,  // Include a brief financial summary
            "metrics_data": {         // Optional: Pre-calculated metrics from frontend
                "assets_and_liabilities": { ... },
                "income_analysis": { ... },
                ...
            },
            "stream": true            // Optional: Stream the response
        }
        
    Returns:
        JSON response with AI-generated insights or error message, or a streamed response
    """
    try:
        # Get the user ID from the JWT token
        user_id = get_jwt_identity()
        
        # Get request parameters
        try:
            data = request.get_json() or {}
        except Exception:
            data = {}
        
        print(f"DEBUG: /api/insights/generate received data keys: {list(data.keys())}")
        
        include_summary = data.get('include_summary', False)
        metrics_data_from_frontend = data.get('metrics_data')
        
        # Handle stream flag flexibly (boolean or string)
        stream_val = data.get('stream', False)
        stream_response = str(stream_val).lower() == 'true' if isinstance(stream_val, str) else bool(stream_val)
        
        print(f"DEBUG: Generating insights for user {user_id}, stream={stream_response}")
        
        # Use metrics from frontend if provided, otherwise fetch from database
        if metrics_data_from_frontend:
            print(f"DEBUG: Using pre-calculated metrics from frontend")
            metrics_data = metrics_data_from_frontend
            # Add client_id for consistency
            metrics_data['client_id'] = user_id
        else:
            print(f"DEBUG: Fetching metrics from database (fallback)")
            # Get all user metrics for AI analysis (original behavior)
            metrics_data = get_all_user_metrics()
            
        if stream_response:
            def generate():
                # Generator for streaming response
                for chunk in generate_ai_insights_stream(metrics_data):
                    yield chunk
            
            return Response(
                stream_with_context(generate()), 
                mimetype='text/plain; charset=utf-8',
                headers={"X-Accel-Buffering": "no"}
            )
        
        # Generate AI insights (non-streaming)
        insights = generate_ai_insights(metrics_data)
        
        # Prepare response data
        response_data = {
            "insights": insights,
            "user_id": user_id,
            "timestamp": "2025-10-30T06:30:00Z",  # Current timestamp
            "used_cached_metrics": metrics_data_from_frontend is not None
        }
        
        # Include financial summary if requested
        if include_summary:
            summary = generate_financial_summary(metrics_data)
            response_data["summary"] = summary
        
        print(f"DEBUG: Successfully generated insights for user {user_id}")
        
        # Return successful response
        return jsonify(response_data)
        
    except Exception as e:
        # Handle errors gracefully
        print(f"Error generating insights: {e}")
        
        # Return a user-friendly error message
        return jsonify({
            "error": "Unable to generate insights at this time",
            "message": "Please try again later or contact your financial advisor",
            "user_id": user_id if 'user_id' in locals() else None
        }), 500

@app.route('/api/admin/client/<int:client_id>/insights/generate', methods=['POST'])
@jwt_required()
@admin_required
def generate_insights_for_client_admin(client_id):
    """
    API endpoint to generate AI-powered financial insights for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to generate insights for
        
    Headers:
        Authorization: Bearer <jwt_token>
        
    Request Body (optional):
        {
            "include_summary": true  // Include a brief financial summary
        }
        
    Returns:
        JSON response with AI-generated insights or error message
    """
    try:
        # Get the admin user ID from the JWT token for audit logging
        admin_user_id = get_jwt_identity()
        
        # Get request parameters
        try:
            data = request.get_json() or {}
        except Exception:
            data = {}
        
        include_summary = data.get('include_summary', False)
        # Handle stream flag flexibly
        stream_val = data.get('stream', False)
        stream_response = str(stream_val).lower() == 'true' if isinstance(stream_val, str) else bool(stream_val)
        
        print(f"DEBUG: Admin {admin_user_id} generating insights for client {client_id}, stream={stream_response}")
        
        # Get all metrics for the specified client
        from metrics import get_all_metrics_for_client
        metrics_data = get_all_metrics_for_client(client_id)
        
        if not metrics_data:
            return jsonify({
                "error": "No financial data available for this client",
                "client_id": client_id
            }), 404
        
        # Add client_id to metrics data for consistency
        metrics_data['client_id'] = client_id
        
        if stream_response:
            def generate():
                # Generator for streaming response
                for chunk in generate_ai_insights_stream(metrics_data):
                    yield chunk
            
            return Response(
                stream_with_context(generate()), 
                mimetype='text/plain; charset=utf-8',
                headers={"X-Accel-Buffering": "no"}
            )
        
        # Generate AI insights using the same function as client-side
        insights = generate_ai_insights(metrics_data)
        
        # Prepare response data
        response_data = {
            "insights": insights,
            "client_id": client_id,
            "generated_by_admin": admin_user_id,
            "timestamp": "2025-10-30T06:30:00Z",  # Current timestamp
        }
        
        # Include financial summary if requested
        if include_summary:
            summary = generate_financial_summary(metrics_data)
            response_data["summary"] = summary
        
        print(f"DEBUG: Successfully generated insights for client {client_id} by admin {admin_user_id}")
        
        # Return successful response
        return jsonify(response_data)
        
    except Exception as e:
        # Handle errors gracefully
        print(f"Error generating insights for client {client_id}: {e}")
        
        # Return a user-friendly error message
        return jsonify({
            "error": "Unable to generate insights for this client at this time",
            "message": "Please try again later or check if the client has financial data",
            "client_id": client_id
        }), 500

@app.route('/api/export-data', methods=['GET'])
@jwt_required()
def export_financial_data():
    """
    Export all financial metrics and chart data for the authenticated user as an Excel file.
    """
    try:
        # Get user identity for data isolation
        client_id = get_jwt_identity()

        # Aggregate all financial data
        all_metrics = get_all_user_metrics_for_export(client_id)
        chart_data = get_chart_data_for_export(client_id)

        # Create Excel file with multiple sheets
        from openpyxl import Workbook
        wb = Workbook()

        # Sheet 1: Financial Metrics
        metrics_sheet = wb.active
        metrics_sheet.title = "Financial Metrics"
        create_metrics_sheet(metrics_sheet, all_metrics)

        # Sheet 2: Income Chart
        income_sheet = wb.create_sheet("Income Chart")
        create_chart_sheet(income_sheet, chart_data['income'], "Income Breakdown")

        # Sheet 3: Expense Chart
        expense_sheet = wb.create_sheet("Expense Chart")
        create_chart_sheet(expense_sheet, chart_data['expense'], "Expense Distribution")

        # Prepare response with Excel file
        from io import BytesIO
        from datetime import datetime
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        from flask import send_file
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f"wisdom_index_financial_data_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        current_app.logger.error(f"Export error for client {client_id}: {e}")
        return jsonify({"error": "Failed to export data"}), 500

# Target Management API Endpoints

@app.route('/api/targets', methods=['GET'])
@jwt_required()
def get_all_targets():
    """
    Get all target values for the authenticated user.
    
    Returns:
        JSON response with all targets or error message
    """
    try:
        targets = get_all_targets_for_user()
        user_id = get_jwt_identity()
        
        return jsonify({
            "targets": targets,
            "user_id": user_id
        })
    except Exception as e:
        print(f"Error in /api/targets: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/targets', methods=['POST'])
@jwt_required()
def update_all_targets():
    """
    Update multiple target values for the authenticated user.
    
    Request Body:
        {
            "targets": {
                "net-worth": 1000000,
                "portfolio-value": 750000,
                ...
            }
        }
        
    Returns:
        JSON response with success status or error message
    """
    try:
        data = request.get_json()
        targets_dict = data.get('targets', {})
        
        # Validate target values
        for metric_name, target_value in targets_dict.items():
            if target_value is not None:
                try:
                    target_value = float(target_value)
                    if target_value < 0:
                        return jsonify({"error": f"Target value for {metric_name} cannot be negative"}), 400
                except (ValueError, TypeError):
                    return jsonify({"error": f"Invalid target value for {metric_name}"}), 400
        
        success = update_multiple_targets_for_user(targets_dict)
        
        if success:
            user_id = get_jwt_identity()
            return jsonify({
                "message": "Targets updated successfully",
                "user_id": user_id
            })
        else:
            return jsonify({"error": "Failed to update targets"}), 500
            
    except Exception as e:
        print(f"Error in /api/targets POST: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/targets/<metric_name>', methods=['PUT'])
@jwt_required()
def update_single_target(metric_name):
    """
    Update a single target value for the authenticated user.
    
    Args:
        metric_name (str): The name of the metric
        
    Request Body:
        {
            "target_value": 1000000
        }
        
    Returns:
        JSON response with success status or error message
    """
    try:
        # Validate metric name
        valid_metrics = [
            'net-worth', 'portfolio-value', 'real-estate-value', 'debt', 'equity', 'fixed-income', 'cash',
            'earned-income', 'social-security-income', 'pension-income', 'real-estate-income', 'business-income', 'total-income',
            'current-year-giving', 'current-year-savings', 'current-year-debt', 'current-year-taxes', 'current-year-living-expenses', 'total-expenses', 'margin',
            'life-insurance', 'disability', 'ltc', 'umbrella', 'business-insurance', 'flood-insurance', 'at-risk',
            'retirement-ratio', 'survivor-ratio', 'education-ratio', 'new-cars-ratio', 'ltc-ratio', 'ltd-ratio',
            'savings-ratio', 'giving-ratio', 'reserves-ratio', 'debt-ratio', 'diversification-ratio'
        ]
        
        if metric_name not in valid_metrics:
            return jsonify({"error": f"Invalid metric name: {metric_name}"}), 400
        
        data = request.get_json()
        target_value = data.get('target_value')
        
        if target_value is not None:
            try:
                target_value = float(target_value)
                if target_value < 0:
                    return jsonify({"error": "Target value cannot be negative"}), 400
            except (ValueError, TypeError):
                return jsonify({"error": "Invalid target value"}), 400
        
        success = update_metric_target_for_user(metric_name, target_value)
        
        if success:
            user_id = get_jwt_identity()
            return jsonify({
                "message": f"Target for {metric_name} updated successfully",
                "user_id": user_id
            })
        else:
            return jsonify({"error": "Failed to update target"}), 500
            
    except Exception as e:
        print(f"Error in /api/targets/{metric_name}: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/targets/<metric_name>', methods=['DELETE'])
@jwt_required()
def delete_single_target(metric_name):
    """
    Delete the most recent target value for a specific metric for the authenticated user.
    
    Args:
        metric_name (str): The name of the metric (received with underscores from frontend)
        
    Returns:
        JSON response with success status or error message
    """
    try:
        # Convert underscore format to hyphen format for database (frontend sends underscores)
        # E.g., 'net_worth' -> 'net-worth'
        db_metric_name = metric_name.replace('_', '-')
        
        # Validate metric name
        valid_metrics = [
            'net-worth', 'portfolio-value', 'real-estate-value', 'debt', 'equity', 'fixed-income', 'cash',
            'earned-income', 'social-security-income', 'pension-income', 'real-estate-income', 'business-income', 'total-income',
            'current-year-giving', 'current-year-savings', 'current-year-debt', 'current-year-taxes', 'current-year-living-expenses', 'total-expenses', 'margin',
            'life-insurance', 'disability', 'ltc', 'umbrella', 'business-insurance', 'flood-insurance', 'at-risk',
            'retirement-ratio', 'survivor-ratio', 'education-ratio', 'new-cars-ratio', 'ltc-ratio', 'ltd-ratio',
            'savings-ratio', 'giving-ratio', 'reserves-ratio', 'debt-ratio', 'diversification-ratio'
        ]
        
        if db_metric_name not in valid_metrics:
            return jsonify({"error": f"Invalid metric name: {metric_name}"}), 400
        
        success = delete_metric_target_for_user(db_metric_name)
        
        if success:
            user_id = get_jwt_identity()
            return jsonify({
                "message": f"Target for {db_metric_name} deleted successfully",
                "user_id": user_id
            })
        else:
            return jsonify({"error": "Target not found or failed to delete"}), 404
            
    except Exception as e:
        print(f"Error in /api/targets/{metric_name} DELETE: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/targets', methods=['DELETE'])
@jwt_required()
def delete_all_targets():
    """
    Delete all target values for the authenticated user.
    
    Returns:
        JSON response with success status or error message
    """
    try:
        success = delete_all_targets_for_user()
        
        if success:
            user_id = get_jwt_identity()
            return jsonify({
                "message": "All targets deleted successfully",
                "user_id": user_id
            })
        else:
            return jsonify({"error": "Failed to delete targets"}), 500
            
    except Exception as e:
        print(f"Error in /api/targets DELETE: {e}")
        return jsonify({"error": "Internal server error"}), 500

# Metric Detail Modal API Endpoints

@app.route('/api/metrics/<metric_name>/details', methods=['GET'])
@jwt_required()
def get_metric_details_endpoint(metric_name):
    """Get metric details including formula, description, and tables used."""
    try:
        # Get metric details
        details = get_metric_details(metric_name)
        if not details:
            return jsonify({"error": "Metric not found"}), 404
        
        # Get current metric value
        metric_function_map = {
            'net-worth': calculate_net_worth_for_user,
            'portfolio-value': calculate_portfolio_value_for_user,
            'real-estate-value': calculate_real_estate_value_for_user,
            'debt': calculate_debt_for_user,
            'equity': calculate_equity_for_user,
            'fixed-income': calculate_fixed_income_for_user,
            'cash': calculate_cash_for_user,
            'earned-income': calculate_earned_income_for_user,
            'social-security-income': calculate_social_security_income_for_user,
            'pension-income': calculate_pension_income_for_user,
            'real-estate-income': calculate_real_estate_income_for_user,
            'business-income': calculate_business_income_for_user,
            'total-income': calculate_total_income_for_user,
            'current-year-giving': calculate_current_year_giving_for_user,
            'current-year-savings': calculate_current_year_savings_for_user,
            'current-year-debt': calculate_current_year_debt_for_user,
            'current-year-taxes': calculate_current_year_taxes_for_user,
            'current-year-living-expenses': calculate_current_year_living_expenses_for_user,
            'total-expenses': calculate_total_expenses_for_user,
            'margin': calculate_margin_for_user,
            'life-insurance': calculate_life_insurance_for_user,
            'disability': calculate_disability_for_user,
            'ltc': calculate_ltc_for_user,
            'umbrella': calculate_umbrella_for_user,
            'business-insurance': calculate_business_insurance_for_user,
            'flood-insurance': calculate_flood_insurance_for_user,
            'at-risk': calculate_at_risk_for_user,
            'retirement-ratio': calculate_retirement_ratio_for_user,
            'survivor-ratio': calculate_survivor_ratio_for_user,
            'education-ratio': calculate_education_ratio_for_user,
            'new-cars-ratio': calculate_new_cars_ratio_for_user,
            'ltc-ratio': calculate_ltc_ratio_for_user,
            'ltd-ratio': calculate_ltd_ratio_for_user,
            'savings-ratio': calculate_savings_ratio_for_user,
            'giving-ratio': calculate_giving_ratio_for_user,
            'reserves-ratio': calculate_reserves_ratio_for_user,
            'debt-ratio': calculate_debt_ratio_for_user,
            'diversification-ratio': calculate_diversification_ratio_for_user
        }
        
        # Get current value
        current_value = None
        formatted_value = None
        
        if metric_name in metric_function_map:
            try:
                current_value = metric_function_map[metric_name]()
                # Format the value
                if current_value is not None:
                    if details['category'] in ('planning', 'wisdom-index'):
                        # Ratio formatting
                        if current_value >= 1000000:
                            formatted_value = f"{current_value / 1000000:.1f}m"
                        elif current_value >= 1000:
                            formatted_value = f"{current_value / 1000:.1f}k"
                        else:
                            formatted_value = f"{current_value:.2f}"
                    else:
                        # Currency formatting
                        abs_value = abs(current_value)
                        sign = '-' if current_value < 0 else ''
                        if abs_value >= 1000000000:
                            formatted_value = f"${sign}{abs_value / 1000000000:.1f}b"
                        elif abs_value >= 1000000:
                            formatted_value = f"${sign}{abs_value / 1000000:.1f}m"
                        elif abs_value >= 1000:
                            formatted_value = f"${sign}{abs_value / 1000:.1f}k"
                        else:
                            formatted_value = f"${sign}{abs_value:.0f}"
                else:
                    formatted_value = '-'
            except Exception as e:
                print(f"Error calculating {metric_name}: {e}")
                current_value = None
                formatted_value = '-'
        
        return jsonify({
            'metric_name': metric_name,
            'title': details['title'],
            'category': details['category'],
            'value': current_value,
            'formatted_value': formatted_value,
            'formula': details['formula'],
            'description': details['description'],
            'tables': details['tables']  # Send table names as array of strings
        })
        
    except Exception as e:
        print(f"Error getting metric details for {metric_name}: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/data/<table_name>', methods=['GET'])
@jwt_required()
def get_table_data_endpoint(table_name):
    """Get raw table data for authenticated user."""
    try:
        # Get query parameters
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 50))
        metric_name = request.args.get('metric_name')
        
        # Validate pagination parameters
        if page < 1:
            page = 1
        if limit < 1 or limit > 100:
            limit = 50
        
        # Get table data
        table_data = get_table_data_for_user(table_name, page=page, limit=limit, metric_name=metric_name)
        
        if table_data is None:
            return jsonify({"error": "Table not found or access denied"}), 404
        
        return jsonify(table_data)
        
    except ValueError as e:
        return jsonify({"error": "Invalid parameters"}), 400
    except Exception as e:
        print(f"Error getting table data for {table_name}: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/tables/<metric_name>', methods=['GET'])
@jwt_required()
def get_tables_for_metric_endpoint(metric_name):
    """Get list of tables used by a specific metric."""
    try:
        # Get metric details
        details = get_metric_details(metric_name)
        if not details:
            return jsonify({"error": "Metric not found"}), 404
        
        # Get detailed information for each table
        tables_info = []
        for table_name in details['tables']:
            try:
                table_data = get_table_data_for_user(table_name, page=1, limit=1)
                if table_data:
                    tables_info.append({
                        'table_name': table_name,
                        'display_name': table_data['display_name'],
                        'columns': table_data['columns'],
                        'row_count': table_data['pagination']['total']
                    })
            except Exception as e:
                print(f"Error getting info for table {table_name}: {e}")
        
        return jsonify({
            'metric_name': metric_name,
            'tables': tables_info
        })
        
    except Exception as e:
        print(f"Error getting tables for metric {metric_name}: {e}")
        return jsonify({"error": "Internal server error"}), 500

# Admin API Endpoints

@app.route('/api/admin/clients', methods=['GET'])
@jwt_required()
@admin_required
def get_all_clients():
    """
    Get all clients for admin dashboard.
    
    Returns:
        JSON response with all client information
    """
    try:
        from database import get_db_connection, close_db_connection
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        query = """
            SELECT
                c.client_id,
                c.first_name,
                c.last_name,
                c.hh_date_of_birth,
                u.username,
                u.email as user_email,
                u.created_at as user_created_at,
                u.last_login,
                CASE WHEN u.user_id IS NOT NULL THEN true ELSE false END as has_account
            FROM core.clients c
            LEFT JOIN core.users u ON c.client_id = u.client_id
            ORDER BY c.client_id
        """
        
        cursor.execute(query)
        clients = cursor.fetchall()
        cursor.close()
        
        # Format results
        clients_list = []
        for client in clients:
            clients_list.append({
                'client_id': client[0],
                'first_name': client[1],
                'last_name': client[2],
                'date_of_birth': client[3],
                'username': client[4],
                'user_email': client[5],
                'account_created': client[6],
                'last_login': client[7],
                'has_account': client[8]
            })
        
        return jsonify({
            "clients": clients_list,
            "total_clients": len(clients_list)
        })
        
    except Exception as e:
        print(f"Error in /api/admin/clients: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/client/<int:client_id>/metrics', methods=['GET'])
@jwt_required()
@admin_required
def get_client_metrics_admin(client_id):
    """
    Get all metrics for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to view
        
    Returns:
        JSON response with all client metrics
    """
    try:
        from metrics import get_all_metrics_for_client
        
        metrics_data = get_all_metrics_for_client(client_id)
        
        return jsonify({
            "client_id": client_id,
            "metrics": metrics_data
        })
        
    except Exception as e:
        print(f"Error in /api/admin/client/{client_id}/metrics: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/tables/<table_name>', methods=['GET'])
@jwt_required()
@admin_required
def get_table_data_admin(table_name):
    """
    Get data from a specific database table.
    
    Args:
        table_name (str): Name of the table to query
        
    Query Parameters:
        client_id (int, optional): Filter by client_id
        limit (int, optional): Limit results (default: 100)
        offset (int, optional): Offset for pagination (default: 0)
        
    Returns:
        JSON response with table data
    """
    try:
        from database import get_db_connection, close_db_connection
        
        # Whitelist of allowed tables
        allowed_tables = {
            'clients': 'core.clients',
            'users': 'core.users',
            'holdings': 'core.holdings',
            'real_estate_assets': 'core.real_estate_assets',
            'businesses': 'core.businesses',
            'investment_deposit_accounts': 'core.investment_deposit_accounts',
            'personal_property_accounts': 'core.personal_property_accounts',
            'liability_note_accounts': 'core.liability_note_accounts',
            'incomes': 'core.incomes',
            'expenses': 'core.expenses',
            'savings': 'core.savings',
            'life_insurance_annuity_accounts': 'core.life_insurance_annuity_accounts',
            'disability_ltc_insurance_accounts': 'core.disability_ltc_insurance_accounts',
            'property_casualty_insurance_accounts': 'core.property_casualty_insurance_accounts',
            'metric_targets': 'core.metric_targets',
            'charities': 'core.charities',
            'entity_interests': 'core.entity_interests',
            'account_history': 'core.account_history',
            'facts': 'core.facts',
            'flows': 'core.flows',
            'values': 'core.values',
            'vw_expense_summary': 'core.vw_expense_summary'
        }
        
        if table_name not in allowed_tables:
            return jsonify({"error": f"Table '{table_name}' not allowed"}), 400
        
        # Get query parameters
        client_id = request.args.get('client_id', type=int)
        limit = min(request.args.get('limit', 100, type=int), 1000)  # Max 1000 rows
        offset = request.args.get('offset', 0, type=int)
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Build query based on table and filters
        base_query = f"SELECT * FROM {allowed_tables[table_name]}"
        count_query = f"SELECT COUNT(*) FROM {allowed_tables[table_name]}"
        
        params = []
        
        # Add client_id filter if provided and table has client_id column
        if client_id is not None:
            base_query += " WHERE client_id = %s"
            count_query += " WHERE client_id = %s"
            params.append(client_id)
        
        # Add pagination
        base_query += " ORDER BY 1 LIMIT %s OFFSET %s"
        params.extend([limit, offset])
        
        # Execute queries
        cursor.execute(base_query, params)
        data = cursor.fetchall()
        
        # Get total count
        cursor.execute(count_query, params[:1])  # Only use client_id param for count
        total_count = cursor.fetchone()[0]
        
        # Get column names
        cursor.execute(f"""
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_schema = 'core' AND table_name = %s
            ORDER BY ordinal_position
        """, (table_name,))
        
        columns_info = cursor.fetchall()
        columns = [{'name': col[0], 'type': col[1]} for col in columns_info]
        cursor.close()
        
        # Format data
        formatted_data = []
        for row in data:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(columns):
                    col_name = columns[i]['name']
                    col_type = columns[i]['type']
                    
                    # Format based on data type
                    if col_type in ('timestamp', 'timestamptz') and value:
                        row_dict[col_name] = value.isoformat()
                    elif col_type in ('numeric', 'decimal') and value:
                        row_dict[col_name] = float(value)
                    else:
                        row_dict[col_name] = value
            formatted_data.append(row_dict)
        
        return jsonify({
            "table_name": table_name,
            "columns": columns,
            "data": formatted_data,
            "pagination": {
                "limit": limit,
                "offset": offset,
                "total": total_count,
                "has_more": offset + len(data) < total_count
            }
        })
        
    except Exception as e:
        print(f"Error in /api/admin/tables/{table_name}: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/analytics', methods=['GET'])
@jwt_required()
@admin_required
def get_admin_analytics():
    """
    Get aggregate analytics across all clients.
    
    Returns:
        JSON response with aggregate metrics
    """
    try:
        from database import get_db_connection, close_db_connection
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Get aggregate metrics
        analytics = {}
        
        # Total AUM (Assets Under Management)
        cursor.execute("""
            SELECT
                COALESCE(SUM(h.value), 0) as holdings_value,
                COALESCE(SUM(rea.total_value), 0) as real_estate_value,
                COALESCE(SUM(b.amount), 0) as business_value,
                COALESCE(SUM(ip.total_value), 0) as investment_value,
                COALESCE(SUM(pp.total_value), 0) as personal_property_value
            FROM core.clients c
            LEFT JOIN core.holdings h ON c.client_id = h.client_id
            LEFT JOIN core.real_estate_assets rea ON c.client_id = rea.client_id
            LEFT JOIN core.businesses b ON c.client_id = b.client_id
            LEFT JOIN core.investment_deposit_accounts ip ON c.client_id = ip.client_id
            LEFT JOIN core.personal_property_accounts pp ON c.client_id = pp.client_id
        """)
        
        asset_data = cursor.fetchone()
        analytics['total_aum'] = sum(asset_data) if asset_data else 0
        analytics['asset_breakdown'] = {
            'holdings': asset_data[0] if asset_data else 0,
            'real_estate': asset_data[1] if asset_data else 0,
            'businesses': asset_data[2] if asset_data else 0,
            'investments': asset_data[3] if asset_data else 0,
            'personal_property': asset_data[4] if asset_data else 0
        }
        
        # Client statistics
        cursor.execute("SELECT COUNT(*) FROM core.clients")
        analytics['total_clients'] = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM core.users WHERE "isAdmin" = false')
        analytics['active_client_accounts'] = cursor.fetchone()[0]
        
        # Income and expense aggregates
        cursor.execute("""
            SELECT
                COALESCE(SUM(i.current_year_amount), 0) as total_income,
                COALESCE(SUM(e.annual_amount), 0) as total_expenses
            FROM core.clients c
            LEFT JOIN core.incomes i ON c.client_id = i.client_id
            LEFT JOIN core.expenses e ON c.client_id = e.client_id
        """)
        
        income_expense = cursor.fetchone()
        analytics['total_income'] = income_expense[0] if income_expense else 0
        analytics['total_expenses'] = income_expense[1] if income_expense else 0
        analytics['total_margin'] = analytics['total_income'] - analytics['total_expenses']
        
        # Target completion rates
        cursor.execute("""
            SELECT
                COUNT(DISTINCT mt.client_id) as clients_with_targets,
                COUNT(DISTINCT c.client_id) as total_clients,
                COUNT(mt.metric_name) as total_targets_set
            FROM core.clients c
            LEFT JOIN core.metric_targets mt ON c.client_id = mt.client_id
        """)
        
        target_data = cursor.fetchone()
        analytics['target_statistics'] = {
            'clients_with_targets': target_data[0] if target_data else 0,
            'total_clients': target_data[1] if target_data else 0,
            'total_targets_set': target_data[2] if target_data else 0,
            'target_adoption_rate': (target_data[0] / target_data[1] * 100) if target_data and target_data[1] > 0 else 0
        }
        
        cursor.close()
        
        return jsonify(analytics)
        
    except Exception as e:
        print(f"Error in /api/admin/analytics: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/client/<int:client_id>/targets', methods=['GET'])
@jwt_required()
@admin_required
def get_client_targets_admin(client_id):
    """
    Get all targets for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to get targets for
        
    Returns:
        JSON response with client targets
    """
    try:
        from metrics import get_all_targets_for_client
        
        targets = get_all_targets_for_client(client_id)
        
        return jsonify({
            "targets": targets,
            "client_id": client_id
        })
        
    except Exception as e:
        print(f"Error in /api/admin/client/{client_id}/targets GET: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/client/<int:client_id>/targets', methods=['POST'])
@jwt_required()
@admin_required
def update_client_targets_admin(client_id):
    """
    Update targets for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to update targets for
        
    Request Body:
        {
            "targets": {
                "net-worth": 1000000,
                "portfolio-value": 750000,
                ...
            }
        }
        
    Returns:
        JSON response with update status
    """
    try:
        from metrics import update_targets_for_client
        
        data = request.get_json()
        targets_dict = data.get('targets', {})
        
        # Validate target values
        for metric_name, target_value in targets_dict.items():
            if target_value is not None:
                try:
                    target_value = float(target_value)
                    if target_value < 0:
                        return jsonify({"error": f"Target value for {metric_name} cannot be negative"}), 400
                except (ValueError, TypeError):
                    return jsonify({"error": f"Invalid target value for {metric_name}"}), 400
        
        # Update targets for the specified client
        success = update_targets_for_client(client_id, targets_dict)
        
        if success:
            return jsonify({
                "message": f"Targets updated successfully for client {client_id}",
                "client_id": client_id
            })
        else:
            return jsonify({"error": "Failed to update targets"}), 500
            
    except Exception as e:
        print(f"Error in /api/admin/client/{client_id}/targets: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/client/<int:client_id>/targets/<metric_name>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_client_target_admin(client_id, metric_name):
    """
    Delete a specific target for a client (admin access).
    
    Args:
        client_id (int): The client ID
        metric_name (str): The metric name
        
    Returns:
        JSON response with delete status
    """
    try:
        from metrics import delete_target_for_client
        
        # Convert underscore format to hyphen format if needed
        db_metric_name = metric_name.replace('_', '-')
        
        success = delete_target_for_client(client_id, db_metric_name)
        
        if success:
            return jsonify({
                "message": f"Target for {db_metric_name} deleted successfully",
                "client_id": client_id
            })
        else:
            return jsonify({"error": "Target not found or failed to delete"}), 404
            
    except Exception as e:
        print(f"Error in /api/admin/client/{client_id}/targets/{metric_name} DELETE: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/client/<int:client_id>/targets', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_all_client_targets_admin(client_id):
    """
    Delete all targets for a specific client (admin access).
    
    Args:
        client_id (int): The client ID
        
    Returns:
        JSON response with delete status
    """
    try:
        from metrics import delete_all_targets_for_client
        
        success = delete_all_targets_for_client(client_id)
        
        if success:
            return jsonify({
                "message": f"All targets deleted successfully for client {client_id}",
                "client_id": client_id
            })
        else:
            return jsonify({"error": "Failed to delete targets"}), 500
            
    except Exception as e:
        print(f"Error in /api/admin/client/{client_id}/targets DELETE: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/clients-summary', methods=['GET'])
@jwt_required()
@admin_required
def get_clients_summary():
    """
    Get all clients with key metrics summary for admin dashboard.
    Optimized to use a single batch query instead of N+1 queries.
    
    Returns:
        JSON response with client summaries
    """
    try:
        from database import get_db_connection, close_db_connection
        from metrics import get_key_metrics_for_all_clients_batch
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Get basic client info
        cursor.execute("""
            SELECT
                c.client_id,
                c.first_name,
                c.last_name,
                u.email,
                u.username,
                u.last_login,
                CASE WHEN u.user_id IS NOT NULL THEN true ELSE false END as has_account
            FROM core.clients c
            LEFT JOIN core.users u ON c.client_id = u.client_id
            ORDER BY c.client_id
        """)
        
        clients_data = cursor.fetchall()
        cursor.close()
        
        # Get all metrics in a single batch query
        all_metrics = get_key_metrics_for_all_clients_batch()
        
        clients_summary = []
        
        for client_row in clients_data:
            client_id = client_row[0]
            
            # Get metrics from batch result, or use defaults if not found
            metrics = all_metrics.get(client_id, {
                'net_worth': 0,
                'portfolio_value': 0,
                'total_income': 0,
                'total_expenses': 0,
                'margin': 0,
                'life_insurance': 0,
                'retirement_ratio': None
            })
            
            client_summary = {
                'client_id': client_id,
                'first_name': client_row[1],
                'last_name': client_row[2],
                'email': client_row[3],
                'username': client_row[4],
                'last_login': client_row[5],
                'has_account': client_row[6],
                'metrics': metrics
            }
            
            clients_summary.append(client_summary)
        
        return jsonify({
            "clients": clients_summary,
            "total_clients": len(clients_summary)
        })
        
    except Exception as e:
        print(f"Error in /api/admin/clients-summary: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

# Security Settings API Endpoints

@app.route('/api/admin/security-code', methods=['GET'])
@jwt_required()
@superuser_required
def get_security_code_status():
    """
    Get the current security code status and last updated information.
    
    Returns:
        JSON response with security code status
    """
    try:
        from database import get_db_connection, close_db_connection
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Get security code info
        cursor.execute("""
            SELECT
                sc.config_value,
                sc.updated_at,
                sc_log.config_value as updated_by_username
            FROM core.system_config sc
            LEFT JOIN core.system_config sc_log ON sc_log.config_key = 'admin_security_code_updated_by'
            WHERE sc.config_key = 'admin_security_code'
        """)
        
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            security_code = result[0]
            updated_at = result[1]
            updated_by_username = result[2]
            
            # Get user info of who last updated it
            updated_by_user = None
            if updated_by_username:
                try:
                    cursor = connection.cursor()
                    cursor.execute("""
                        SELECT user_id, email FROM core.users
                        WHERE username = %s AND "isAdmin" = true
                    """, (updated_by_username,))
                    user_result = cursor.fetchone()
                    if user_result:
                        updated_by_user = {
                            'user_id': user_result[0],
                            'username': updated_by_username,
                            'email': user_result[1]
                        }
                    cursor.close()
                except Exception as e:
                    print(f"Error getting user info for security code update: {e}")
            
            return jsonify({
                "security_code_exists": True,
                "security_code": security_code,  # Include the actual security code
                "security_code_length": len(security_code) if security_code else 0,
                "last_updated": updated_at.isoformat() if updated_at else None,
                "updated_by": updated_by_user
            })
        else:
            return jsonify({
                "security_code_exists": False,
                "security_code": None,  # Include null for consistency
                "security_code_length": 0,
                "last_updated": None,
                "updated_by": None
            })
        
    except Exception as e:
        print(f"Error in /api/admin/security-code: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/security-code', methods=['PUT'])
@jwt_required()
@superuser_required
def update_security_code():
    """
    Update the admin security code.
    
    Request Body:
        {
            "new_security_code": "NewSecureCode123!",
            "confirm_security_code": "NewSecureCode123!"
        }
        
    Returns:
        JSON response with update status
    """
    try:
        data = request.get_json()
        new_security_code = data.get('new_security_code')
        confirm_security_code = data.get('confirm_security_code')
        
        # Validate input
        if not new_security_code or not confirm_security_code:
            return jsonify({"error": "Both new security code and confirmation are required"}), 400
        
        if new_security_code != confirm_security_code:
            return jsonify({"error": "Security codes do not match"}), 400
        
        # Get current admin user ID
        user_id = get_jwt_identity()
        
        # Update security code
        success, message = update_admin_security_code(new_security_code, int(user_id))
        
        if success:
            return jsonify({
                "message": message,
                "user_id": user_id
            })
        else:
            return jsonify({"error": message}), 400
            
    except Exception as e:
        print(f"Error in /api/admin/security-code PUT: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/security-code/validate', methods=['POST'])
@jwt_required()
@superuser_required
def validate_security_code():
    """
    Validate a security code without changing it.
    This can be used to test a new security code before applying it.
    
    Request Body:
        {
            "security_code": "TestCode123!"
        }
        
    Returns:
        JSON response with validation result
    """
    try:
        data = request.get_json()
        security_code = data.get('security_code')
        
        # Validate security code
        if not security_code:
            return jsonify({
                "valid": False,
                "errors": ["Security code is required"]
            }), 400
        
        errors = []
        
        # Check minimum length
        if len(security_code) < 8:
            errors.append("Security code must be at least 8 characters long")
        
        # Check for complexity (optional but recommended)
        has_upper = any(c.isupper() for c in security_code)
        has_lower = any(c.islower() for c in security_code)
        has_digit = any(c.isdigit() for c in security_code)
        has_special = any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in security_code)
        
        complexity_score = sum([has_upper, has_lower, has_digit, has_special])
        
        if complexity_score < 3:
            errors.append("Security code should contain at least 3 of: uppercase letters, lowercase letters, numbers, special characters")
        
        # Check if it's the same as current code
        current_code = get_admin_security_code()
        if security_code == current_code:
            errors.append("New security code must be different from the current one")
        
        return jsonify({
            "valid": len(errors) == 0,
            "errors": errors,
            "complexity_score": complexity_score,
            "complexity_requirements": {
                "has_uppercase": has_upper,
                "has_lowercase": has_lower,
                "has_digit": has_digit,
                "has_special": has_special
            }
        })
        
    except Exception as e:
        print(f"Error in /api/admin/security-code/validate: {e}")
        return jsonify({"error": "Internal server error"}), 500

# User Permissions Management API Endpoints

@app.route('/api/admin/users', methods=['GET'])
@jwt_required()
@superuser_required
def get_all_admin_users():
    """
    Get all admin users for user permissions management.
    
    Returns:
        JSON response with all admin users and their roles
    """
    try:
        from database import get_db_connection, close_db_connection
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        query = """
            SELECT
                user_id,
                client_id,
                username,
                email,
                "isAdmin",
                "isSuperuser",
                created_at,
                last_login
            FROM core.users
            WHERE "isAdmin" = true
            ORDER BY created_at DESC
        """
        
        cursor.execute(query)
        admin_users = cursor.fetchall()
        cursor.close()
        
        # Format results
        users_list = []
        for user in admin_users:
            users_list.append({
                'user_id': user[0],
                'client_id': user[1],
                'username': user[2],
                'email': user[3],
                'isAdmin': user[4],
                'isSuperuser': user[5],
                'created_at': user[6].isoformat() if user[6] else None,
                'last_login': user[7].isoformat() if user[7] else None
            })
        
        return jsonify({
            "admin_users": users_list,
            "total_admin_users": len(users_list)
        })
        
    except Exception as e:
        print(f"Error in /api/admin/users: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/users/<int:user_id>/role', methods=['PUT'])
@jwt_required()
@superuser_required
def update_admin_user_role(user_id):
    """
    Update admin user role (superuser status).
    
    Args:
        user_id (int): The user ID to update
        
    Request Body:
        {
            "isSuperuser": true,
            "reason": "Promoting to superuser for system management"
        }
        
    Returns:
        JSON response with update status
    """
    try:
        from database import get_db_connection, close_db_connection
        
        data = request.get_json()
        is_superuser = data.get('isSuperuser')
        reason = data.get('reason', '')
        
        if is_superuser is None:
            return jsonify({"error": "isSuperuser field is required"}), 400
        
        if user_id == get_jwt_identity():
            return jsonify({"error": "Cannot modify your own role"}), 400
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Check if user exists and is admin
        cursor.execute("""
            SELECT "isAdmin", "isSuperuser" FROM core.users
            WHERE user_id = %s
        """, (user_id,))
        
        user_data = cursor.fetchone()
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        if not user_data[0]:  # isAdmin
            return jsonify({"error": "User is not an admin"}), 400
        
        # Update the user's superuser status
        cursor.execute("""
            UPDATE core.users
            SET "isSuperuser" = %s
            WHERE user_id = %s
        """, (is_superuser, user_id))
        
        connection.commit()
        cursor.close()
        
        # Log the role change for audit purposes
        try:
            cursor = connection.cursor()
            
            # Get usernames for both the changed user and the user making the change
            cursor.execute("""
                SELECT username FROM core.users WHERE user_id = %s
            """, (user_id,))
            target_result = cursor.fetchone()
            target_username = target_result[0] if target_result else "Unknown"
            
            cursor.execute("""
                SELECT username FROM core.users WHERE user_id = %s
            """, (get_jwt_identity(),))
            changer_result = cursor.fetchone()
            changer_username = changer_result[0] if changer_result else "Unknown"
            
            log_query = """
                INSERT INTO core.system_config (config_key, config_value, description)
                VALUES (%s, %s, %s)
                ON CONFLICT (config_key)
                DO UPDATE SET
                    config_value = EXCLUDED.config_value,
                    updated_at = CURRENT_TIMESTAMP
            """
            
            log_key = f"user_role_change_{user_id}_{int(datetime.now().timestamp())}"
            log_value = json.dumps({
                "changed_by": changer_username,
                "user_id": user_id,
                "username": target_username,
                "old_is_superuser": user_data[1],
                "new_is_superuser": is_superuser,
                "reason": reason,
                "timestamp": datetime.now().isoformat()
            })
            
            cursor.execute(log_query, (log_key, log_value, "User role change audit log"))
            connection.commit()
            cursor.close()
        except Exception as log_error:
            print(f"Error logging role change: {log_error}")
        
        action = "promoted to" if is_superuser else "demoted from"
        return jsonify({
            "message": f"User successfully {action} superuser",
            "user_id": user_id,
            "isSuperuser": is_superuser
        })
        
    except Exception as e:
        print(f"Error in /api/admin/users/{user_id}/role: {e}")
        if connection:
            connection.rollback()
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
@superuser_required
def delete_admin_user(user_id):
    """
    Delete an admin user record from the database.
    
    Args:
        user_id (int): The user ID to delete
        
    Request Body:
        {
            "reason": "Account no longer requires admin access"
        }
        
    Returns:
        JSON response with deletion status
    """
    from database import get_db_connection, close_db_connection
    connection = None
    cursor = None
    
    try:
        # Try to get JSON data, but handle empty body gracefully
        try:
            data = request.get_json() or {}
        except Exception:
            data = {}
        
        reason = data.get('reason', 'Admin account deleted')
        
        if user_id == get_jwt_identity():
            return jsonify({"error": "Cannot delete your own account"}), 400
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Check if user exists and is admin
        cursor.execute("""
            SELECT "isAdmin", "isSuperuser", email, username FROM core.users
            WHERE user_id = %s
        """, (user_id,))
        
        user_data = cursor.fetchone()
        if not user_data:
            return jsonify({"error": "User not found"}), 404
        
        if not user_data[0]:  # isAdmin
            return jsonify({"error": "User is not an admin"}), 404
        
        # Check if this is the last superuser
        if user_data[1]:  # isSuperuser
            cursor.execute("""
                SELECT COUNT(*) FROM core.users WHERE "isSuperuser" = true
            """)
            superuser_count = cursor.fetchone()[0]
            
            if superuser_count <= 1:
                return jsonify({"error": "Cannot delete the last superuser admin"}), 400
        
        # Store user info for audit logging before deletion
        user_email = user_data[2]
        user_username = user_data[3]
        
        # Delete the admin user record from the database
        cursor.execute("""
            DELETE FROM core.users
            WHERE user_id = %s
        """, (user_id,))
        
        # Check if the deletion was successful
        if cursor.rowcount == 0:
            connection.rollback()
            return jsonify({"error": "Failed to delete user"}), 500
        
        connection.commit()
        
        # Log the account deletion for audit purposes
        try:
            # Create a new cursor for logging
            log_cursor = connection.cursor()
            
            # Get username of the user making the change
            log_cursor.execute("""
                SELECT username FROM core.users WHERE user_id = %s
            """, (get_jwt_identity(),))
            changer_result = log_cursor.fetchone()
            changer_username = changer_result[0] if changer_result else "Unknown"
            
            log_key = f"user_deleted_{user_id}_{int(datetime.now().timestamp())}"
            log_value = json.dumps({
                "changed_by": changer_username,
                "deleted_user_id": user_id,
                "deleted_email": user_email,
                "deleted_username": user_username,
                "old_is_admin": user_data[0],
                "old_is_superuser": user_data[1],
                "reason": reason,
                "timestamp": datetime.now().isoformat()
            })
            
            log_cursor.execute("""
                INSERT INTO core.system_config (config_key, config_value, description)
                VALUES (%s, %s, %s)
            """, (log_key, log_value, "User deletion audit log"))
            connection.commit()
            log_cursor.close()
        except Exception as log_error:
            print(f"Error logging user deletion: {log_error}")
        
        return jsonify({
            "message": "Admin user successfully deleted from the system",
            "user_id": user_id,
            "deleted_email": user_email,
            "deleted_username": user_username
        })
            
    except Exception as e:
        print(f"Error in /api/admin/users/{user_id}: {e}")
        if connection:
            connection.rollback()
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/activity-logs', methods=['GET'])
@jwt_required()
@superuser_required
def get_user_activity_logs():
    """
    Get user activity logs for monitoring.
    
    Query Parameters:
        limit (int, optional): Number of logs to return (default: 50)
        offset (int, optional): Pagination offset (default: 0)
        
    Returns:
        JSON response with activity logs
    """
    try:
        from database import get_db_connection, close_db_connection
        
        # Get user identity for logging
        user_id = get_jwt_identity()
        
        # Get query parameters
        limit = min(request.args.get('limit', 50, type=int), 200)  # Max 200 logs
        offset = request.args.get('offset', 0, type=int)
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Get activity logs from system_config table
        try:
            # Build query with proper escaping (safe since we control the values)
            limit = min(limit, 200)  # Ensure limit doesn't exceed 200
            query = f"""
                SELECT config_key, config_value, description, updated_at
                FROM core.system_config
                WHERE config_key LIKE 'user_role_change_%'
                   OR config_key LIKE 'user_deleted_%'
                   OR config_key LIKE 'policy_change_%'
                   OR config_key LIKE 'ai_config_change_%'
                   OR config_key LIKE '%admin_%'
                   OR config_key = 'admin_security_code'
                   OR config_key = 'admin_security_code_updated_by'
                ORDER BY updated_at DESC NULLS LAST
                LIMIT {limit}
            """
            if offset > 0:
                query += f" OFFSET {offset}"
            
            cursor.execute(query)
            
            logs = cursor.fetchall()
            
            # Get total count
            count_query = """
                SELECT COUNT(*) FROM core.system_config
                WHERE config_key LIKE 'user_role_change_%'
                   OR config_key LIKE 'user_deleted_%'
                   OR config_key LIKE 'policy_change_%'
                   OR config_key LIKE 'ai_config_change_%'
                   OR config_key LIKE '%admin_%'
                   OR config_key = 'admin_security_code'
                   OR config_key = 'admin_security_code_updated_by'
            """
            cursor.execute(count_query)
            count_result = cursor.fetchone()
            total_count = count_result[0] if count_result and len(count_result) > 0 else 0
            
        except Exception as query_error:
            print(f"Database query error in activity logs: {query_error}")
            logs = []
            total_count = 0
        
        cursor.close()
        
        # Format logs with safer processing
        formatted_logs = []
        for i, log in enumerate(logs):
            try:
                # Ensure log is a tuple with at least 4 elements
                if not isinstance(log, (list, tuple)) or len(log) < 4:
                    print(f"Skipping malformed log {i}: {log}")
                    continue
                
                # Extract values safely
                config_key = log[0] if len(log) > 0 else 'unknown'
                config_value = log[1] if len(log) > 1 else None
                description = log[2] if len(log) > 2 else 'No description'
                updated_at = log[3] if len(log) > 3 else None
                
                # Parse JSON data safely
                log_data = {}
                if config_value:
                    try:
                        log_data = json.loads(config_value)
                    except (json.JSONDecodeError, TypeError):
                        log_data = {'raw_value': str(config_value)}
                
                # Special handling for security code related logs
                if config_key == 'admin_security_code':
                    # This is the actual security code value
                    log_data = {
                        'type': 'security_code_update',
                        'security_code': config_value,
                        'message': f'Admin security code updated to: {config_value}'
                    }
                    description = 'Admin security code updated'
                elif config_key == 'admin_security_code_updated_by':
                    # This is the username who updated the security code
                    log_data = {
                        'type': 'security_code_updated_by',
                        'username': config_value,
                        'message': f'Security code updated by: {config_value}'
                    }
                    description = 'Username who last updated security code'
                
                # Special handling for AI configuration change logs
                elif config_key.startswith('ai_config_change_'):
                    # This is an AI configuration change log
                    username = log_data.get('changed_by_username', f"user {log_data.get('changed_by_user_id', 'Unknown')}")
                    log_data = {
                        'type': 'ai_config_change',
                        'changed_by_user_id': log_data.get('changed_by_user_id'),
                        'changed_by_username': log_data.get('changed_by_username'),
                        'changed_at': log_data.get('changed_at'),
                        'changes': log_data.get('changes', {}),
                        'message': f'AI configuration changed by {username}'
                    }
                    description = 'AI configuration change audit log'
                
                # Create log entry
                log_entry = {
                    'log_key': str(config_key),
                    'description': str(description),
                    'timestamp': updated_at.isoformat() if updated_at else None,
                    'data': log_data
                }
                
                formatted_logs.append(log_entry)
                print(f"Successfully processed log {i}: {log_entry['log_key']}")
                
            except Exception as e:
                print(f"Error formatting log entry {i}: {e}")
                # Create a minimal safe entry
                try:
                    formatted_logs.append({
                        'log_key': f'error_log_{i}',
                        'description': f'Error processing log: {str(e)}',
                        'timestamp': None,
                        'data': {'error': str(e)}
                    })
                except Exception as fallback_error:
                    print(f"Fallback failed for log {i}: {fallback_error}")
                    continue
        
        return jsonify({
            "logs": formatted_logs,
            "pagination": {
                "limit": limit,
                "offset": offset,
                "total": total_count,
                "has_more": offset + len(logs) < total_count
            }
        })
        
    except Exception as e:
        print(f"Error in /api/admin/activity-logs: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/permissions/policies', methods=['GET'])
@jwt_required()
@superuser_required
def get_access_control_policies():
    """
    Get current access control policies.
    
    Returns:
        JSON response with access control policies
    """
    try:
        from database import get_db_connection, close_db_connection
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Get system policies from system_config
        cursor.execute("""
            SELECT config_key, config_value, description, updated_at
            FROM core.system_config
            WHERE config_key LIKE '%policy%' OR config_key LIKE '%permission%'
            ORDER BY config_key
        """)
        
        policies = cursor.fetchall()
        cursor.close()
        
        # Format policies
        formatted_policies = {}
        for policy in policies:
            policy_name = policy[0].replace('_policy', '').replace('access_', '')
            try:
                policy_value = json.loads(policy[1]) if policy[1] and policy[1].startswith('{') else policy[1]
            except json.JSONDecodeError:
                policy_value = policy[1]
            
            formatted_policies[policy_name] = {
                'value': policy_value,
                'description': policy[2],
                'last_updated': policy[3].isoformat() if policy[3] else None
            }
        
        return jsonify({
            "policies": formatted_policies
        })
        
    except Exception as e:
        print(f"Error in /api/admin/permissions/policies: {e}")
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

@app.route('/api/admin/permissions/policies/<policy_name>', methods=['PUT'])
@jwt_required()
@superuser_required
def update_access_control_policy(policy_name):
    """
    Update an access control policy.
    
    Args:
        policy_name (str): The name of the policy to update
        
    Request Body:
        {
            "policy_value": {"enabled": true, "restrictions": [...]},
            "description": "Updated policy description"
        }
        
    Returns:
        JSON response with update status
    """
    try:
        from database import get_db_connection, close_db_connection
        
        data = request.get_json()
        policy_value = data.get('policy_value')
        description = data.get('description', f'Access control policy for {policy_name}')
        
        if policy_value is None:
            return jsonify({"error": "policy_value is required"}), 400
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Convert policy value to JSON string if it's a dict
        if isinstance(policy_value, dict):
            policy_value_str = json.dumps(policy_value)
        else:
            policy_value_str = str(policy_value)
        
        # Update or insert policy
        cursor.execute("""
            INSERT INTO core.system_config (config_key, config_value, description)
            VALUES (%s, %s, %s)
            ON CONFLICT (config_key)
            DO UPDATE SET
                config_value = EXCLUDED.config_value,
                description = EXCLUDED.description,
                updated_at = CURRENT_TIMESTAMP
        """, (f'access_{policy_name}_policy', policy_value_str, description))
        
        connection.commit()
        cursor.close()
        
        # Log the policy change
        try:
            cursor = connection.cursor()
            log_key = f"policy_change_{policy_name}_{int(datetime.now().timestamp())}"
            log_value = json.dumps({
                "changed_by": get_jwt_identity(),
                "policy_name": policy_name,
                "old_value": "Previous policy value",  # In a real implementation, you'd fetch the old value
                "new_value": policy_value,
                "timestamp": datetime.now().isoformat()
            })
            
            cursor.execute("""
                INSERT INTO core.system_config (config_key, config_value, description)
                VALUES (%s, %s, %s)
            """, (log_key, log_value, "Policy change audit log"))
            connection.commit()
            cursor.close()
        except Exception as log_error:
            print(f"Error logging policy change: {log_error}")
        
        return jsonify({
            "message": f"Policy '{policy_name}' updated successfully",
            "policy_name": policy_name,
            "policy_value": policy_value
        })
        
    except Exception as e:
        print(f"Error in /api/admin/permissions/policies/{policy_name}: {e}")
        if connection:
            connection.rollback()
        return jsonify({"error": "Internal server error"}), 500
    finally:
        if connection:
            close_db_connection(connection)

# Account History API Endpoints

@app.route('/api/accounts', methods=['GET'])
@jwt_required()
def get_user_accounts_endpoint():
    """
    Get all accounts for the authenticated user.
    
    Returns:
        JSON response with account information
    """
    try:
        accounts = get_user_accounts()
        user_id = get_jwt_identity()
        
        return jsonify({
            "accounts": accounts,
            "total_accounts": len(accounts),
            "user_id": user_id
        })
    except Exception as e:
        print(f"Error in /api/accounts: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/accounts/<account_id>/history', methods=['GET'])
@jwt_required()
def get_account_history_endpoint(account_id):
    """
    Get historical data for a specific account.
    
    Args:
        account_id (str): The account ID to get history for
        
    Query Parameters:
        start_date (str, optional): Filter from date
        end_date (str, optional): Filter to date
        limit (int, optional): Number of records to return (default: 100, max: 1000)
        offset (int, optional): Pagination offset (default: 0)
        
    Returns:
        JSON response with account history data
    """
    try:
        # Get query parameters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = min(int(request.args.get('limit', 100)), 1000)  # Max 1000 records
        offset = int(request.args.get('offset', 0))
        
        # Get account history
        history_data = get_account_history_for_user(account_id, start_date, end_date, limit, offset)
        user_id = get_jwt_identity()
        
        return jsonify({
            "account_id": account_id,
            "history": history_data.get('history', []),
            "pagination": history_data.get('pagination', {}),
            "user_id": user_id
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/accounts/{account_id}/history: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/accounts/history', methods=['POST'])
@jwt_required()
def get_multiple_account_history_endpoint():
    """
    Get historical data for multiple accounts.
    
    Request Body:
        {
            "account_ids": ["string"],
            "start_date": "string",
            "end_date": "string"
        }
        
    Returns:
        JSON response with multiple account histories
    """
    try:
        data = request.get_json()
        account_ids = data.get('account_ids', [])
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if not account_ids:
            return jsonify({"error": "account_ids is required"}), 400
        
        # Get multiple account histories
        histories = get_multiple_account_history(account_ids, start_date, end_date)
        user_id = get_jwt_identity()
        
        return jsonify({
            "histories": histories,
            "user_id": user_id
        })
    except Exception as e:
        print(f"Error in /api/accounts/history: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/accounts/<account_id>/summary', methods=['GET'])
@jwt_required()
def get_account_summary_endpoint(account_id):
    """
    Get summary statistics for a specific account.
    
    Args:
        account_id (str): The account ID to get summary for
        
    Returns:
        JSON response with account summary statistics
    """
    try:
        summary = get_account_summary_for_user(account_id)
        user_id = get_jwt_identity()
        
        return jsonify({
            "account_id": account_id,
            "summary": summary,
            "user_id": user_id
        })
    except Exception as e:
        print(f"Error in /api/accounts/{account_id}/summary: {e}")
        return jsonify({"error": "Internal server error"}), 500

@app.route('/api/admin/ai-config', methods=['GET'])
@jwt_required()
@superuser_required
def get_ai_config():
    """
    Get current AI configuration values for superuser management.
    
    Query Parameters:
        include_api_key (bool, optional): Include the actual API key value in the response (default: false)
    
    Returns:
        JSON response with AI configuration values (base_url, model, optionally api_key)
    """
    try:
        from database import get_db_connection, close_db_connection
        
        # Get query parameter to determine if API key should be included
        include_api_key = request.args.get('include_api_key', 'false').lower() == 'true'
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        if include_api_key:
            # Query to get all AI configuration from system_config table including the API key
            query = """
                SELECT config_key, config_value
                FROM core.system_config
                WHERE config_key IN ('ai_base_url', 'ai_model', 'ai_api_key')
            """
            cursor.execute(query)
        else:
            # Query to get non-sensitive AI configuration from system_config table
            query = """
                SELECT config_key, config_value
                FROM core.system_config
                WHERE config_key IN ('ai_base_url', 'ai_model')
            """
            cursor.execute(query)
        config_results = cursor.fetchall()
        cursor.close()
        
        # Create a dictionary from the results
        ai_config = {}
        for config_key, config_value in config_results:
            if config_key == 'ai_base_url':
                ai_config['ai_base_url'] = config_value
            elif config_key == 'ai_model':
                ai_config['ai_model'] = config_value
            elif config_key == 'ai_api_key' and include_api_key:
                ai_config['ai_api_key'] = config_value
        
        # Check if API key exists in database
        cursor = connection.cursor()
        cursor.execute("""
            SELECT COUNT(*) > 0
            FROM core.system_config
            WHERE config_key = 'ai_api_key' AND config_value IS NOT NULL AND config_value != ''
        """)
        api_key_exists = cursor.fetchone()[0]
        cursor.close()
        
        # Include status indicators for sensitive fields
        ai_config['api_key_set'] = api_key_exists
        if 'ai_base_url' not in ai_config:
            ai_config['ai_base_url'] = ''
        if 'ai_model' not in ai_config:
            ai_config['ai_model'] = ''
        if include_api_key and 'ai_api_key' not in ai_config:
            ai_config['ai_api_key'] = ''
        
        return jsonify(ai_config)
    except Exception as e:
        print(f"Error getting AI configuration: {e}")
        return jsonify({"error": "Failed to retrieve AI configuration"}), 500
    finally:
        if connection:
            close_db_connection(connection)


@app.route('/api/admin/ai-config', methods=['PUT'])
@jwt_required()
@superuser_required
def update_ai_config():
    """
    Update AI configuration values by superuser.
    
    Request Body:
        {
            "ai_base_url": "string",
            "ai_model": "string",
            "ai_api_key": "string" (optional)
        }
    
    Returns:
        JSON response with update status
    """
    try:
        from database import get_db_connection, close_db_connection
        
        data = request.get_json()
        ai_base_url = data.get('ai_base_url', '').strip()
        ai_model = data.get('ai_model', '').strip()
        ai_api_key = data.get('ai_api_key', '').strip()  # Optional, for security updates only
        
        if not ai_base_url or not ai_model:
            return jsonify({"error": "Both base URL and model are required"}), 400
        
        # Validate URL format
        import re
        url_pattern = re.compile(
            r'^https?://'  # http:// or https://
            r'(?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}'  # domain
            r'(?::[0-9]+)?'  # optional port
            r'(?:/.*)?$'  # optional path
        )
        if not url_pattern.match(ai_base_url):
            return jsonify({"error": "Invalid URL format for base URL"}), 400
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Update AI configuration in system_config table
        update_queries = []
        params_list = []
        
        # Update base URL
        update_queries.append("""
            INSERT INTO core.system_config (config_key, config_value, description)
            VALUES ('ai_base_url', %s, 'AI service base URL')
            ON CONFLICT (config_key) DO UPDATE SET
                config_value = EXCLUDED.config_value,
                description = EXCLUDED.description,
                updated_at = CURRENT_TIMESTAMP
        """)
        params_list.append((ai_base_url,))
        
        # Update model
        update_queries.append("""
            INSERT INTO core.system_config (config_key, config_value, description)
            VALUES ('ai_model', %s, 'AI model identifier')
            ON CONFLICT (config_key) DO UPDATE SET
                config_value = EXCLUDED.config_value,
                description = EXCLUDED.description,
                updated_at = CURRENT_TIMESTAMP
        """)
        params_list.append((ai_model,))
        
        # Update API key if provided
        if ai_api_key:
            update_queries.append("""
                INSERT INTO core.system_config (config_key, config_value, description)
                VALUES ('ai_api_key', %s, 'AI service API key')
                ON CONFLICT (config_key) DO UPDATE SET
                    config_value = EXCLUDED.config_value,
                    description = EXCLUDED.description,
                    updated_at = CURRENT_TIMESTAMP
            """)
            params_list.append((ai_api_key,))
        
        # Execute all updates in a transaction
        for i, query in enumerate(update_queries):
            cursor.execute(query, params_list[i])
        
        connection.commit()
        cursor.close()
        
        # Log the change for audit purposes
        try:
            from auth import is_admin_user
            user_id = get_jwt_identity()
            cursor = connection.cursor()
            
            # Get the username of the user making the change
            cursor.execute("""
                SELECT username FROM core.users WHERE user_id = %s
            """, (int(user_id),))
            user_result = cursor.fetchone()
            username = user_result[0] if user_result else "Unknown"
            
            log_key = f"ai_config_change_{int(datetime.now().timestamp())}"
            log_value = json.dumps({
                "changed_by_user_id": user_id,
                "changed_by_username": username,
                "changed_at": datetime.now().isoformat(),
                "changes": {
                    "ai_base_url": ai_base_url,
                    "ai_model": ai_model,
                    "ai_api_key_updated": bool(ai_api_key)
                }
            })
            
            cursor.execute("""
                INSERT INTO core.system_config (config_key, config_value, description)
                VALUES (%s, %s, %s)
            """, (log_key, log_value, "AI configuration change audit log"))
            connection.commit()
            cursor.close()
        except Exception as log_error:
            print(f"Error logging AI config change: {log_error}")
        finally:
            close_db_connection(connection)
        
        return jsonify({
            "message": "AI configuration updated successfully",
            "config": {
                "ai_base_url": ai_base_url,
                "ai_model": ai_model,
                "api_key_updated": bool(ai_api_key)
            }
        })
    except Exception as e:
        print(f"Error updating AI configuration: {e}")
        if 'connection' in locals():
            connection.rollback()
            close_db_connection(connection)
        return jsonify({"error": "Failed to update AI configuration"}), 500


# Admin-specific Account History API Endpoints

@app.route('/api/admin/client/<int:client_id>/accounts', methods=['GET'])
@jwt_required()
@admin_required
def get_admin_client_accounts(client_id):
    """
    Get all accounts for a specific client (admin access).
    
    Args:
        client_id (int): The client ID to get accounts for
        
    Returns:
        JSON response with account information for the specified client
    """
    try:
        accounts = get_user_accounts_for_admin(client_id)

        return jsonify({
            "accounts": accounts,
            "total_accounts": len(accounts),
            "client_id": client_id
        })
    except Exception as e:
        print(f"Error in /api/admin/client/{client_id}/accounts: {e}")
        return jsonify({"error": "Internal server error"}), 500


@app.route('/api/admin/client/<int:client_id>/accounts/<account_id>/history', methods=['GET'])
@jwt_required()
@admin_required
def get_admin_client_account_history(client_id, account_id):
    """
    Get historical data for a specific account of a specific client (admin access).
    
    Args:
        client_id (int): The client ID
        account_id (str): The account ID to get history for
        
    Query Parameters:
        start_date (str, optional): Filter from date
        end_date (str, optional): Filter to date
        limit (int, optional): Number of records to return (default: 100, max: 1000)
        offset (int, optional): Pagination offset (default: 0)
        
    Returns:
        JSON response with account history data
    """
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        limit = min(int(request.args.get('limit', 100)), 1000)  # Max 1000 records
        offset = int(request.args.get('offset', 0))

        history_data = get_account_history_for_admin(
            client_id, account_id, start_date, end_date, limit, offset
        )

        return jsonify({
            "client_id": client_id,
            "account_id": account_id,
            "history": history_data.get('history', []),
            "pagination": history_data.get('pagination', {})
        })
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        print(f"Error in /api/admin/client/{client_id}/accounts/{account_id}/history: {e}")
        return jsonify({"error": "Internal server error"}), 500


@app.route('/health', methods=['GET'])
def health_check():
    """
    Health check endpoint to verify the API is running.
    
    Returns:
        JSON response indicating the service status
    """
    return jsonify({
        "status": "healthy",
        "service": "backend_beta",
        "version": "1.0.0"
    })

@app.errorhandler(404)
def not_found(error):
    """
    Handle 404 errors for undefined routes.
    """
    return jsonify({"error": "Endpoint not found"}), 404

@app.errorhandler(500)
def internal_error(error):
    """
    Handle 500 errors.
    """
    return jsonify({"error": "Internal server error"}), 500

if __name__ == '__main__':
    # Get port from environment variable for Railway, default to 5001 for local development
    port = int(os.environ.get('PORT', 5001))
    
    # Check if running in Railway environment
    if os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('ENVIRONMENT') == 'production':
        # Production environment - use waitress
        from waitress import serve
        serve(app, host='0.0.0.0', port=port, threads=8)
    else:
        # Development environment
        try:
            from waitress import serve
            print(f"Starting development server with waitress on port {port}")
            serve(app, host='0.0.0.0', port=port, threads=4)
        except ImportError:
            # Fallback to Flask development server if waitress is not installed
            print(f"Starting development server with Flask on port {port}")
            app.run(
                host='0.0.0.0',
                port=port,
                debug=True
            )
